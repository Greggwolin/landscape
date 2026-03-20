"""
Parcel Import tools for Landscaper AI.

Provides spreadsheet parsing and bulk parcel creation for land development
projects. Enables Landscaper to read lot rosters from Excel/CSV files and
create phases + parcels with user confirmation.

Tools:
  1. parse_spreadsheet_lots  - Parse a spreadsheet to identify lot roster data
  2. bulk_create_parcels     - Create phases and parcels from structured lot data
  3. get_hierarchy_config    - Check project hierarchy settings to guide import

Architecture note:
  These tools are designed to work conversationally:
  1. User drops spreadsheet on Landscaper
  2. Landscaper calls parse_spreadsheet_lots to identify lot data
  3. Landscaper proposes grouping (phases) based on hierarchy config
  4. User confirms → Landscaper calls bulk_create_parcels
"""

import json
import logging
import os
import re
import tempfile
from typing import Dict, Any, List, Optional

from django.db import connection
from ..tool_executor import register_tool

logger = logging.getLogger(__name__)


def _parse_excel_lot_roster(file_path: str) -> Dict[str, Any]:
    """
    Parse an Excel file to find lot roster data.

    Looks for repeating rows with lot/unit identifiers (addresses, lot numbers)
    plus attributes like SF, acreage, unit size, design type.

    Returns structured lot data or error.
    """
    try:
        import openpyxl
    except ImportError:
        return {'success': False, 'error': 'openpyxl not available for Excel parsing'}

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception as e:
        return {'success': False, 'error': f'Failed to open workbook: {e}'}

    lots = []
    sheet_scanned = None
    header_row_idx = None

    # Scan each sheet for lot roster patterns
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_data = []
        for row in ws.iter_rows(values_only=True, max_row=200, max_col=30):
            rows_data.append([c for c in row])

        if not rows_data:
            continue

        # Strategy: find a section with repeating rows that look like lots
        # Lot indicators: addresses (### Street), "Lot N", "Unit N", "Block N"
        lot_pattern = re.compile(
            r'^\d+\s+\w+\s+(ct|court|st|street|dr|drive|ln|lane|way|rd|road|ave|blvd|pl|circle|cir)'
            r'|^lot\s*\d+'
            r'|^unit\s*\d+'
            r'|^\w+\s+ct\s+unit\s*\d+',
            re.IGNORECASE
        )

        # Find consecutive rows matching lot patterns
        current_group = []
        current_group_start = None
        best_group = []
        best_group_start = None
        header_candidate = None

        for i, row in enumerate(rows_data):
            # Check col B (index 1) — common for proformas to have labels in col B
            label = None
            label_col = None
            for col_idx in [1, 0, 2]:
                if col_idx < len(row) and row[col_idx] is not None:
                    val = str(row[col_idx]).strip()
                    if val and lot_pattern.search(val):
                        label = val
                        label_col = col_idx
                        break

            if label:
                if not current_group:
                    current_group_start = i
                    # Look back 1-3 rows for a header
                    for hdr_offset in range(1, 4):
                        if i - hdr_offset >= 0:
                            hdr_row = rows_data[i - hdr_offset]
                            hdr_vals = [str(c).strip() if c else '' for c in hdr_row]
                            # Header should have multiple non-empty text cells
                            text_cells = [v for v in hdr_vals if v and not v.replace('.', '').replace('-', '').isdigit()]
                            if len(text_cells) >= 2:
                                header_candidate = hdr_row
                                break

                current_group.append((i, label_col, row))
            else:
                if len(current_group) > len(best_group):
                    best_group = current_group
                    best_group_start = current_group_start
                current_group = []
                current_group_start = None

        # Check final group
        if len(current_group) > len(best_group):
            best_group = current_group
            best_group_start = current_group_start

        if len(best_group) >= 2:
            sheet_scanned = sheet_name

            for row_idx, label_col, row in best_group:
                lot_label = str(row[label_col]).strip() if label_col < len(row) and row[label_col] else ''

                # Extract numeric values from adjacent columns
                numerics = []
                for col_idx in range(len(row)):
                    if col_idx == label_col:
                        continue
                    val = row[col_idx]
                    if isinstance(val, (int, float)) and val > 0:
                        numerics.append((col_idx, val))

                # Heuristic: identify lot_sf, unit_sf, acres from value ranges
                lot_sf = None
                unit_sf = None
                acres = None

                for col_idx, val in numerics:
                    if 1000 <= val <= 50000:
                        # Could be lot SF or unit SF
                        if lot_sf is None:
                            lot_sf = val
                        elif unit_sf is None:
                            unit_sf = val
                    elif 0.01 <= val <= 100:
                        acres = val

                # Use header row to disambiguate if available
                if header_candidate:
                    for col_idx, val in numerics:
                        if col_idx < len(header_candidate) and header_candidate[col_idx]:
                            hdr = str(header_candidate[col_idx]).lower()
                            if 'lot' in hdr and ('sf' in hdr or 'sqft' in hdr or 'size' in hdr or 'area' in hdr):
                                lot_sf = val
                            elif 'unit' in hdr and ('sf' in hdr or 'size' in hdr):
                                unit_sf = val
                            elif 'livable' in hdr or 'living' in hdr:
                                unit_sf = val
                            elif 'acre' in hdr:
                                acres = val

                lots.append({
                    'label': lot_label,
                    'lot_sf': lot_sf,
                    'unit_sf': unit_sf,
                    'acres': acres,
                    'row_index': row_idx + 1,  # 1-based
                })

            break  # Found lots in this sheet, stop scanning

    wb.close()

    if not lots:
        return {
            'success': False,
            'error': 'No lot roster found. Looked for repeating rows with addresses or lot/unit identifiers.',
            'sheets_scanned': wb.sheetnames if hasattr(wb, 'sheetnames') else [],
        }

    # Try to detect natural groupings (by address prefix/street)
    groups = {}
    for lot in lots:
        label = lot['label']
        # Group by street name: "806 Lemhi Ct" → "Lemhi Ct", "Waahni Ct Unit 1" → "Waahni Ct"
        street = re.sub(r'^\d+\s+', '', label)  # Remove leading number
        street = re.sub(r'\s+unit\s*\d+.*$', '', street, flags=re.IGNORECASE)  # Remove trailing "Unit N"
        street = street.strip()
        if street not in groups:
            groups[street] = []
        groups[street].append(lot)

    return {
        'success': True,
        'sheet': sheet_scanned,
        'total_lots': len(lots),
        'lots': lots,
        'detected_groups': {k: len(v) for k, v in groups.items()},
        'groups_detail': {k: [l['label'] for l in v] for k, v in groups.items()},
    }


@register_tool('parse_spreadsheet_lots')
def parse_spreadsheet_lots(doc_id: int = None, **kwargs):
    """
    Parse a spreadsheet (Excel/CSV) to identify lot roster data.

    Reads the uploaded document from core_doc, parses it to find repeating
    rows with lot/unit identifiers (addresses, lot numbers) and extracts
    attributes like lot SF, unit SF, and acreage.

    Also detects natural groupings (by street name) that could map to phases.

    Args:
        doc_id: The core_doc ID of the uploaded spreadsheet

    Returns:
        Structured lot data with detected groupings for phase mapping.
    """
    project_context = kwargs.get('project_context', {})
    project_id = project_context.get('project_id')
    tool_input = kwargs.get('tool_input', {})

    doc_id = doc_id or tool_input.get('doc_id')

    if not doc_id:
        return {'success': False, 'error': 'doc_id is required'}
    if not project_id:
        return {'success': False, 'error': 'project_id not available in context'}

    # Get file info from core_doc
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT doc_name, mime_type, storage_uri
                FROM landscape.core_doc
                WHERE doc_id = %s AND project_id = %s AND is_deleted = false
            """, [doc_id, project_id])
            row = cursor.fetchone()

        if not row:
            return {'success': False, 'error': f'Document {doc_id} not found for this project'}

        doc_name, mime_type, storage_uri = row

        # Check if it's a spreadsheet
        spreadsheet_types = [
            'application/vnd.ms-excel',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'application/vnd.ms-excel.sheet.macroEnabled.12',
            'text/csv',
        ]
        is_spreadsheet = (
            mime_type in spreadsheet_types
            or doc_name.lower().endswith(('.xlsx', '.xlsm', '.xls', '.csv'))
        )

        if not is_spreadsheet:
            return {
                'success': False,
                'error': f'Document "{doc_name}" is not a spreadsheet (type: {mime_type}). '
                         f'Please upload an Excel (.xlsx, .xlsm) or CSV file.',
            }

        # Download file from storage URI to temp location
        if storage_uri and storage_uri.startswith('http'):
            import urllib.request
            suffix = os.path.splitext(doc_name)[1] or '.xlsx'
            with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
                tmp_path = tmp.name
            try:
                urllib.request.urlretrieve(storage_uri, tmp_path)
                result = _parse_excel_lot_roster(tmp_path)
            finally:
                if os.path.exists(tmp_path):
                    os.unlink(tmp_path)
        else:
            return {
                'success': False,
                'error': f'Cannot access file storage for doc {doc_id}. Storage URI: {storage_uri}',
            }

        if result.get('success'):
            result['doc_id'] = doc_id
            result['doc_name'] = doc_name

        return result

    except Exception as e:
        logger.error(f"Error parsing spreadsheet lots for doc {doc_id}: {e}")
        return {'success': False, 'error': str(e)}


@register_tool('get_hierarchy_config')
def get_hierarchy_config(**kwargs):
    """
    Get the project's hierarchy configuration (which levels are enabled).

    Returns level1_enabled, level2_enabled, and their labels so Landscaper
    can guide the user on how to structure the parcel import (with or without
    areas/phases).
    """
    project_context = kwargs.get('project_context', {})
    project_id = project_context.get('project_id')

    if not project_id:
        return {'success': False, 'error': 'project_id not available in context'}

    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT
                    tier_1_label, tier_2_label, tier_3_label,
                    COALESCE(level1_enabled, true) as level1_enabled,
                    COALESCE(level2_enabled, true) as level2_enabled,
                    auto_number
                FROM landscape.tbl_project_config
                WHERE project_id = %s
            """, [project_id])
            row = cursor.fetchone()

            # Also get current phase/area counts
            cursor.execute("""
                SELECT COUNT(*) FROM landscape.tbl_area WHERE project_id = %s
            """, [project_id])
            area_count = cursor.fetchone()[0]

            cursor.execute("""
                SELECT COUNT(*) FROM landscape.tbl_phase
                WHERE area_id IN (SELECT area_id FROM landscape.tbl_area WHERE project_id = %s)
            """, [project_id])
            phase_count = cursor.fetchone()[0]

            cursor.execute("""
                SELECT COUNT(*) FROM landscape.tbl_parcel WHERE project_id = %s
            """, [project_id])
            parcel_count = cursor.fetchone()[0]

        if not row:
            return {
                'success': True,
                'level1_label': 'Area', 'level2_label': 'Phase', 'level3_label': 'Parcel',
                'level1_enabled': True, 'level2_enabled': True,
                'existing_areas': area_count, 'existing_phases': phase_count,
                'existing_parcels': parcel_count,
            }

        return {
            'success': True,
            'level1_label': row[0] or 'Area',
            'level2_label': row[1] or 'Phase',
            'level3_label': row[2] or 'Parcel',
            'level1_enabled': bool(row[3]),
            'level2_enabled': bool(row[4]),
            'auto_number': bool(row[5]),
            'existing_areas': area_count,
            'existing_phases': phase_count,
            'existing_parcels': parcel_count,
        }

    except Exception as e:
        logger.error(f"Error getting hierarchy config for project {project_id}: {e}")
        return {'success': False, 'error': str(e)}


@register_tool('stage_parcel_lots')
def stage_parcel_lots(doc_id: int = None, lots: list = None, phase_mapping: dict = None, **kwargs):
    """
    Stage parsed lot data into ai_extraction_staging for Workbench review.

    Takes lots from parse_spreadsheet_lots output and creates staging rows
    so the user can review/edit each lot before committing to tbl_parcel.

    Args:
        doc_id: The core_doc ID of the source spreadsheet
        lots: List of lot dicts from parse_spreadsheet_lots (label, lot_sf, unit_sf, acres)
        phase_mapping: Optional dict mapping group keys to phase names

    Returns:
        Staging summary with extraction IDs.
    """
    project_context = kwargs.get('project_context', {})
    project_id = project_context.get('project_id')
    tool_input = kwargs.get('tool_input', {})

    doc_id = doc_id or tool_input.get('doc_id')
    lots = lots or tool_input.get('lots', [])
    phase_mapping = phase_mapping or tool_input.get('phase_mapping', {})

    if not doc_id:
        return {'success': False, 'error': 'doc_id is required'}
    if not project_id:
        return {'success': False, 'error': 'project_id not available in context'}
    if not lots:
        return {'success': False, 'error': 'lots list is required (from parse_spreadsheet_lots)'}

    try:
        staged_ids = []
        with connection.cursor() as cursor:
            for idx, lot in enumerate(lots):
                label = lot.get('label', f'Lot {idx + 1}')
                lot_sf = lot.get('lot_sf')
                unit_sf = lot.get('unit_sf')
                acres = lot.get('acres')

                # Compute acres from lot_sf if not provided
                if acres is None and lot_sf:
                    acres = round(lot_sf / 43560.0, 4)

                # Build parcel data dict (mirrors tbl_parcel columns)
                parcel_data = {
                    'parcel_code': label,
                    'lot_area': lot_sf,
                    'acres_gross': acres,
                    'units_total': lot.get('units', 1),
                    'family_name': lot.get('family_name', 'Residential'),
                    'type_code': lot.get('type_code', ''),
                    'lot_product': f"{unit_sf:.0f} SF livable" if unit_sf else None,
                }

                # Determine phase group for scope_label
                street = re.sub(r'^\d+\s+', '', label)
                street = re.sub(r'\s+unit\s*\d+.*$', '', street, flags=re.IGNORECASE).strip()
                phase_name = phase_mapping.get(street, street)
                parcel_data['_phase_group'] = phase_name

                # Stage as a single JSON row (same pattern as rent roll units)
                cursor.execute("""
                    INSERT INTO landscape.ai_extraction_staging (
                        project_id, doc_id, field_key, extracted_value,
                        confidence_score, source_snippet, status, scope,
                        scope_label, array_index, target_table, target_field,
                        db_write_type, property_type, extraction_type, created_at
                    ) VALUES (
                        %s, %s, %s, %s, %s, %s, 'pending', %s, %s, %s, %s, %s, %s, %s, %s, NOW()
                    )
                    RETURNING extraction_id
                """, [
                    project_id,
                    doc_id,
                    'parcel_code',  # Primary field key
                    json.dumps(parcel_data),
                    0.85,  # Spreadsheet parse confidence
                    f"Row {lot.get('row_index', idx + 1)}: {label}",
                    'lot_inventory',  # scope
                    f"Lot {idx + 1}: {label}",  # scope_label
                    idx,  # array_index
                    'tbl_parcel',
                    'parcel_code',
                    'row_lot_inventory',
                    'land_development',
                    'spreadsheet_import',
                ])
                row = cursor.fetchone()
                if row:
                    staged_ids.append(row[0])

        return {
            'success': True,
            'staged_count': len(staged_ids),
            'staging_ids': staged_ids,
            'doc_id': doc_id,
            'message': (
                f"Staged {len(staged_ids)} lots for review in the Ingestion Workbench. "
                f"Open the Workbench to review and commit."
            ),
        }

    except Exception as e:
        logger.error(f"Error staging parcel lots for project {project_id}: {e}")
        return {'success': False, 'error': str(e)}


@register_tool('bulk_create_parcels', is_mutation=True)
def bulk_create_parcels(parcels: list = None, phase_mapping: dict = None, **kwargs):
    """
    Bulk-create parcels (and optionally phases) from structured lot data.

    This is a mutation tool — in propose_only mode it returns the proposed
    creates for user confirmation. On execution it creates the records.

    Args:
        parcels: List of parcel dicts, each with:
            - label: str (address or lot identifier, stored as parcel_code)
            - lot_sf: float (lot area in SF)
            - unit_sf: float (livable area in SF, stored as lot_product note)
            - acres: float (lot acreage)
            - phase_group: str (group key for phase assignment)
            - family_name: str (e.g., "Residential")
            - type_code: str (e.g., "SFD")
            - units: int (default 1)
        phase_mapping: Dict mapping group keys to phase names, e.g.:
            {"Lemhi Ct": "Block 8 - Lemhi Ct", "Waahni Ct": "Block 10 - Waahni Ct"}
            If None, all parcels go to a single auto-created phase.

    Returns:
        Created phase and parcel IDs with summary.
    """
    project_context = kwargs.get('project_context', {})
    project_id = project_context.get('project_id')
    tool_input = kwargs.get('tool_input', {})
    propose_only = kwargs.get('propose_only', True)

    parcels = parcels or tool_input.get('parcels', [])
    phase_mapping = phase_mapping or tool_input.get('phase_mapping', {})

    if not parcels:
        return {'success': False, 'error': 'parcels list is required'}
    if not project_id:
        return {'success': False, 'error': 'project_id not available in context'}

    # Build proposal summary
    groups = {}
    for p in parcels:
        grp = p.get('phase_group', 'Default')
        if grp not in groups:
            groups[grp] = []
        groups[grp].append(p)

    proposal = {
        'total_parcels': len(parcels),
        'phases_to_create': [],
        'parcels_by_phase': {},
    }

    for grp_key, grp_parcels in groups.items():
        phase_name = phase_mapping.get(grp_key, grp_key)
        proposal['phases_to_create'].append(phase_name)
        proposal['parcels_by_phase'][phase_name] = [
            {
                'label': p.get('label', ''),
                'lot_sf': p.get('lot_sf'),
                'unit_sf': p.get('unit_sf'),
                'acres': p.get('acres'),
                'units': p.get('units', 1),
                'family_name': p.get('family_name', 'Residential'),
            }
            for p in grp_parcels
        ]

    if propose_only:
        return {
            'success': True,
            'proposed': True,
            'action': 'bulk_create_parcels',
            'proposal': proposal,
            'message': (
                f"Propose creating {len(proposal['phases_to_create'])} phase(s) "
                f"and {len(parcels)} parcel(s). Confirm to apply."
            ),
        }

    # Execute: create phases and parcels
    try:
        created_phases = {}
        created_parcels = []

        with connection.cursor() as cursor:
            # Ensure at least one area exists (required FK)
            cursor.execute("""
                SELECT area_id, area_no FROM landscape.tbl_area
                WHERE project_id = %s ORDER BY area_no LIMIT 1
            """, [project_id])
            area_row = cursor.fetchone()

            if not area_row:
                # Create a default area
                cursor.execute("""
                    INSERT INTO landscape.tbl_area (project_id, area_no, label)
                    VALUES (%s, 1, 'Default')
                    RETURNING area_id, area_no
                """, [project_id])
                area_row = cursor.fetchone()

            area_id, area_no = area_row

            # Create phases
            for grp_key, grp_parcels in groups.items():
                phase_name = phase_mapping.get(grp_key, grp_key)

                # Get next phase_no
                cursor.execute("""
                    SELECT COALESCE(MAX(phase_no), 0) + 1
                    FROM landscape.tbl_phase
                    WHERE area_id = %s
                """, [area_id])
                next_phase_no = cursor.fetchone()[0]

                cursor.execute("""
                    INSERT INTO landscape.tbl_phase (area_id, phase_no, label)
                    VALUES (%s, %s, %s)
                    RETURNING phase_id, phase_no
                """, [area_id, next_phase_no, phase_name])
                phase_row = cursor.fetchone()
                created_phases[grp_key] = {
                    'phase_id': phase_row[0],
                    'phase_no': phase_row[1],
                    'label': phase_name,
                }

                # Create parcels for this phase
                for p in grp_parcels:
                    lot_sf = p.get('lot_sf')
                    unit_sf = p.get('unit_sf')
                    acres = p.get('acres')
                    # Convert lot_sf to acres if acres not provided
                    if acres is None and lot_sf:
                        acres = lot_sf / 43560.0

                    cursor.execute("""
                        INSERT INTO landscape.tbl_parcel (
                            project_id, area_id, phase_id,
                            parcel_code, lot_area, acres_gross,
                            units_total, family_name, type_code,
                            lot_product
                        ) VALUES (
                            %s, %s, %s,
                            %s, %s, %s,
                            %s, %s, %s,
                            %s
                        )
                        RETURNING parcel_id
                    """, [
                        project_id, area_id, phase_row[0],
                        p.get('label', ''),
                        lot_sf,
                        acres,
                        p.get('units', 1),
                        p.get('family_name', 'Residential'),
                        p.get('type_code', ''),
                        # Store unit_sf in lot_product as note until column exists
                        f"{unit_sf:.0f} SF livable" if unit_sf else None,
                    ])
                    parcel_row = cursor.fetchone()
                    created_parcels.append({
                        'parcel_id': parcel_row[0],
                        'label': p.get('label', ''),
                        'phase': phase_name,
                    })

        return {
            'success': True,
            'created_phases': list(created_phases.values()),
            'created_parcels': created_parcels,
            'summary': {
                'phases_created': len(created_phases),
                'parcels_created': len(created_parcels),
            },
            'message': (
                f"Created {len(created_phases)} phase(s) and "
                f"{len(created_parcels)} parcel(s) successfully."
            ),
        }

    except Exception as e:
        logger.error(f"Error bulk creating parcels for project {project_id}: {e}")
        return {'success': False, 'error': str(e)}
