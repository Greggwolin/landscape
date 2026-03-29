"""
Base class for DB-driven report preview generators.

Each generator implements generate_preview() which returns a dict matching
the ReportPreviewResponse TypeScript interface consumed by ReportViewer.tsx.

Preview response shape:
{
    'title': str,
    'subtitle': str | None,
    'as_of_date': str | None,
    'message': str | None,
    'sections': [
        {
            'heading': str,
            'type': 'table' | 'kpi_cards' | 'text',
            # For type='table':
            'columns': [{'key': str, 'label': str, 'align': str, 'format': str}],
            'rows': [dict],
            'totals': dict | None,
            # For type='kpi_cards':
            'cards': [{'label': str, 'value': str}],
            # For type='text':
            'content': str,
        }
    ]
}
"""

import io
from datetime import datetime
from decimal import Decimal
from django.db import connection


class PreviewBaseGenerator:
    """Base class for all report preview generators."""

    report_code = ''
    report_name = ''

    def __init__(self, project_id: int):
        self.project_id = project_id
        self._project_cache = None

    def generate_preview(self) -> dict:
        """
        Return preview data dict matching ReportPreviewData interface.

        Subclasses MUST implement this.
        """
        raise NotImplementedError(
            f"{self.__class__.__name__} must implement generate_preview()"
        )

    def generate_pdf(self) -> bytes:
        """
        Generic PDF export using reportlab.
        Renders the preview data (KPIs, tables, text) into a formatted PDF.
        Subclasses can override for custom layouts.
        """
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
            PageBreak, HRFlowable,
        )
        from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER

        preview = self.generate_preview()
        buf = io.BytesIO()

        page_size = letter

        doc = SimpleDocTemplate(
            buf,
            pagesize=page_size,
            rightMargin=0.5 * inch,
            leftMargin=0.5 * inch,
            topMargin=0.5 * inch,
            bottomMargin=0.5 * inch,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'ReportTitle', parent=styles['Heading1'],
            fontSize=13, spaceAfter=1, leading=15,
        )
        subtitle_style = ParagraphStyle(
            'ReportSubtitle', parent=styles['Normal'],
            fontSize=9, textColor=colors.grey, spaceAfter=2, leading=11,
        )
        heading_style = ParagraphStyle(
            'SectionHeading', parent=styles['Heading2'],
            fontSize=9, spaceBefore=6, spaceAfter=2, leading=11,
        )
        body_style = ParagraphStyle(
            'BodyText', parent=styles['Normal'],
            fontSize=8, spaceAfter=3, leading=10,
        )

        elements = []

        # Title
        elements.append(Paragraph(preview.get('title', self.report_name), title_style))
        if preview.get('subtitle'):
            elements.append(Paragraph(preview['subtitle'], subtitle_style))
        elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
        elements.append(Spacer(1, 2))

        # Sections
        for section in preview.get('sections', []):
            if section.get('heading'):
                elements.append(Paragraph(section['heading'], heading_style))

            if section['type'] == 'kpi_cards':
                cards = section.get('cards', [])
                if cards:
                    row_data = [[c['label'] for c in cards], [c['value'] for c in cards]]
                    t = Table(row_data, hAlign='LEFT')
                    t.setStyle(TableStyle([
                        ('FONTSIZE', (0, 0), (-1, 0), 7),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.grey),
                        ('FONTSIZE', (0, 1), (-1, 1), 10),
                        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 1),
                        ('TOPPADDING', (0, 1), (-1, 1), 0),
                        ('LEFTPADDING', (0, 0), (-1, -1), 8),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                    ]))
                    elements.append(t)
                    elements.append(Spacer(1, 2))

            elif section['type'] == 'table':
                columns = section.get('columns', [])
                rows = section.get('rows', [])
                totals = section.get('totals')

                if columns and rows:
                    num_cols = len(columns)
                    page_w = page_size[0]  # actual page width (portrait or landscape)
                    avail_width = page_w - 1.0 * inch  # minus left+right margins

                    # --- Content-aware column width calculation ---
                    # Measure max content length per column to allocate proportionally
                    col_max_len = []
                    for ci, c in enumerate(columns):
                        max_len = len(c['label'])
                        for row in rows[:50]:  # sample first 50 rows
                            val = row.get(c['key'])
                            cell_text = self._pdf_format_cell(val, c.get('format'))
                            max_len = max(max_len, len(str(cell_text)))
                        col_max_len.append(max_len)

                    # Ensure minimum width: numeric/currency cols need at least
                    # 10 chars ($1,234,567) to avoid wrapping
                    for ci, c in enumerate(columns):
                        fmt = c.get('format', '')
                        if fmt in ('currency', 'number', 'percentage'):
                            col_max_len[ci] = max(col_max_len[ci], 10)
                        else:
                            col_max_len[ci] = max(col_max_len[ci], 6)

                    # Give text columns (left-aligned) a small weight boost for
                    # readability, but not so much that numeric columns get squeezed
                    weighted_lens = []
                    for ci, c in enumerate(columns):
                        w = col_max_len[ci]
                        if c.get('align', 'left') == 'left' and c.get('format') not in ('currency', 'number', 'percentage'):
                            w *= 1.2
                        weighted_lens.append(w)

                    total_weight = sum(weighted_lens)
                    col_widths = [(w / total_weight) * avail_width for w in weighted_lens]

                    # --- Paragraph styles for wrappable cells (compact) ---
                    cell_style_left = ParagraphStyle(
                        'CellLeft', fontSize=7, fontName='Helvetica',
                        leading=9, alignment=TA_LEFT,
                    )
                    cell_style_right = ParagraphStyle(
                        'CellRight', fontSize=7, fontName='Helvetica',
                        leading=9, alignment=TA_RIGHT,
                    )
                    cell_style_left_bold = ParagraphStyle(
                        'CellLeftBold', fontSize=7, fontName='Helvetica-Bold',
                        leading=9, alignment=TA_LEFT,
                    )
                    cell_style_right_bold = ParagraphStyle(
                        'CellRightBold', fontSize=7, fontName='Helvetica-Bold',
                        leading=9, alignment=TA_RIGHT,
                    )
                    header_style_left = ParagraphStyle(
                        'HdrLeft', fontSize=7, fontName='Helvetica-Bold',
                        leading=9, alignment=TA_LEFT,
                    )
                    header_style_right = ParagraphStyle(
                        'HdrRight', fontSize=7, fontName='Helvetica-Bold',
                        leading=9, alignment=TA_RIGHT,
                    )

                    def _para(text, bold=False, align='left'):
                        """Wrap cell text in a Paragraph for proper word-wrapping."""
                        # Escape XML special chars for reportlab
                        safe = str(text).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                        if bold and align == 'right':
                            return Paragraph(safe, cell_style_right_bold)
                        if bold:
                            return Paragraph(safe, cell_style_left_bold)
                        if align == 'right':
                            return Paragraph(safe, cell_style_right)
                        return Paragraph(safe, cell_style_left)

                    # Build header row with Paragraphs
                    header = []
                    for c in columns:
                        align = c.get('align', 'left')
                        sty = header_style_right if align == 'right' else header_style_left
                        safe = str(c['label']).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                        header.append(Paragraph(safe, sty))
                    table_data = [header]

                    # Track _rowStyle per data row (offset +1 for header)
                    row_styles = []
                    for row in rows:
                        rs = row.get('_rowStyle', '')
                        row_styles.append(rs)
                        is_bold = rs in ('header', 'subtotal', 'total')
                        formatted = []
                        for ci, c in enumerate(columns):
                            val = row.get(c['key'])
                            cell_text = self._pdf_format_cell(val, c.get('format'))
                            col_align = c.get('align', 'left')
                            formatted.append(_para(cell_text, bold=is_bold, align=col_align))
                        table_data.append(formatted)

                    if totals:
                        totals_formatted = []
                        for i, c in enumerate(columns):
                            if i == 0:
                                totals_formatted.append(_para('Total', bold=True, align='left'))
                            else:
                                val = self._pdf_format_cell(totals.get(c['key']), c.get('format'))
                                totals_formatted.append(_para(val, bold=True, align=c.get('align', 'left')))
                        table_data.append(totals_formatted)

                    t = Table(table_data, colWidths=col_widths, repeatRows=1)
                    style_cmds = [
                        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.93, 0.93, 0.93)),
                        ('LINEBELOW', (0, 0), (-1, 0), 0.5, colors.grey),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.97, 0.97, 0.97)]),
                        ('LEFTPADDING', (0, 0), (-1, -1), 3),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                        ('TOPPADDING', (0, 0), (-1, -1), 2),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ]

                    # Per-row styling based on _rowStyle
                    for ri, rs in enumerate(row_styles):
                        data_row = ri + 1  # +1 for header row
                        if rs == 'header':
                            style_cmds.append(('BACKGROUND', (0, data_row), (-1, data_row), colors.Color(0.93, 0.93, 0.93)))
                        elif rs == 'indent':
                            style_cmds.append(('LEFTPADDING', (0, data_row), (0, data_row), 20))
                        elif rs == 'total':
                            style_cmds.append(('LINEABOVE', (0, data_row), (-1, data_row), 1.5, colors.black))
                            style_cmds.append(('LINEBELOW', (0, data_row), (-1, data_row), 1.5, colors.black))

                    if totals:
                        style_cmds.append(('LINEABOVE', (0, -1), (-1, -1), 1.5, colors.black))

                    t.setStyle(TableStyle(style_cmds))
                    elements.append(t)
                    elements.append(Spacer(1, 2))

            elif section['type'] == 'map':
                # Render static map image using Pillow / staticmap
                import logging
                logger = logging.getLogger(__name__)
                try:
                    map_img = self._render_map_image(section)
                    if map_img:
                        from reportlab.platypus import Image as RLImage
                        map_w = min(page_size[0] - 1.0 * inch, 7.5 * inch)
                        map_h = map_w * 0.45  # compact aspect
                        elements.append(RLImage(map_img, width=map_w, height=map_h))
                        elements.append(Spacer(1, 2))
                    else:
                        logger.warning("_render_map_image returned None for section: %s", section.get('heading'))
                        elements.append(Paragraph(
                            f"[Map: {section.get('heading', 'Comparable Locations')} — image unavailable]",
                            body_style
                        ))
                except Exception as e:
                    logger.error("Map rendering failed: %s", e, exc_info=True)
                    elements.append(Paragraph(
                        f"[Map rendering error: {e}]", body_style
                    ))

            elif section['type'] == 'text':
                content = section.get('content', '')
                elements.append(Paragraph(content, body_style))

        if not elements:
            elements.append(Paragraph(
                preview.get('message', 'No data available.'), body_style
            ))

        doc.build(elements)
        return buf.getvalue()

    @staticmethod
    def _pdf_format_cell(value, fmt: str = None) -> str:
        """Format a cell value for PDF output."""
        if value is None:
            return '—'
        if value == '':
            return ''
        if fmt == 'currency':
            try:
                num = float(value)
                if num == 0:
                    return '—'
                return f"${num:,.0f}"
            except (ValueError, TypeError):
                return str(value)
        if fmt == 'number':
            try:
                num = float(value)
                if num == 0:
                    return '—'
                return f"{num:,.0f}"
            except (ValueError, TypeError):
                return str(value)
        if fmt == 'percentage':
            try:
                num = float(value)
                if num == 0:
                    return '—'
                return f"{num:.1f}%"
            except (ValueError, TypeError):
                return str(value)
        return str(value)

    @staticmethod
    def _render_map_image(section: dict):
        """Render a static map image with real OSM map tiles using staticmap.

        Falls back to a Pillow-only coordinate plot if staticmap is unavailable.
        Returns an io.BytesIO PNG image buffer, or None if rendering fails.
        """
        from PIL import ImageDraw, ImageFont

        markers = section.get('markers', [])
        if not markers:
            return None

        W, H = 1050, 600

        def hex_to_rgb(h):
            h = h.lstrip('#')
            return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))

        try:
            font_number = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 12)
            font_name = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 10)
        except (OSError, IOError):
            font_number = ImageFont.load_default()
            font_name = font_number

        # --- try staticmap (real OSM tiles) ---
        try:
            from staticmap import StaticMap, CircleMarker

            m = StaticMap(W, H, url_template='https://tile.openstreetmap.org/{z}/{x}/{y}.png')

            # Add circle markers so staticmap calculates correct bounds/zoom
            for marker in markers:
                lng, lat = marker['coordinates']
                color = hex_to_rgb(marker.get('color', '#888888'))
                is_subject = marker.get('id') == 'subject'
                size = 10 if is_subject else 8
                m.add_marker(CircleMarker((lng, lat), color, size))

            # Render the base map with tiles
            img = m.render()
            draw = ImageDraw.Draw(img)

            # Now draw our styled markers on top of the rendered tile image
            subject_markers = [mk for mk in markers if mk.get('id') == 'subject']
            comp_markers = [mk for mk in markers if mk.get('id') != 'subject']

            R_COMP = 14
            R_SUBJ = 16

            def to_px(lng, lat):
                """Convert lng/lat to pixel on the rendered staticmap image."""
                # Use staticmap's internal projection
                px_x, px_y = m._x_to_px(m._lon_to_x(lng)), m._y_to_px(m._lat_to_y(lat))
                return int(px_x), int(px_y)

            # Draw comps first, subject on top
            for mk in comp_markers:
                lng, lat = mk['coordinates']
                px, py = to_px(lng, lat)
                color = hex_to_rgb(mk.get('color', '#888888'))
                label = mk.get('label', '')
                name = mk.get('name', '')

                # White halo behind circle for visibility on any tile
                draw.ellipse(
                    [(px - R_COMP - 2, py - R_COMP - 2), (px + R_COMP + 2, py + R_COMP + 2)],
                    fill=(255, 255, 255), outline=None
                )
                draw.ellipse(
                    [(px - R_COMP, py - R_COMP), (px + R_COMP, py + R_COMP)],
                    fill=color, outline=(255, 255, 255), width=2
                )
                draw.text((px, py), label, fill=(255, 255, 255), font=font_number, anchor='mm')
                if name:
                    # Text with white outline for readability on map tiles
                    for dx in (-1, 0, 1):
                        for dy in (-1, 0, 1):
                            if dx or dy:
                                draw.text((px + R_COMP + 6 + dx, py - 2 + dy), name,
                                          fill=(255, 255, 255), font=font_name, anchor='lm')
                    draw.text((px + R_COMP + 6, py - 2), name, fill=(30, 30, 40), font=font_name, anchor='lm')

            for mk in subject_markers:
                lng, lat = mk['coordinates']
                px, py = to_px(lng, lat)
                color = hex_to_rgb(mk.get('color', '#2d8cf0'))

                draw.ellipse(
                    [(px - R_SUBJ - 2, py - R_SUBJ - 2), (px + R_SUBJ + 2, py + R_SUBJ + 2)],
                    fill=(255, 255, 255), outline=None
                )
                draw.ellipse(
                    [(px - R_SUBJ, py - R_SUBJ), (px + R_SUBJ, py + R_SUBJ)],
                    fill=color, outline=(255, 255, 255), width=2
                )
                draw.text((px, py), 'S', fill=(255, 255, 255), font=font_number, anchor='mm')
                for dx in (-1, 0, 1):
                    for dy in (-1, 0, 1):
                        if dx or dy:
                            draw.text((px + R_SUBJ + 6 + dx, py - 2 + dy), 'Subject',
                                      fill=(255, 255, 255), font=font_name, anchor='lm')
                draw.text((px + R_SUBJ + 6, py - 2), 'Subject', fill=(30, 30, 40), font=font_name, anchor='lm')

            buf = io.BytesIO()
            img.save(buf, format='PNG')
            buf.seek(0)
            return buf

        except (ImportError, Exception) as exc:
            # staticmap not installed or tile fetch failed — fall back to Pillow coordinate plot
            import logging
            logging.getLogger(__name__).warning("staticmap failed (%s), using fallback map renderer", exc)
            from PIL import Image
            img = Image.new('RGB', (W, H), (245, 245, 248))
            draw = ImageDraw.Draw(img)

            all_lngs = [mk['coordinates'][0] for mk in markers]
            all_lats = [mk['coordinates'][1] for mk in markers]
            min_lng, max_lng = min(all_lngs), max(all_lngs)
            min_lat, max_lat = min(all_lats), max(all_lats)
            lng_range = (max_lng - min_lng) or 0.01
            lat_range = (max_lat - min_lat) or 0.01
            min_lng -= lng_range * 0.25
            max_lng += lng_range * 0.25
            min_lat -= lat_range * 0.25
            max_lat += lat_range * 0.25
            lng_range = max_lng - min_lng
            lat_range = max_lat - min_lat
            PAD = 40

            def to_px_fb(lng, lat):
                x = PAD + (lng - min_lng) / lng_range * (W - 2 * PAD)
                y = (H - PAD) - (lat - min_lat) / lat_range * (H - 2 * PAD)
                return int(x), int(y)

            for mk in markers:
                lng, lat = mk['coordinates']
                px, py = to_px_fb(lng, lat)
                color = hex_to_rgb(mk.get('color', '#888888'))
                draw.ellipse([(px - 12, py - 12), (px + 12, py + 12)],
                             fill=color, outline=(60, 60, 70), width=2)
                draw.text((px, py), mk.get('label', ''), fill=(255, 255, 255),
                          font=font_number, anchor='mm')

            buf = io.BytesIO()
            img.save(buf, format='PNG')
            buf.seek(0)
            return buf

    def generate_excel(self) -> bytes:
        """
        Generic Excel export using openpyxl.
        Renders the preview data (KPIs, tables, text) into a formatted workbook.
        Subclasses can override for custom layouts.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        preview = self.generate_preview()
        wb = Workbook()
        ws = wb.active
        ws.title = (preview.get('title', self.report_name))[:31]  # Excel 31-char limit

        # Styles
        title_font = Font(size=14, bold=True)
        subtitle_font = Font(size=10, color='666666')
        heading_font = Font(size=11, bold=True)
        header_font = Font(size=9, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell_font = Font(size=9)
        total_font = Font(size=9, bold=True)
        total_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        thin_border = Border(
            bottom=Side(style='thin', color='CCCCCC'),
        )

        row_num = 1

        # Title
        ws.cell(row=row_num, column=1, value=preview.get('title', self.report_name)).font = title_font
        row_num += 1
        if preview.get('subtitle'):
            ws.cell(row=row_num, column=1, value=preview['subtitle']).font = subtitle_font
            row_num += 1
        row_num += 1  # blank row

        for section in preview.get('sections', []):
            if section.get('heading'):
                ws.cell(row=row_num, column=1, value=section['heading']).font = heading_font
                row_num += 1

            if section['type'] == 'kpi_cards':
                cards = section.get('cards', [])
                for i, card in enumerate(cards):
                    col = i + 1
                    ws.cell(row=row_num, column=col, value=card['label']).font = Font(size=8, color='888888')
                    ws.cell(row=row_num + 1, column=col, value=card['value']).font = Font(size=12, bold=True)
                row_num += 3  # labels + values + blank

            elif section['type'] == 'table':
                columns = section.get('columns', [])
                rows = section.get('rows', [])
                totals = section.get('totals')

                if columns:
                    # Headers
                    for i, col in enumerate(columns):
                        cell = ws.cell(row=row_num, column=i + 1, value=col['label'])
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(
                            horizontal='right' if col.get('align') == 'right' else 'left',
                            wrap_text=True,
                        )
                    row_num += 1

                    # Data rows
                    for data_row in rows:
                        for i, col in enumerate(columns):
                            val = data_row.get(col['key'])
                            cell = ws.cell(row=row_num, column=i + 1, value=val)
                            cell.font = cell_font
                            cell.border = thin_border
                            if col.get('align') == 'right':
                                cell.alignment = Alignment(horizontal='right')
                            # Apply number format
                            fmt = col.get('format')
                            if fmt == 'currency' and isinstance(val, (int, float)):
                                cell.number_format = '$#,##0'
                            elif fmt == 'number' and isinstance(val, (int, float)):
                                cell.number_format = '#,##0'
                            elif fmt == 'percentage' and isinstance(val, (int, float)):
                                cell.number_format = '0.0"%"'
                        row_num += 1

                    # Totals row
                    if totals:
                        for i, col in enumerate(columns):
                            if i == 0:
                                cell = ws.cell(row=row_num, column=1, value='Total')
                            else:
                                val = totals.get(col['key'])
                                cell = ws.cell(row=row_num, column=i + 1, value=val)
                                fmt = col.get('format')
                                if fmt == 'currency' and isinstance(val, (int, float)):
                                    cell.number_format = '$#,##0'
                                elif fmt == 'number' and isinstance(val, (int, float)):
                                    cell.number_format = '#,##0'
                            cell.font = total_font
                            cell.fill = total_fill
                        row_num += 1

                    # Auto-size columns (approximate)
                    for i, col in enumerate(columns):
                        col_letter = get_column_letter(i + 1)
                        max_len = len(col['label'])
                        for data_row in rows[:20]:  # sample first 20 rows
                            val = str(data_row.get(col['key'], ''))
                            max_len = max(max_len, len(val))
                        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)

                row_num += 1  # blank row after table

            elif section['type'] == 'text':
                ws.cell(row=row_num, column=1, value=section.get('content', '')).font = cell_font
                row_num += 2

        # Message if no sections
        if not preview.get('sections') and preview.get('message'):
            ws.cell(row=row_num, column=1, value=preview['message']).font = cell_font

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ─── Helper: project info ────────────────────────────────────────────

    def get_project(self) -> dict:
        """Fetch basic project info. Cached per instance."""
        if self._project_cache:
            return self._project_cache

        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT project_id, project_name, project_type_code,
                       city, state
                FROM landscape.tbl_project
                WHERE project_id = %s
            """, [self.project_id])
            row = cursor.fetchone()
            if not row:
                self._project_cache = {
                    'project_id': self.project_id,
                    'project_name': f'Project {self.project_id}',
                }
                return self._project_cache

            cols = [c.name for c in cursor.description]
            self._project_cache = dict(zip(cols, row))
        return self._project_cache

    # ─── Helper: SQL execution ───────────────────────────────────────────

    def execute_query(self, sql: str, params: list = None) -> list[dict]:
        """Execute SQL and return list of dicts. Returns [] on SQL error."""
        import logging
        try:
            with connection.cursor() as cursor:
                cursor.execute(sql, params or [])
                cols = [c.name for c in cursor.description]
                return [dict(zip(cols, row)) for row in cursor.fetchall()]
        except Exception as e:
            logging.getLogger(__name__).warning("Report query failed: %s", e)
            return []

    def execute_scalar(self, sql: str, params: list = None):
        """Execute SQL and return single value. Returns None on SQL error."""
        import logging
        try:
            with connection.cursor() as cursor:
                cursor.execute(sql, params or [])
                row = cursor.fetchone()
                return row[0] if row else None
        except Exception as e:
            logging.getLogger(__name__).warning("Report scalar query failed: %s", e)
            return None

    def get_today_str(self) -> str:
        """Return today's date as 'Mon DD, YYYY' string for PDF subtitles."""
        return datetime.now().strftime('%b %d, %Y')

    # ─── Helper: formatting ──────────────────────────────────────────────

    @staticmethod
    def fmt_currency(value, decimals: int = 0) -> str:
        """Format as currency string."""
        if value is None:
            return '—'
        num = float(value)
        if decimals == 0:
            return f"${num:,.0f}"
        return f"${num:,.{decimals}f}"

    @staticmethod
    def fmt_number(value, decimals: int = 0) -> str:
        """Format as number string."""
        if value is None:
            return '—'
        num = float(value)
        if decimals == 0:
            return f"{num:,.0f}"
        return f"{num:,.{decimals}f}"

    @staticmethod
    def fmt_pct(value, decimals: int = 1) -> str:
        """Format as percentage string."""
        if value is None:
            return '—'
        return f"{float(value):.{decimals}f}%"

    @staticmethod
    def safe_div(numerator, denominator, default=0):
        """Safe division avoiding ZeroDivisionError."""
        if not denominator:
            return default
        return float(numerator) / float(denominator)

    # ─── Helper: section builders ────────────────────────────────────────

    def make_table_section(
        self,
        heading: str,
        columns: list[dict],
        rows: list[dict],
        totals: dict = None,
    ) -> dict:
        """Build a table section dict."""
        section = {
            'heading': heading,
            'type': 'table',
            'columns': columns,
            'rows': rows,
        }
        if totals:
            section['totals'] = totals
        return section

    def make_kpi_section(self, heading: str, cards: list[dict]) -> dict:
        """Build a KPI cards section dict."""
        return {
            'heading': heading,
            'type': 'kpi_cards',
            'cards': cards,
        }

    def make_text_section(self, heading: str, content: str) -> dict:
        """Build a text section dict."""
        return {
            'heading': heading,
            'type': 'text',
            'content': content,
        }

    def make_map_section(self, heading: str, center: list, markers: list) -> dict:
        """Build a map section dict for static map rendering.

        Args:
            heading: Section title
            center: [lng, lat] center point
            markers: list of dicts with keys: id, coordinates ([lng,lat]),
                     color, label, name (optional), variant (optional)
        """
        return {
            'heading': heading,
            'type': 'map',
            'center': center,
            'markers': markers,
        }

    def make_kpi_card(self, label: str, value) -> dict:
        """Build a single KPI card dict."""
        return {'label': label, 'value': str(value)}
