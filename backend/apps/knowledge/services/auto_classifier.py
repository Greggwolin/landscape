"""
Auto-Classification Service for Knowledge Library Uploads

When a file is uploaded via the Knowledge Library drop zone, this service:
1. Extracts text from the file bytes (PDF, DOCX, XLSX, TXT)
2. Classifies the document type (Offering, Operations, Market Data, etc.)
3. Infers geographic tags (state, city, MSA, county)
4. Infers property type (MF, LAND, RET, etc.)
5. Stores all results in the appropriate tables

Uses rule-based keyword matching — no API calls required.
"""

import logging
import os
import re
import tempfile
from typing import Dict, Any, List, Optional, Tuple

from django.db import connection

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────
# Text extraction from raw file bytes
# ─────────────────────────────────────────────

# PDF
try:
    import fitz  # PyMuPDF
    HAS_PYMUPDF = True
except ImportError:
    HAS_PYMUPDF = False

# DOCX
try:
    from docx import Document as DocxDocument
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False

# Excel
try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def extract_text_from_bytes(
    file_bytes: bytes,
    filename: str,
    mime_type: str = '',
) -> Tuple[Optional[str], Optional[str]]:
    """
    Extract text from raw file bytes.

    Returns (extracted_text, error_message).
    """
    if not file_bytes:
        return None, "Empty file"

    ext = os.path.splitext(filename)[1].lower()
    if not mime_type:
        mime_type = _mime_from_ext(ext)

    suffix = ext or '.bin'
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name

        if mime_type == 'application/pdf' or ext == '.pdf':
            return _extract_pdf(tmp_path), None
        elif ext == '.docx' or 'wordprocessingml' in mime_type:
            return _extract_docx(tmp_path), None
        elif ext == '.xlsx' or 'spreadsheetml' in mime_type:
            return _extract_xlsx(tmp_path), None
        elif ext in ('.txt', '.md', '.csv', '.json', '.xml') or mime_type.startswith('text/'):
            return _extract_text(tmp_path), None
        else:
            return None, f"Unsupported file type: {ext} ({mime_type})"

    except Exception as e:
        return None, f"Extraction failed: {e}"
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass


def _extract_pdf(path: str) -> Optional[str]:
    if not HAS_PYMUPDF:
        return None
    parts = []
    with fitz.open(path) as doc:
        for page in doc:
            parts.append(page.get_text())
    return "\n\n".join(parts).strip() or None


def _extract_docx(path: str) -> Optional[str]:
    if not HAS_DOCX:
        return None
    doc = DocxDocument(path)
    return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip()).strip() or None


def _extract_xlsx(path: str) -> Optional[str]:
    if not HAS_OPENPYXL:
        return None
    wb = load_workbook(path, read_only=True, data_only=True)
    parts = []
    for name in wb.sheetnames:
        lines = [f"=== Sheet: {name} ==="]
        for row in wb[name].iter_rows(values_only=True):
            if any(c is not None for c in row):
                lines.append('\t'.join(str(c) if c is not None else '' for c in row))
        if len(lines) > 1:
            parts.append('\n'.join(lines))
    wb.close()
    return '\n\n'.join(parts).strip() or None


def _extract_text(path: str) -> Optional[str]:
    with open(path, 'r', encoding='utf-8', errors='ignore') as f:
        return f.read().strip() or None


def _mime_from_ext(ext: str) -> str:
    return {
        '.pdf': 'application/pdf',
        '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        '.txt': 'text/plain',
        '.csv': 'text/csv',
        '.md': 'text/markdown',
        '.json': 'application/json',
    }.get(ext, 'application/octet-stream')


# ─────────────────────────────────────────────
# Document type classification (keyword-based)
# ─────────────────────────────────────────────

# Maps each document_type (from tbl_extraction_mapping) to keyword patterns.
# Score = number of unique keyword hits.
DOC_TYPE_PATTERNS: Dict[str, List[str]] = {
    'Offering': [
        'offering memorandum', 'investment summary', 'investment highlights',
        'confidential offering', 'executive summary', 'asking price',
        'cap rate', 'price per unit', 'investment thesis', 'offering price',
        'broker opinion', 'property overview', 'rent comparable',
        'financial overview', 'debt assumption', 'loan assumption',
    ],
    'Operations': [
        'rent roll', 'unit mix', 'occupancy', 'tenant', 'lease',
        'operating statement', 'income statement', 'trailing twelve',
        't-12', 't12', 'net operating income', 'vacancy', 'effective gross',
        'gross potential rent', 'concessions', 'loss to lease',
    ],
    'Market Data': [
        'market survey', 'market analysis', 'comp survey', 'comparable',
        'submarket', 'demographics', 'population', 'employment',
        'absorption', 'vacancy rate', 'market rent', 'market trends',
        'economic overview', 'supply pipeline', 'new construction',
    ],
    'Property Data': [
        'appraisal', 'site plan', 'survey', 'plat', 'zoning',
        'environmental', 'phase i', 'phase ii', 'soil report',
        'geotechnical', 'title report', 'property condition',
        'engineering', 'inspection', 'as-built',
    ],
    'Diligence': [
        'due diligence', 'inspection report', 'title commitment',
        'alta survey', 'environmental assessment', 'property condition assessment',
        'pca', 'phase i esa', 'title insurance', 'lien search',
    ],
    'Accounting': [
        'financial statement', 'balance sheet', 'profit and loss',
        'general ledger', 'chart of accounts', 'accounts payable',
        'accounts receivable', 'tax return', 'depreciation schedule',
        'capital expenditure', 'capex',
    ],
    'Agreements': [
        'purchase agreement', 'purchase and sale', 'letter of intent',
        'loi', 'contract', 'amendment', 'addendum', 'closing statement',
        'hud-1', 'settlement statement', 'escrow',
    ],
}


def classify_document_type(
    text: str,
    filename: str = '',
) -> Dict[str, Any]:
    """
    Classify document type using keyword matching.

    Returns dict with:
      - doc_type: best-matching type or 'general'
      - confidence: 0.0–1.0
      - matched_patterns: list of matched keywords
    """
    if not text:
        return {'doc_type': 'general', 'confidence': 0.0, 'matched_patterns': []}

    # Use first 5000 chars for classification (covers most document headers)
    sample = text[:5000].lower()
    # Also check filename
    fname_lower = filename.lower()

    best_type = 'general'
    best_score = 0
    best_matches: List[str] = []

    for doc_type, patterns in DOC_TYPE_PATTERNS.items():
        matches = []
        for pattern in patterns:
            if pattern in sample or pattern in fname_lower:
                matches.append(pattern)

        if len(matches) > best_score:
            best_score = len(matches)
            best_type = doc_type
            best_matches = matches

    # Confidence: normalize by pattern count, cap at 1.0
    max_patterns = len(DOC_TYPE_PATTERNS.get(best_type, []))
    confidence = min(best_score / max(max_patterns * 0.3, 1), 1.0) if best_score > 0 else 0.0

    return {
        'doc_type': best_type,
        'confidence': round(confidence, 2),
        'matched_patterns': best_matches,
    }


# ─────────────────────────────────────────────
# Geographic tag inference
# ─────────────────────────────────────────────

US_STATES = {
    'alabama': 'AL', 'alaska': 'AK', 'arizona': 'AZ', 'arkansas': 'AR',
    'california': 'CA', 'colorado': 'CO', 'connecticut': 'CT', 'delaware': 'DE',
    'florida': 'FL', 'georgia': 'GA', 'hawaii': 'HI', 'idaho': 'ID',
    'illinois': 'IL', 'indiana': 'IN', 'iowa': 'IA', 'kansas': 'KS',
    'kentucky': 'KY', 'louisiana': 'LA', 'maine': 'ME', 'maryland': 'MD',
    'massachusetts': 'MA', 'michigan': 'MI', 'minnesota': 'MN', 'mississippi': 'MS',
    'missouri': 'MO', 'montana': 'MT', 'nebraska': 'NE', 'nevada': 'NV',
    'new hampshire': 'NH', 'new jersey': 'NJ', 'new mexico': 'NM', 'new york': 'NY',
    'north carolina': 'NC', 'north dakota': 'ND', 'ohio': 'OH', 'oklahoma': 'OK',
    'oregon': 'OR', 'pennsylvania': 'PA', 'rhode island': 'RI', 'south carolina': 'SC',
    'south dakota': 'SD', 'tennessee': 'TN', 'texas': 'TX', 'utah': 'UT',
    'vermont': 'VT', 'virginia': 'VA', 'washington': 'WA', 'west virginia': 'WV',
    'wisconsin': 'WI', 'wyoming': 'WY',
}

# Reverse: abbreviation → full name
STATE_ABBREV_TO_NAME = {v: k.title() for k, v in US_STATES.items()}

# Major metro areas / MSAs to detect
MAJOR_METROS = [
    'Phoenix', 'Los Angeles', 'New York', 'Chicago', 'Houston', 'Dallas',
    'San Antonio', 'San Diego', 'San Jose', 'Austin', 'Jacksonville',
    'San Francisco', 'Denver', 'Nashville', 'Seattle', 'Washington DC',
    'Atlanta', 'Boston', 'Miami', 'Tampa', 'Orlando', 'Minneapolis',
    'Charlotte', 'Portland', 'Las Vegas', 'Tucson', 'Scottsdale', 'Mesa',
    'Peoria', 'Chandler', 'Gilbert', 'Glendale', 'Tempe', 'Surprise',
    'Raleigh', 'Salt Lake City', 'Kansas City', 'Indianapolis',
    'Columbus', 'Pittsburgh', 'Sacramento', 'St. Louis', 'Milwaukee',
    'Albuquerque', 'Oklahoma City', 'Boise', 'Richmond', 'Louisville',
]


def infer_geo_tags(text: str) -> List[Dict[str, str]]:
    """
    Extract geographic references from document text.

    Returns list of dicts: [{'geo_level': 'state', 'geo_value': 'AZ', 'geo_source': 'ai_extracted'}, ...]
    """
    if not text:
        return []

    # Use first 8000 chars (enough for headers, property summary, location sections)
    sample = text[:8000]
    sample_lower = sample.lower()
    tags: List[Dict[str, str]] = []
    seen = set()

    # 1. Match "City, ST" patterns (e.g., "Phoenix, AZ" or "Los Angeles, CA")
    # Only match capitalized words (proper nouns) before the comma, on the same line
    city_state_pattern = re.compile(
        r'\b([A-Z][a-z]+(?: [A-Z][a-z]+){0,3}),\s*([A-Z]{2})\b'
    )
    for match in city_state_pattern.finditer(sample):
        city_name = match.group(1).strip()
        state_abbr = match.group(2)
        if state_abbr in STATE_ABBREV_TO_NAME:
            state_key = ('state', state_abbr)
            if state_key not in seen:
                tags.append({'geo_level': 'state', 'geo_value': state_abbr, 'geo_source': 'ai_extracted'})
                seen.add(state_key)
            city_key = ('city', city_name)
            if city_key not in seen:
                tags.append({'geo_level': 'city', 'geo_value': city_name, 'geo_source': 'ai_extracted'})
                seen.add(city_key)

    # 2. Match full state names (e.g., "State of California", "Arizona market")
    for state_name, abbr in US_STATES.items():
        if state_name in sample_lower:
            key = ('state', abbr)
            if key not in seen:
                tags.append({'geo_level': 'state', 'geo_value': abbr, 'geo_source': 'ai_extracted'})
                seen.add(key)

    # 3. Match county names (e.g., "Maricopa County", "Los Angeles County")
    # Only match 1-3 capitalized words immediately before "County"
    county_pattern = re.compile(r'\b([A-Z][a-z]+(?:\s+[A-Z][a-z]+){0,2})\s+County\b')
    for match in county_pattern.finditer(sample):
        county = match.group(1).strip()
        key = ('county', county)
        if key not in seen and len(county) > 3:
            tags.append({'geo_level': 'county', 'geo_value': f"{county} County", 'geo_source': 'ai_extracted'})
            seen.add(key)

    # 4. Match MSA names (e.g., "Phoenix MSA", "Phoenix-Mesa-Chandler MSA")
    # Only match capitalized proper nouns (optionally hyphenated) before MSA/Metro
    # Skip common preceding words like "the", "growing", etc.
    msa_pattern = re.compile(r'(?:[Tt]he\s+)?([A-Z][a-z]+(?:[-][A-Z][a-z]+){0,3})\s+(?:MSA|Metropolitan|Metro(?:\s+Area)?)\b')
    for match in msa_pattern.finditer(sample):
        msa = match.group(1).strip()
        key = ('msa', msa)
        if key not in seen and len(msa) > 3:
            tags.append({'geo_level': 'msa', 'geo_value': f"{msa} MSA", 'geo_source': 'ai_extracted'})
            seen.add(key)

    # 5. Match major city names mentioned in text
    for city in MAJOR_METROS:
        if city.lower() in sample_lower:
            key = ('city', city)
            if key not in seen:
                tags.append({'geo_level': 'city', 'geo_value': city, 'geo_source': 'ai_extracted'})
                seen.add(key)

    return tags


# ─────────────────────────────────────────────
# Property type inference
# ─────────────────────────────────────────────

PROPERTY_TYPE_PATTERNS: Dict[str, List[str]] = {
    'MF': [
        'multifamily', 'apartment', 'unit mix', 'rent roll', 'garden style',
        'mid-rise', 'high-rise', 'studio', 'one-bedroom', 'two-bedroom',
        'three-bedroom', 'average rent', 'per unit', 'occupancy rate',
        'student housing', 'senior living', 'affordable housing',
    ],
    'LAND': [
        'land development', 'subdivision', 'master planned', 'lot',
        'parcel', 'acreage', 'entitled', 'grading', 'infrastructure',
        'platted', 'builder', 'homebuilder', 'single family',
        'density', 'zoning', 'entitlement',
    ],
    'RET': [
        'retail', 'shopping center', 'mall', 'tenant', 'anchor',
        'inline', 'pad site', 'triple net', 'nnn', 'cam',
        'percentage rent', 'retail sales',
    ],
    'OFF': [
        'office', 'class a office', 'class b office', 'coworking',
        'flex space', 'office park', 'rsf', 'usable sf',
        'full service gross', 'modified gross',
    ],
    'IND': [
        'industrial', 'warehouse', 'distribution', 'logistics',
        'manufacturing', 'clear height', 'dock high', 'cross dock',
        'cold storage', 'fulfillment',
    ],
    'HTL': [
        'hotel', 'hospitality', 'room count', 'revpar',
        'average daily rate', 'adr', 'occupancy', 'flag',
        'franchise', 'select service', 'full service hotel',
    ],
    'MXU': [
        'mixed use', 'mixed-use', 'live-work', 'residential over retail',
        'ground floor retail', 'podium', 'vertical mixed',
    ],
}


def infer_property_type(text: str) -> Dict[str, Any]:
    """
    Infer property type from document text using keyword matching.

    Returns dict with:
      - property_type: code (MF, LAND, etc.) or None
      - confidence: 0.0–1.0
    """
    if not text:
        return {'property_type': None, 'confidence': 0.0}

    sample = text[:8000].lower()

    best_type = None
    best_score = 0

    for ptype, patterns in PROPERTY_TYPE_PATTERNS.items():
        score = sum(1 for p in patterns if p in sample)
        if score > best_score:
            best_score = score
            best_type = ptype

    max_patterns = len(PROPERTY_TYPE_PATTERNS.get(best_type or '', []))
    confidence = min(best_score / max(max_patterns * 0.3, 1), 1.0) if best_score > 0 else 0.0

    # Only return a result if we have reasonable confidence
    if confidence < 0.15:
        return {'property_type': None, 'confidence': 0.0}

    return {
        'property_type': best_type,
        'confidence': round(confidence, 2),
    }


# ─────────────────────────────────────────────
# Orchestrator: classify + persist
# ─────────────────────────────────────────────

def auto_classify_document(
    doc_id: int,
    file_bytes: bytes,
    filename: str,
    mime_type: str = '',
    project_id: Optional[int] = None,
) -> Dict[str, Any]:
    """
    Full auto-classification pipeline for a newly uploaded document.

    1. Extract text → store in core_doc_text
    2. Classify document type → update core_doc.doc_type
    3. Infer geo tags → insert into doc_geo_tag
    4. Infer property type (informational, stored in result)

    Returns classification results dict.
    """
    result: Dict[str, Any] = {
        'doc_type': 'general',
        'doc_type_confidence': 0.0,
        'property_type': None,
        'property_type_confidence': 0.0,
        'geo_tags': [],
        'text_extracted': False,
        'text_length': 0,
    }

    # Step 1: Extract text
    text, error = extract_text_from_bytes(file_bytes, filename, mime_type)

    if error:
        logger.warning("Text extraction failed for doc %s (%s): %s", doc_id, filename, error)
        return result

    if not text:
        logger.info("No text extracted from doc %s (%s)", doc_id, filename)
        return result

    result['text_extracted'] = True
    result['text_length'] = len(text)

    # Step 1b: Store extracted text in core_doc_text
    try:
        word_count = len(text.split())
        with connection.cursor() as cursor:
            cursor.execute("""
                INSERT INTO landscape.core_doc_text (doc_id, extracted_text, word_count, extraction_method, extracted_at, updated_at)
                VALUES (%s, %s, %s, 'auto_upload', NOW(), NOW())
                ON CONFLICT (doc_id) DO UPDATE SET
                    extracted_text = EXCLUDED.extracted_text,
                    word_count = EXCLUDED.word_count,
                    extraction_method = 'auto_upload',
                    updated_at = NOW()
            """, [doc_id, text, word_count])
    except Exception as e:
        logger.warning("Failed to store extracted text for doc %s: %s", doc_id, e)

    # Step 2: Classify document type
    classification = classify_document_type(text, filename)
    result['doc_type'] = classification['doc_type']
    result['doc_type_confidence'] = classification['confidence']
    result['matched_patterns'] = classification.get('matched_patterns', [])

    # Update core_doc.doc_type if classification is confident enough
    if classification['confidence'] >= 0.2 and classification['doc_type'] != 'general':
        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                    UPDATE landscape.core_doc
                    SET doc_type = %s, updated_at = NOW()
                    WHERE doc_id = %s
                """, [classification['doc_type'], doc_id])
        except Exception as e:
            logger.warning("Failed to update doc_type for doc %s: %s", doc_id, e)

    # Step 3: Infer geo tags
    geo_tags = infer_geo_tags(text)

    # Also add project-inferred geo tags if this doc is linked to a project
    if project_id:
        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT jurisdiction_city, jurisdiction_state
                    FROM landscape.tbl_project
                    WHERE project_id = %s
                """, [project_id])
                row = cursor.fetchone()
                if row:
                    city, state = row
                    if state and ('state', state) not in {(t['geo_level'], t['geo_value']) for t in geo_tags}:
                        geo_tags.append({'geo_level': 'state', 'geo_value': state, 'geo_source': 'inferred'})
                    if city and ('city', city) not in {(t['geo_level'], t['geo_value']) for t in geo_tags}:
                        geo_tags.append({'geo_level': 'city', 'geo_value': city, 'geo_source': 'inferred'})
        except Exception as e:
            logger.warning("Failed to get project geo for doc %s: %s", doc_id, e)

    # Persist geo tags
    for tag in geo_tags:
        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                    INSERT INTO landscape.doc_geo_tag (doc_id, geo_level, geo_value, geo_source, created_at)
                    VALUES (%s, %s, %s, %s, NOW())
                    ON CONFLICT (doc_id, geo_level, geo_value) DO NOTHING
                """, [doc_id, tag['geo_level'], tag['geo_value'], tag['geo_source']])
        except Exception as e:
            logger.warning("Failed to insert geo tag for doc %s: %s", doc_id, e)

    result['geo_tags'] = [{'level': t['geo_level'], 'value': t['geo_value']} for t in geo_tags]

    # Step 4: Infer property type
    pt = infer_property_type(text)
    result['property_type'] = pt['property_type']
    result['property_type_confidence'] = pt['confidence']

    logger.info(
        "[AUTO_CLASSIFY] doc=%s file=%s → type=%s (%.0f%%), property=%s (%.0f%%), geo=%d tags, text=%d chars",
        doc_id, filename, result['doc_type'], result['doc_type_confidence'] * 100,
        result['property_type'], result['property_type_confidence'] * 100,
        len(geo_tags), len(text)
    )

    return result
