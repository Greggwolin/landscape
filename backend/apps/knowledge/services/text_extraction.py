"""
Extract text from documents stored in UploadThing.
Supports PDF, DOCX, TXT, and common text formats.
"""
import os
import tempfile
import requests
from typing import Optional, Tuple
from urllib.parse import urlparse

# PDF extraction
try:
    import fitz  # PyMuPDF
    HAS_PYMUPDF = True
except ImportError:
    HAS_PYMUPDF = False

# PDF table extraction
try:
    import pdfplumber
    HAS_PDFPLUMBER = True
except ImportError:
    HAS_PDFPLUMBER = False

# DOCX extraction
try:
    from docx import Document as DocxDocument
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False

# Excel extraction
try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def extract_text_from_url(storage_uri: str, mime_type: str = None) -> Tuple[Optional[str], Optional[str]]:
    """
    Download document from URL and extract text.

    Args:
        storage_uri: UploadThing URL or any accessible URL
        mime_type: MIME type hint (optional, will infer from URL)

    Returns:
        Tuple of (extracted_text, error_message)
    """
    if not storage_uri:
        return None, "No storage URI provided"

    # Infer mime type from URL if not provided
    if not mime_type:
        mime_type = _infer_mime_type(storage_uri)

    try:
        # Download file to temp location
        response = requests.get(storage_uri, timeout=60)
        response.raise_for_status()

        with tempfile.NamedTemporaryFile(delete=False, suffix=_get_extension(mime_type)) as tmp:
            tmp.write(response.content)
            tmp_path = tmp.name

        try:
            # Extract based on type
            if mime_type == 'application/pdf':
                return _extract_pdf(tmp_path), None
            elif mime_type in ('application/vnd.openxmlformats-officedocument.wordprocessingml.document',):
                return _extract_docx(tmp_path), None
            elif mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
                return _extract_xlsx(tmp_path), None
            elif mime_type == 'application/vnd.ms-excel':
                return None, "Legacy .xls format not supported. Please convert to .xlsx"
            elif mime_type.startswith('text/') or mime_type in ('application/json', 'application/xml'):
                return _extract_text(tmp_path), None
            else:
                return None, f"Unsupported mime type: {mime_type}"
        finally:
            os.unlink(tmp_path)

    except requests.RequestException as e:
        return None, f"Download failed: {str(e)}"
    except Exception as e:
        return None, f"Extraction failed: {str(e)}"


def extract_text_and_page_count_from_url(
    storage_uri: str,
    mime_type: str = None
) -> Tuple[Optional[str], Optional[int], Optional[str]]:
    """
    Download document from URL and extract text with optional PDF page count.

    Returns:
        Tuple of (extracted_text, page_count, error_message)
    """
    if not storage_uri:
        return None, None, "No storage URI provided"

    if not mime_type:
        mime_type = _infer_mime_type(storage_uri)

    try:
        response = requests.get(storage_uri, timeout=60)
        response.raise_for_status()

        with tempfile.NamedTemporaryFile(delete=False, suffix=_get_extension(mime_type)) as tmp:
            tmp.write(response.content)
            tmp_path = tmp.name

        try:
            if mime_type == 'application/pdf':
                text, page_count = _extract_pdf_with_page_count(tmp_path)
                return text, page_count, None
            if mime_type in ('application/vnd.openxmlformats-officedocument.wordprocessingml.document',):
                return _extract_docx(tmp_path), None, None
            if mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
                return _extract_xlsx(tmp_path), None, None
            if mime_type == 'application/vnd.ms-excel':
                return None, None, "Legacy .xls format not supported. Please convert to .xlsx"
            if mime_type.startswith('text/') or mime_type in ('application/json', 'application/xml'):
                return _extract_text(tmp_path), None, None
            return None, None, f"Unsupported mime type: {mime_type}"
        finally:
            os.unlink(tmp_path)

    except requests.RequestException as e:
        return None, None, f"Download failed: {str(e)}"
    except Exception as e:
        return None, None, f"Extraction failed: {str(e)}"


def _extract_pdf(file_path: str) -> Optional[str]:
    """Extract text from PDF using PyMuPDF."""
    text, _ = _extract_pdf_with_page_count(file_path)
    return text


def _extract_pdf_with_page_count(file_path: str) -> Tuple[Optional[str], Optional[int]]:
    """Extract text and page count from PDF using PyMuPDF + pdfplumber for tables."""
    if not HAS_PYMUPDF:
        raise ImportError("PyMuPDF (fitz) not installed. Run: pip install PyMuPDF")

    text_parts = []
    page_count = 0

    # Step 1: Extract text with PyMuPDF (good for prose)
    with fitz.open(file_path) as doc:
        page_count = len(doc)
        for page in doc:
            text_parts.append(page.get_text())

    # Step 2: Extract tables with pdfplumber and append to page text
    if HAS_PDFPLUMBER:
        try:
            with pdfplumber.open(file_path) as pdf:
                for page_idx, page in enumerate(pdf.pages):
                    table_text = _extract_tables_from_page(page, page_idx)
                    if table_text and page_idx < len(text_parts):
                        text_parts[page_idx] = text_parts[page_idx] + "\n\n" + table_text
        except Exception:
            pass  # Fall back to PyMuPDF-only text

    text = "\n\n".join(text_parts).strip() or None
    return text, page_count


def _extract_tables_from_page(page, page_idx: int) -> Optional[str]:
    """Extract tables from a single PDF page using pdfplumber and format as text."""
    try:
        tables = page.extract_tables()
        if not tables:
            return None

        table_texts = []
        for table_idx, table in enumerate(tables):
            if not table or len(table) < 2:
                continue

            # Filter out rows that are completely empty
            filtered_rows = []
            for row in table:
                if row and any(cell and str(cell).strip() for cell in row):
                    filtered_rows.append(row)

            if len(filtered_rows) < 2:
                continue

            # Format as pipe-delimited table with markers
            lines = [f"[TABLE page={page_idx + 1} table={table_idx + 1}]"]

            for row in filtered_rows:
                cells = [str(cell).strip() if cell else '' for cell in row]
                lines.append(' | '.join(cells))

            lines.append("[/TABLE]")
            table_texts.append('\n'.join(lines))

        return '\n\n'.join(table_texts) if table_texts else None
    except Exception:
        return None


def _extract_docx(file_path: str) -> Optional[str]:
    """Extract text from DOCX."""
    if not HAS_DOCX:
        raise ImportError("python-docx not installed. Run: pip install python-docx")

    doc = DocxDocument(file_path)
    text_parts = [para.text for para in doc.paragraphs if para.text.strip()]

    return "\n\n".join(text_parts).strip() or None


def _extract_text(file_path: str) -> Optional[str]:
    """Extract from plain text files."""
    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
        return f.read().strip() or None


def _extract_xlsx(file_path: str) -> Optional[str]:
    """Extract text from Excel .xlsx files."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")

    wb = load_workbook(file_path, read_only=True, data_only=True)
    text_parts = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet_lines = [f"=== Sheet: {sheet_name} ==="]

        for row in sheet.iter_rows(values_only=True):
            # Filter out empty rows
            if any(cell is not None for cell in row):
                # Convert each cell to string, handle None
                row_text = '\t'.join(str(cell) if cell is not None else '' for cell in row)
                sheet_lines.append(row_text)

        if len(sheet_lines) > 1:  # More than just the header
            text_parts.append('\n'.join(sheet_lines))

    wb.close()
    return '\n\n'.join(text_parts).strip() or None


def _infer_mime_type(url: str) -> str:
    """Infer MIME type from URL extension."""
    path = urlparse(url).path.lower()

    if path.endswith('.pdf'):
        return 'application/pdf'
    elif path.endswith('.docx'):
        return 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
    elif path.endswith('.doc'):
        return 'application/msword'
    elif path.endswith('.xlsx'):
        return 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    elif path.endswith('.xls'):
        return 'application/vnd.ms-excel'
    elif path.endswith('.txt'):
        return 'text/plain'
    elif path.endswith('.md'):
        return 'text/markdown'
    elif path.endswith('.json'):
        return 'application/json'
    elif path.endswith('.csv'):
        return 'text/csv'
    else:
        return 'application/octet-stream'


def _get_extension(mime_type: str) -> str:
    """Get file extension for mime type."""
    mapping = {
        'application/pdf': '.pdf',
        'application/vnd.openxmlformats-officedocument.wordprocessingml.document': '.docx',
        'application/msword': '.doc',
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': '.xlsx',
        'application/vnd.ms-excel': '.xls',
        'text/plain': '.txt',
        'text/markdown': '.md',
    }
    return mapping.get(mime_type, '.bin')
