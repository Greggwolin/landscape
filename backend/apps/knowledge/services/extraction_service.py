"""
Extraction Service - AI-based data extraction from documents

This service handles:
1. Extracting structured data from documents using RAG + Claude
2. Staging extracted values for user validation
3. Applying validated values to target tables
"""

import json
import logging
import os
from typing import Dict, List, Any, Optional
from decimal import Decimal
from django.db import connection
from django.conf import settings
from anthropic import Anthropic
from .opex_utils import upsert_opex_entry

logger = logging.getLogger(__name__)


# =============================================================================
# Rent Roll Extraction Helpers
# =============================================================================

def parse_rent_roll_tags(tags_value: str) -> dict:
    """
    Parse Tags column into structured data.

    Tags may contain:
    - Property classification: "retail", "Office", "Residential Unit"
    - Program flags: "Sec. 8", "Section 8"
    - Status flags: "UD Mirage" (unlawful detainer), "under payment plan"
    - Recertification: "January-recertification"
    """
    if not tags_value:
        return {}

    tags = [t.strip() for t in str(tags_value).split(',')]
    result = {
        'all_tags': tags,
        'section_8': any('sec' in t.lower() and '8' in t for t in tags) or any('hud' in t.lower() for t in tags),
        'payment_plan': any('payment plan' in t.lower() for t in tags),
        'unlawful_detainer': any('ud' in t.lower().split() or 'unlawful' in t.lower() or 'eviction' in t.lower() for t in tags),
        'property_type': None,
        'recertification': None,
    }

    # Extract property type
    property_types = ['retail', 'office', 'residential', 'commercial', 'industrial']
    for tag in tags:
        for pt in property_types:
            if pt in tag.lower():
                result['property_type'] = tag
                break
        if result['property_type']:
            break

    # Extract recertification
    for tag in tags:
        if 'recertification' in tag.lower() or 'recert' in tag.lower():
            result['recertification'] = tag
            break

    return result


def map_occupancy_status(source_status: str, tenant_name: str = None) -> str:
    """
    Map source status values to standard occupancy_status (lowercase canonical).

    Mapping:
    - Current, Occupied, Active -> 'occupied'
    - Vacant, Vacant-Unrented -> 'vacant'
    - Notice, Notice Given -> 'notice'
    - Down, Offline -> 'down'

    If status is empty, infer from tenant name presence.
    """
    if not source_status:
        # Infer from tenant name
        if tenant_name and tenant_name.strip() and tenant_name.lower() not in ['current tenant', 'tenant', 'n/a', 'na', '--', '-']:
            return 'occupied'
        return 'vacant'

    status_lower = str(source_status).lower().strip()

    # Direct matches - return lowercase canonical values
    if status_lower in ('current', 'occupied', 'active', 'leased'):
        return 'occupied'
    elif 'vacant' in status_lower or status_lower in ('empty', 'available'):
        return 'vacant'
    elif 'notice' in status_lower or status_lower in ('ntv', 'move out'):
        return 'notice'
    elif status_lower in ('down', 'offline', 'maintenance', 'renovation'):
        return 'down'
    else:
        # Default to occupied if unclear but has a value
        return 'occupied'


def extract_tenant_name(value: str) -> Optional[str]:
    """
    Extract tenant name, never defaulting to placeholder.

    Returns None for known placeholders or empty values.
    """
    if not value:
        return None

    value = str(value).strip()

    # Reject known placeholders
    placeholders = [
        'current tenant', 'tenant', 'occupied', 'n/a', 'na', '--', '-',
        'vacant', 'unknown', 'tbd', 'to be determined', 'resident'
    ]
    if value.lower() in placeholders:
        return None

    # Reject if it looks like a status rather than a name
    if value.lower() in ['current', 'active', 'leased']:
        return None

    return value


def _get_anthropic_client() -> Anthropic:
    """Get Anthropic client with API key from .env or settings."""
    api_key = None

    # First, try to read directly from backend/.env file
    env_file = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))), '.env')
    if os.path.exists(env_file):
        with open(env_file) as f:
            for line in f:
                if line.strip().startswith('ANTHROPIC_API_KEY='):
                    api_key = line.split('=', 1)[1].strip()
                    break

    # Fallback to system env or Django settings
    if not api_key:
        api_key = os.getenv('ANTHROPIC_API_KEY') or getattr(settings, 'ANTHROPIC_API_KEY', None)

    if not api_key:
        raise ValueError("ANTHROPIC_API_KEY not found. Set it in backend/.env or environment.")

    return Anthropic(api_key=api_key)


# =============================================================================
# Conflict Detection Helpers
# =============================================================================

def normalize_value_for_comparison(value: Any) -> str:
    """
    Normalize value for comparison to detect conflicts.
    Handles numeric precision, string casing, whitespace.
    """
    if value is None:
        return ''

    # Handle JSON values (stored as jsonb)
    if isinstance(value, dict):
        value = json.dumps(value, sort_keys=True)
    elif isinstance(value, list):
        value = json.dumps(value, sort_keys=True)

    # Convert to string
    str_val = str(value).strip()

    # Empty string is treated as "no value" - return empty
    if not str_val:
        return ''

    # Normalize numbers (handle float precision issues)
    try:
        # Remove common formatting characters
        cleaned = str_val.replace(',', '').replace('$', '').replace('%', '').strip()
        num = float(cleaned)
        # Round to 2 decimal places for comparison
        return f"{num:.2f}"
    except (ValueError, TypeError):
        pass

    # Normalize strings - lowercase and strip whitespace
    return str_val.lower().strip()


def is_empty_value(value: Any) -> bool:
    """
    Check if a value is empty/null for conflict detection.
    Returns True if value is None, empty string, 'null', etc.
    """
    if value is None:
        return True
    if isinstance(value, str):
        normalized = value.strip().lower()
        return normalized in ('', 'null', 'none', 'n/a', '-')
    return False


class PlaceholderDetector:
    """Detect placeholder/dummy data patterns in existing records."""

    # Suspicious patterns
    PLACEHOLDER_DATES = [
        '2020-01-01', '2000-01-01', '1970-01-01',
        '2099-12-31', '9999-12-31',
    ]

    @classmethod
    def analyze_units(cls, existing_units: List[Dict[str, Any]], field: str) -> Dict[str, Any]:
        """
        Analyze a field across all units for placeholder patterns.

        Returns:
            {
                'is_placeholder': bool,
                'pattern': str or None,
                'unique_values': int,
                'total_units': int,
                'recommendation': str
            }
        """
        if not existing_units:
            return {'is_placeholder': False, 'pattern': None}

        values = [u.get(field) for u in existing_units if u.get(field) is not None]

        if not values:
            return {
                'is_placeholder': False,
                'pattern': 'all_null',
                'unique_values': 0,
                'total_units': len(existing_units),
                'recommendation': 'fill_blanks'
            }

        unique = set(values)
        total = len(values)

        # All identical values -- suspicious for dates/rents
        if len(unique) == 1:
            single_value = list(unique)[0]

            # Check if it's a known placeholder date
            if str(single_value) in cls.PLACEHOLDER_DATES:
                return {
                    'is_placeholder': True,
                    'pattern': 'all_identical_placeholder',
                    'value': single_value,
                    'unique_values': 1,
                    'total_units': total,
                    'recommendation': 'replace_all'
                }

            # All same date is suspicious for lease_start/lease_end
            if field in ('lease_start', 'lease_end', 'move_in_date'):
                return {
                    'is_placeholder': True,
                    'pattern': 'all_identical_date',
                    'value': single_value,
                    'unique_values': 1,
                    'total_units': total,
                    'recommendation': 'replace_all'
                }

            # All same rent is suspicious if > 10 units and mixed unit types
            if field in ('current_rent', 'market_rent') and total > 10:
                return {
                    'is_placeholder': True,
                    'pattern': 'all_identical_rent',
                    'value': single_value,
                    'unique_values': 1,
                    'total_units': total,
                    'recommendation': 'review_individually'
                }

        return {
            'is_placeholder': False,
            'pattern': None,
            'unique_values': len(unique),
            'total_units': total,
            'recommendation': None
        }

    @classmethod
    def analyze_rent_roll(cls, existing_units: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Analyze entire rent roll for placeholder patterns.

        Returns analysis for key fields.
        """
        fields_to_check = [
            'lease_start', 'lease_end', 'move_in_date',
            'current_rent', 'market_rent', 'tenant_name'
        ]

        analysis: Dict[str, Any] = {}
        placeholder_fields = []

        for field in fields_to_check:
            result = cls.analyze_units(existing_units, field)
            analysis[field] = result
            if result.get('is_placeholder'):
                placeholder_fields.append(field)

        return {
            'fields': analysis,
            'placeholder_detected': len(placeholder_fields) > 0,
            'placeholder_fields': placeholder_fields
        }


def deduplicate_extractions(extractions: Dict[str, Any]) -> Dict[str, Any]:
    """
    ISSUE 1 FIX: Deduplicate extractions within a single batch.

    The Claude extraction sometimes returns the same field multiple times
    from different source snippets. This function keeps only the best value
    for each field_key.

    For each duplicate:
    - Higher confidence wins
    - If confidence equal, prefer more specific/longer value
    - For property_class, prefer standard terms

    Args:
        extractions: Dict mapping field_key -> extraction dict
                    or field_key -> list (for array scopes)

    Returns:
        Deduplicated extractions dict
    """
    # Since extractions is already a dict keyed by field_key,
    # true duplicates can't exist at this level. However, the issue
    # might be that multiple batches run and produce overlapping fields.
    # The ON CONFLICT clause in SQL handles cross-batch deduplication.
    #
    # But if the JSON response somehow has multiple values for same field
    # (which shouldn't be possible with JSON), or if values are in arrays
    # when they should be single values, we handle that here.

    deduplicated = {}

    for field_key, extraction in extractions.items():
        if not extraction:
            continue

        # Skip non-dict extractions (already handled by JSON structure)
        if not isinstance(extraction, dict):
            deduplicated[field_key] = extraction
            continue

        # If this is a list value that should be single (based on field type),
        # take the best one
        value = extraction.get('value')
        if isinstance(value, list) and len(value) > 0:
            # Check if this looks like duplicate single values
            # (vs legitimate array like unit_types)
            if all(isinstance(v, (str, int, float, bool)) for v in value):
                # Take the most specific/longest value
                best_value = _select_best_value(value, field_key)
                extraction = dict(extraction)
                extraction['value'] = best_value

        deduplicated[field_key] = extraction

    return deduplicated


def _select_best_value(values: List[Any], field_key: str) -> Any:
    """
    Select the best value from a list of duplicates.
    Used when extraction returns multiple values for same field.
    """
    if not values:
        return None
    if len(values) == 1:
        return values[0]

    # Filter out empty values
    non_empty = [v for v in values if not is_empty_value(v)]
    if not non_empty:
        return values[0]
    if len(non_empty) == 1:
        return non_empty[0]

    # For property_class, prefer standard terms
    if field_key.lower() == 'property_class':
        standard_classes = ['multifamily', 'office', 'retail', 'industrial',
                          'land', 'hotel', 'mixed use', 'residential']
        for v in non_empty:
            if str(v).lower() in standard_classes:
                return v

    # For state fields, prefer 2-letter abbreviations
    if field_key.lower() in ('state', 'property_state'):
        for v in non_empty:
            if isinstance(v, str) and len(v) == 2 and v.isalpha():
                return v.upper()

    # Default: prefer longer/more specific values (for text)
    # For numbers, they should all be the same - take first
    if all(isinstance(v, (int, float)) for v in non_empty):
        return non_empty[0]

    # For text, prefer longer values (more complete)
    return max(non_empty, key=lambda v: len(str(v)))


def get_existing_accepted_value(
    project_id: int,
    field_key: str,
    scope_label: Optional[str] = None,
    array_index: Optional[int] = None
) -> Optional[Dict[str, Any]]:
    """
    Get existing accepted/applied extraction for this field.
    Returns None if no accepted value exists.

    Only returns values that have been accepted/applied (not pending).
    These are the values we check for conflicts against.
    """
    # Handle nullable scope_label and array_index for the query
    scope_label_val = scope_label or ''
    array_index_val = array_index if array_index is not None else 0

    query = """
        SELECT
            extraction_id,
            doc_id,
            extracted_value,
            confidence_score,
            status,
            validated_at,
            created_at
        FROM landscape.ai_extraction_staging
        WHERE project_id = %s
          AND field_key = %s
          AND COALESCE(scope_label, '') = %s
          AND COALESCE(array_index, 0) = %s
          AND status IN ('accepted', 'applied', 'validated')
        ORDER BY validated_at DESC NULLS LAST, created_at DESC
        LIMIT 1
    """

    with connection.cursor() as cursor:
        cursor.execute(query, [project_id, field_key, scope_label_val, array_index_val])
        row = cursor.fetchone()

    if not row:
        return None

    return {
        'extraction_id': row[0],
        'doc_id': row[1],
        'value': row[2],
        'confidence': float(row[3]) if row[3] else 0.0,
        'status': row[4],
        'validated_at': row[5],
        'created_at': row[6],
    }


def check_for_conflict(
    project_id: int,
    field_key: str,
    new_value: Any,
    new_doc_id: int,
    new_confidence: float,
    scope_label: Optional[str] = None,
    array_index: Optional[int] = None
) -> Optional[Dict[str, Any]]:
    """
    Check if a new extraction conflicts with an existing accepted value.

    Returns:
        None if no conflict (field is new, existing is empty, or same value)
        Dict with conflict info if non-empty values differ
    """
    existing = get_existing_accepted_value(
        project_id, field_key, scope_label, array_index
    )

    if not existing:
        # No existing accepted value - no conflict
        return None

    # ISSUE 2 FIX: Empty existing value is NOT a conflict
    # This handles the case where a field was previously extracted as empty/null
    # and now we have an actual value - that's an update, not a conflict
    if is_empty_value(existing['value']):
        return None

    # Empty new value is NOT a conflict (nothing to compare)
    if is_empty_value(new_value):
        return None

    # Normalize both values for comparison
    norm_existing = normalize_value_for_comparison(existing['value'])
    norm_new = normalize_value_for_comparison(new_value)

    if norm_existing == norm_new:
        # Same value - no conflict
        return None

    # Values differ - REAL conflict detected!
    return {
        'type': 'value_mismatch',
        'existing_extraction_id': existing['extraction_id'],
        'existing_value': existing['value'],
        'existing_doc_id': existing['doc_id'],
        'existing_confidence': existing['confidence'],
        'existing_status': existing['status'],
        'new_value': new_value,
        'new_doc_id': new_doc_id,
        'new_confidence': new_confidence,
    }


def get_doc_name(doc_id: int) -> Optional[str]:
    """Get document name by ID."""
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT doc_name FROM landscape.core_doc WHERE doc_id = %s
        """, [doc_id])
        row = cursor.fetchone()
        return row[0] if row else None


# Extraction type configurations
EXTRACTION_CONFIGS = {
    'unit_mix': {
        'target_table': 'tbl_mf_unit',
        'prompt_template': '''Extract unit mix data from the document. For each unit type, provide:
- unit_type: The unit type name (e.g., "1BR/1BA", "2BR/2BA", "Studio")
- bedrooms: Number of bedrooms (integer)
- bathrooms: Number of bathrooms (decimal, e.g., 1.0, 1.5, 2.0)
- sf_avg: Average square footage (integer)
- unit_count: Number of units of this type (integer)
- rent_market: Market rent per unit (decimal)

Return as JSON array of objects. Only include data you find in the documents.''',
        'required_fields': ['unit_type', 'bedrooms', 'sf_avg', 'unit_count'],
    },
    'rent_roll': {
        'target_table': 'tbl_mf_unit',
        'prompt_template': '''Extract ALL data from this rent roll. Do not skip any columns.

For each unit, extract these STANDARD fields:
- unit_number: The unit identifier (e.g., "101", "A-102")
- unit_type: The floorplan type (e.g., "1BR/1BA", "2x2") - derive from bedrooms/bathrooms if not explicit
- bedrooms: Number of bedrooms
- bathrooms: Number of bathrooms
- square_feet: Square footage (may be labeled "Sqft", "SF", "Size")
- current_rent: Current rent amount (may be labeled "Rent", "Contract Rent")
- lease_start: Lease start date (YYYY-MM-DD format) - may be "Lease From", "Move In"
- lease_end: Lease end date (YYYY-MM-DD format) - may be "Lease To", "Expiration"
- tenant_name: Actual tenant name from the source - NEVER substitute "Current Tenant" as a placeholder
- source_status: Original status value from source (e.g., "Current", "Vacant-Unrented") - we will map it later

CRITICAL: Also capture ALL OTHER COLUMNS as extra_data. Common examples:
- tags: Any Tags/Labels column (may contain "Sec. 8", "retail", "payment plan", etc.)
- delinquent_rent: Delinquency amounts
- rent_received: Monthly payment amounts
- additional_tags: Secondary tag columns
- Any other columns present in the source

For the Tags column, also parse into flags:
- section_8: true if tags mention "Sec. 8", "Section 8", "HUD"
- payment_plan: true if tags mention "payment plan" or "under payment plan"
- unlawful_detainer: true if tags mention "UD", "unlawful detainer", "eviction"
- recertification: Extract recertification month/date if mentioned

Return format for each unit:
{
  "unit_number": "204",
  "bedrooms": 2,
  "bathrooms": 2,
  "square_feet": 1035,
  "current_rent": 2200,
  "lease_start": "2024-08-01",
  "lease_end": "2025-07-31",
  "tenant_name": "Alana J. Sims",
  "source_status": "Current",
  "extra_data": {
    "tags": "Residential Unit, under payment plan",
    "section_8": false,
    "payment_plan": true,
    "unlawful_detainer": false,
    "delinquent_rent": 800.00,
    "rent_received": 2300,
    "additional_tags": "under payment plan"
  }
}

Return as JSON array. Include ALL units. Do not skip any columns from the source.''',
        'required_fields': ['unit_number'],
    },
    'opex': {
        'target_table': 'tbl_operating_expenses',
        'prompt_template': '''Extract operating expense data from the document. For each expense category, provide:
- category: The expense category name (e.g., "Property Tax", "Insurance", "Utilities")
- subcategory: Subcategory if applicable
- amount: Annual amount (decimal)
- per_unit: Amount per unit if shown
- year: The year this data is for (YYYY format)
- notes: Any relevant notes

Return as JSON array of objects. Only include data you find in the documents.''',
        'required_fields': ['category', 'amount'],
    },
    'market_comps': {
        'target_table': 'tbl_rental_comp',
        'prompt_template': '''Extract rental comparable data from the document. For each comparable property, provide:
- property_name: Name of the comparable property
- address: Property address
- city: City
- state: State
- year_built: Year built
- total_units: Total number of units
- avg_rent: Average rent
- avg_sf: Average square footage
- rent_psf: Rent per square foot
- occupancy: Occupancy rate (decimal, e.g., 0.95 for 95%)
- distance_miles: Distance from subject property if shown

Return as JSON array of objects. Only include data you find in the documents.''',
        'required_fields': ['property_name', 'avg_rent'],
    },
    'acquisition': {
        'target_table': 'tbl_acquisition',
        'prompt_template': '''Extract acquisition/purchase price data from the document. Provide:
- purchase_price: Total purchase price
- price_per_unit: Price per unit if shown
- price_psf: Price per square foot if shown
- cap_rate: Capitalization rate (decimal, e.g., 0.05 for 5%)
- closing_date: Expected closing date (YYYY-MM-DD format)
- deposit_amount: Earnest money/deposit amount
- due_diligence_days: Due diligence period in days
- financing_contingency: Financing contingency description

Return as JSON object with available fields. Only include data you find in the documents.''',
        'required_fields': ['purchase_price'],
    },
}


class ExtractionService:
    """Service for AI-based data extraction from documents."""

    def __init__(self, project_id: int):
        self.project_id = project_id

    def extract_from_documents(
        self,
        extraction_type: str,
        doc_ids: Optional[List[int]] = None,
        user_prompt: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Extract structured data from project documents.

        Args:
            extraction_type: Type of extraction (unit_mix, rent_roll, opex, etc.)
            doc_ids: Optional list of specific document IDs to search
            user_prompt: Optional additional user instructions

        Returns:
            Dict with extracted_values, source_docs, confidence
        """
        if extraction_type not in EXTRACTION_CONFIGS:
            return {
                'success': False,
                'error': f'Unknown extraction type: {extraction_type}',
                'valid_types': list(EXTRACTION_CONFIGS.keys())
            }

        config = EXTRACTION_CONFIGS[extraction_type]

        # Get relevant document chunks via RAG
        rag_context = self._get_rag_context(extraction_type, doc_ids)

        if not rag_context['chunks']:
            return {
                'success': False,
                'error': 'No relevant document content found for extraction',
                'extraction_type': extraction_type
            }

        # Build extraction prompt
        prompt = self._build_extraction_prompt(config, rag_context, user_prompt)

        # Call Claude for extraction
        try:
            client = _get_anthropic_client()
            response = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=4096,
                system="""You are a data extraction assistant for real estate analysis.
Extract structured data from document content with high precision.
Only extract data that is explicitly stated in the documents.
Return valid JSON that matches the requested format.
If you cannot find data for a field, omit it rather than guessing.""",
                messages=[{"role": "user", "content": prompt}]
            )

            response_text = response.content[0].text

            # Parse JSON from response
            extracted = self._parse_extraction_response(response_text)

            if extracted is None:
                return {
                    'success': False,
                    'error': 'Failed to parse extraction response',
                    'raw_response': response_text[:500]
                }

            # Stage the extracted values
            staging_ids = self._stage_extractions(
                extraction_type=extraction_type,
                config=config,
                extracted_values=extracted,
                source_chunks=rag_context['chunks']
            )

            return {
                'success': True,
                'extraction_type': extraction_type,
                'target_table': config['target_table'],
                'extracted_count': len(extracted) if isinstance(extracted, list) else 1,
                'staging_ids': staging_ids,
                'source_docs': list(set(c['doc_name'] for c in rag_context['chunks'])),
                'preview': extracted[:3] if isinstance(extracted, list) else extracted
            }

        except Exception as e:
            logger.error(f"Extraction error: {e}")
            return {
                'success': False,
                'error': str(e),
                'extraction_type': extraction_type
            }

    def _get_rag_context(
        self,
        extraction_type: str,
        doc_ids: Optional[List[int]] = None
    ) -> Dict[str, Any]:
        """Retrieve relevant document chunks for extraction."""
        # Build search query based on extraction type
        search_queries = {
            'unit_mix': 'unit mix floor plans bedrooms bathrooms square feet units',
            'rent_roll': 'rent roll tenant lease unit rent occupied vacant',
            'opex': 'operating expenses property tax insurance utilities repairs maintenance',
            'market_comps': 'comparable properties rental comps market rent occupancy',
            'acquisition': 'purchase price acquisition cap rate closing earnest money',
        }

        query = search_queries.get(extraction_type, extraction_type)

        with connection.cursor() as cursor:
            # Get embeddings for query (simplified - in production use actual embedding)
            # For now, use text search as fallback
            if doc_ids:
                doc_filter = f"AND d.doc_id IN ({','.join(str(d) for d in doc_ids)})"
            else:
                doc_filter = ""

            cursor.execute(f"""
                SELECT
                    e.embedding_id,
                    e.source_id as doc_id,
                    d.doc_name,
                    e.content_text
                FROM landscape.knowledge_embeddings e
                JOIN landscape.core_doc d ON e.source_id = d.doc_id
                WHERE d.project_id = %s
                AND e.source_type = 'document_chunk'
                {doc_filter}
                AND (
                    e.content_text ILIKE %s
                    OR e.content_text ILIKE %s
                    OR e.content_text ILIKE %s
                )
                ORDER BY e.embedding_id
                LIMIT 20
            """, [
                self.project_id,
                f'%{query.split()[0]}%',
                f'%{query.split()[1] if len(query.split()) > 1 else query.split()[0]}%',
                f'%{query.split()[2] if len(query.split()) > 2 else query.split()[0]}%',
            ])

            rows = cursor.fetchall()

            chunks = []
            for row in rows:
                chunks.append({
                    'embedding_id': row[0],
                    'doc_id': row[1],
                    'doc_name': row[2],
                    'text': row[3],
                })

            return {'chunks': chunks, 'query': query}

    def _build_extraction_prompt(
        self,
        config: Dict,
        rag_context: Dict,
        user_prompt: Optional[str]
    ) -> str:
        """Build the extraction prompt with document context."""
        chunks_text = "\n\n---\n\n".join([
            f"[From: {c['doc_name']}]\n{c['text']}"
            for c in rag_context['chunks']
        ])

        prompt = f"""{config['prompt_template']}

DOCUMENT CONTENT:
{chunks_text}

"""
        if user_prompt:
            prompt += f"\nADDITIONAL INSTRUCTIONS:\n{user_prompt}\n"

        prompt += "\nRespond with ONLY valid JSON, no other text."

        return prompt

    def _parse_extraction_response(self, response_text: str) -> Optional[Any]:
        """Parse JSON from Claude's response."""
        # Try direct parse first
        try:
            return json.loads(response_text)
        except json.JSONDecodeError:
            pass

        # Try to find JSON in response
        import re

        # Look for JSON array
        array_match = re.search(r'\[[\s\S]*\]', response_text)
        if array_match:
            try:
                return json.loads(array_match.group())
            except json.JSONDecodeError:
                pass

        # Look for JSON object
        obj_match = re.search(r'\{[\s\S]*\}', response_text)
        if obj_match:
            try:
                return json.loads(obj_match.group())
            except json.JSONDecodeError:
                pass

        return None

    def _stage_extractions(
        self,
        extraction_type: str,
        config: Dict,
        extracted_values: Any,
        source_chunks: List[Dict]
    ) -> List[int]:
        """Stage extracted values for validation."""
        staging_ids = []

        # Get primary doc_id from source chunks
        doc_id = source_chunks[0]['doc_id'] if source_chunks else None

        # Build source text summary
        source_text = "; ".join([
            f"{c['doc_name']}: {c['text'][:200]}..."
            for c in source_chunks[:3]
        ])

        items = extracted_values if isinstance(extracted_values, list) else [extracted_values]

        with connection.cursor() as cursor:
            for item in items:
                # Calculate confidence based on required fields present
                required = config['required_fields']
                fields_present = sum(1 for f in required if f in item and item[f] is not None)
                confidence = fields_present / len(required) if required else 0.5

                cursor.execute("""
                    INSERT INTO landscape.ai_extraction_staging (
                        project_id,
                        doc_id,
                        target_table,
                        target_field,
                        extracted_value,
                        extraction_type,
                        source_text,
                        confidence_score,
                        status
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'pending')
                    RETURNING extraction_id
                """, [
                    self.project_id,
                    doc_id,
                    config['target_table'],
                    None,  # target_field - null for full row extractions
                    json.dumps(item),
                    extraction_type,
                    source_text[:1000],
                    confidence,
                ])

                staging_ids.append(cursor.fetchone()[0])

        return staging_ids

    def get_pending_extractions(self) -> List[Dict]:
        """Get all pending and conflict extractions for this project."""
        with connection.cursor() as cursor:
            # Get pending and conflict extractions with conflict details
            cursor.execute("""
                SELECT
                    e.extraction_id,
                    e.doc_id,
                    d.doc_name,
                    e.field_key,
                    e.target_table,
                    e.target_field,
                    e.extracted_value,
                    e.extraction_type,
                    e.source_text,
                    e.source_snippet,
                    e.confidence_score,
                    e.status,
                    e.scope,
                    e.scope_label,
                    e.created_at,
                    e.conflict_with_extraction_id,
                    -- Conflict details (from the conflicting accepted extraction)
                    c.extracted_value as conflict_existing_value,
                    c.doc_id as conflict_existing_doc_id,
                    cd.doc_name as conflict_existing_doc_name,
                    c.confidence_score as conflict_existing_confidence
                FROM landscape.ai_extraction_staging e
                LEFT JOIN landscape.core_doc d ON e.doc_id = d.doc_id
                LEFT JOIN landscape.ai_extraction_staging c ON e.conflict_with_extraction_id = c.extraction_id
                LEFT JOIN landscape.core_doc cd ON c.doc_id = cd.doc_id
                WHERE e.project_id = %s
                AND e.status IN ('pending', 'conflict')
                ORDER BY
                    CASE WHEN e.status = 'conflict' THEN 0 ELSE 1 END,
                    e.created_at DESC
            """, [self.project_id])

            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()

            results = []
            for row in rows:
                extraction = dict(zip(columns, row))

                # Build conflict object if this extraction has a conflict
                if extraction.get('conflict_with_extraction_id'):
                    extraction['conflict'] = {
                        'existing_extraction_id': extraction.pop('conflict_with_extraction_id'),
                        'existing_value': extraction.pop('conflict_existing_value'),
                        'existing_doc_id': extraction.pop('conflict_existing_doc_id'),
                        'existing_doc_name': extraction.pop('conflict_existing_doc_name'),
                        'existing_confidence': float(extraction.pop('conflict_existing_confidence')) if extraction.get('conflict_existing_confidence') else None,
                    }
                else:
                    # Remove conflict-related None values
                    extraction.pop('conflict_with_extraction_id', None)
                    extraction.pop('conflict_existing_value', None)
                    extraction.pop('conflict_existing_doc_id', None)
                    extraction.pop('conflict_existing_doc_name', None)
                    extraction.pop('conflict_existing_confidence', None)

                results.append(extraction)

            return results

    def get_extraction_summary(self) -> Dict[str, Any]:
        """Get summary counts for extractions in this project."""
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT
                    COUNT(*) as total,
                    COUNT(*) FILTER (WHERE status = 'pending') as pending,
                    COUNT(*) FILTER (WHERE status = 'conflict') as conflicts,
                    COUNT(*) FILTER (WHERE status IN ('accepted', 'validated')) as accepted,
                    COUNT(*) FILTER (WHERE status = 'applied') as applied,
                    COUNT(*) FILTER (WHERE status = 'rejected') as rejected
                FROM landscape.ai_extraction_staging
                WHERE project_id = %s
            """, [self.project_id])

            row = cursor.fetchone()
            return {
                'total': row[0],
                'pending': row[1],
                'conflicts': row[2],
                'accepted': row[3],
                'applied': row[4],
                'rejected': row[5],
                'awaiting_review': row[1] + row[2],  # pending + conflicts
            }

    def validate_extraction(
        self,
        extraction_id: int,
        action: str,
        validated_value: Optional[Dict] = None,
        validated_by: str = 'user'
    ) -> Dict[str, Any]:
        """
        Validate or reject an extraction.

        Args:
            extraction_id: The staging row ID
            action: 'approve', 'reject', or 'edit'
            validated_value: Modified value if action is 'edit'
            validated_by: User identifier
        """
        with connection.cursor() as cursor:
            if action == 'reject':
                cursor.execute("""
                    UPDATE landscape.ai_extraction_staging
                    SET status = 'rejected',
                        validated_by = %s,
                        validated_at = NOW()
                    WHERE extraction_id = %s
                    AND project_id = %s
                    RETURNING extraction_id
                """, [validated_by, extraction_id, self.project_id])

                if cursor.fetchone():
                    return {'success': True, 'status': 'rejected'}
                return {'success': False, 'error': 'Extraction not found'}

            elif action in ('approve', 'edit'):
                # Get the extraction
                cursor.execute("""
                    SELECT extracted_value, target_table, extraction_type
                    FROM landscape.ai_extraction_staging
                    WHERE extraction_id = %s AND project_id = %s
                """, [extraction_id, self.project_id])

                row = cursor.fetchone()
                if not row:
                    return {'success': False, 'error': 'Extraction not found'}

                original_value = row[0]
                target_table = row[1]
                extraction_type = row[2]

                # Use validated value if provided, otherwise original
                final_value = validated_value if validated_value else original_value

                # Update staging record
                cursor.execute("""
                    UPDATE landscape.ai_extraction_staging
                    SET status = 'validated',
                        validated_value = %s,
                        validated_by = %s,
                        validated_at = NOW()
                    WHERE extraction_id = %s
                    RETURNING extraction_id
                """, [json.dumps(final_value), validated_by, extraction_id])

                return {
                    'success': True,
                    'status': 'validated',
                    'extraction_id': extraction_id,
                    'target_table': target_table,
                    'value': final_value
                }

            return {'success': False, 'error': f'Unknown action: {action}'}

    def apply_validated_extractions(self) -> Dict[str, Any]:
        """Apply all validated extractions to their target tables."""
        results = {
            'applied': [],
            'errors': [],
            'skipped': []
        }

        with connection.cursor() as cursor:
            # Get all validated extractions
            cursor.execute("""
                SELECT
                    extraction_id,
                    target_table,
                    extraction_type,
                    COALESCE(validated_value, extracted_value) as value
                FROM landscape.ai_extraction_staging
                WHERE project_id = %s
                AND status = 'validated'
                ORDER BY created_at
            """, [self.project_id])

            extractions = cursor.fetchall()

            for extraction_id, target_table, extraction_type, value in extractions:
                try:
                    # Apply based on extraction type
                    applied = self._apply_single_extraction(
                        cursor, extraction_id, target_table, extraction_type, value
                    )

                    if applied:
                        # Mark as applied
                        cursor.execute("""
                            UPDATE landscape.ai_extraction_staging
                            SET status = 'applied'
                            WHERE extraction_id = %s
                        """, [extraction_id])

                        results['applied'].append({
                            'extraction_id': extraction_id,
                            'target_table': target_table
                        })
                    else:
                        results['skipped'].append({
                            'extraction_id': extraction_id,
                            'reason': 'No handler for this extraction type'
                        })

                except Exception as e:
                    logger.error(f"Error applying extraction {extraction_id}: {e}")
                    results['errors'].append({
                        'extraction_id': extraction_id,
                        'error': str(e)
                    })

        return results

    def _apply_single_extraction(
        self,
        cursor,
        extraction_id: int,
        target_table: str,
        extraction_type: str,
        value: Dict
    ) -> bool:
        """Apply a single extraction to its target table."""
        if extraction_type == 'unit_mix':
            return self._apply_unit_mix(cursor, value)
        elif extraction_type == 'rent_roll':
            return self._apply_rent_roll(cursor, value)
        elif extraction_type == 'opex':
            return self._apply_opex(cursor, value)
        elif extraction_type == 'market_comps':
            return self._apply_market_comp(cursor, value)
        elif extraction_type == 'acquisition':
            return self._apply_acquisition(cursor, value)

        return False

    def _apply_unit_mix(self, cursor, value: Dict) -> bool:
        """Insert unit mix data into tbl_mf_unit."""
        cursor.execute("""
            INSERT INTO landscape.tbl_mf_unit (
                project_id,
                unit_type,
                bedrooms,
                bathrooms,
                sf_avg,
                unit_count,
                rent_market
            ) VALUES (%s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (project_id, unit_type) DO UPDATE SET
                bedrooms = EXCLUDED.bedrooms,
                bathrooms = EXCLUDED.bathrooms,
                sf_avg = EXCLUDED.sf_avg,
                unit_count = EXCLUDED.unit_count,
                rent_market = EXCLUDED.rent_market
        """, [
            self.project_id,
            value.get('unit_type'),
            value.get('bedrooms'),
            value.get('bathrooms'),
            value.get('sf_avg'),
            value.get('unit_count'),
            value.get('rent_market'),
        ])
        return True

    def _apply_rent_roll(self, cursor, value: Dict) -> bool:
        """Insert rent roll data."""
        # Similar implementation for rent roll
        return True

    def _apply_opex(self, cursor, value: Dict) -> bool:
        """Insert operating expense data."""
        category_label = (
            value.get('category')
            or value.get('expense_category')
            or value.get('label')
            or value.get('subcategory')
        )
        amount = value.get('amount') or value.get('annual_amount') or value.get('total')

        if not category_label or amount is None:
            logger.warning("Skipping opex extraction: missing category or amount (%s)", value)
            return False

        selector = {}
        if value.get('year'):
            selector['statement_discriminator'] = str(value['year'])

        result = upsert_opex_entry(connection, self.project_id, category_label, amount, selector)
        if result.get('success'):
            return True

        logger.warning("Failed to upsert opex: %s", result.get('error'))
        return False

    def _apply_market_comp(self, cursor, value: Dict) -> bool:
        """Insert rental comp data."""
        cursor.execute("""
            INSERT INTO landscape.tbl_rental_comp (
                project_id,
                property_name,
                address,
                city,
                state,
                year_built,
                total_units,
                avg_rent,
                avg_sf,
                rent_psf,
                occupancy,
                distance_miles
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, [
            self.project_id,
            value.get('property_name'),
            value.get('address'),
            value.get('city'),
            value.get('state'),
            value.get('year_built'),
            value.get('total_units'),
            value.get('avg_rent'),
            value.get('avg_sf'),
            value.get('rent_psf'),
            value.get('occupancy'),
            value.get('distance_miles'),
        ])
        return True

    def _apply_acquisition(self, cursor, value: Dict) -> bool:
        """Update acquisition data."""
        # Check if project has acquisition record, update or insert
        cursor.execute("""
            UPDATE landscape.tbl_project
            SET purchase_price = %s
            WHERE project_id = %s
        """, [value.get('purchase_price'), self.project_id])
        return True


def extract_data(project_id: int, extraction_type: str, **kwargs) -> Dict[str, Any]:
    """Convenience function for extraction."""
    service = ExtractionService(project_id)
    return service.extract_from_documents(extraction_type, **kwargs)


def get_pending(project_id: int) -> List[Dict]:
    """Get pending extractions for a project."""
    service = ExtractionService(project_id)
    return service.get_pending_extractions()


def validate(project_id: int, extraction_id: int, action: str, **kwargs) -> Dict[str, Any]:
    """Validate an extraction."""
    service = ExtractionService(project_id)
    return service.validate_extraction(extraction_id, action, **kwargs)


def apply_all(project_id: int) -> Dict[str, Any]:
    """Apply all validated extractions."""
    service = ExtractionService(project_id)
    return service.apply_validated_extractions()


# =============================================================================
# Registry-Based AI Extraction Service (v3)
# =============================================================================

from .field_registry import get_registry, FieldMapping
from .document_classifier import (
    DocumentClassifier,
    DocumentType,
    ClassificationResult,
    classify_document,
    DOCTYPE_TO_EVIDENCE
)
from .subtype_classifier import DocumentSubtypeClassifier


class RegistryBasedExtractor:
    """
    AI extraction service using Field Registry v3 as authoritative contract.

    This extractor:
    1. Auto-classifies documents to determine extractable fields
    2. Uses the Field Registry to get target fields per document type
    3. Builds dynamic extraction prompts from registry metadata
    4. Stages extracted values with field_key linking to registry
    5. Auto-assigns subtype tags when classifier confidence >= 0.6
    """

    def __init__(self, project_id: int, property_type: str = 'multifamily'):
        self.project_id = project_id
        self.property_type = property_type
        self.registry = get_registry()
        self.classifier = DocumentClassifier(project_id)
        self.subtype_classifier = DocumentSubtypeClassifier()

    def extract_from_document(
        self,
        doc_id: int,
        field_keys: Optional[List[str]] = None,
        high_extractability_only: bool = False
    ) -> Dict[str, Any]:
        """
        Extract data from a single document using registry-defined fields.

        Args:
            doc_id: Document ID to extract from
            field_keys: Optional list of specific field keys to extract
            high_extractability_only: If True, only extract fields with high extractability

        Returns:
            Dict with extracted values, staging IDs, and metadata
        """
        # Step 1: Classify the document
        classification = self.classifier.classify_document(doc_id)

        if classification.doc_type == DocumentType.UNKNOWN:
            return {
                'success': False,
                'error': 'Could not classify document type',
                'doc_id': doc_id,
                'classification': {
                    'type': 'unknown',
                    'confidence': classification.confidence
                }
            }

        # Step 2: Get extractable fields from registry
        if field_keys:
            # Use specific fields requested
            fields = self.registry.get_fields_by_keys(field_keys, self.property_type)
        else:
            # Get fields for this document type
            fields = self.registry.get_extraction_fields_for_doc_type(
                classification.evidence_type,
                self.property_type,
                high_extractability_only
            )

        if not fields:
            return {
                'success': False,
                'error': f'No extractable fields for document type: {classification.evidence_type}',
                'doc_id': doc_id,
                'classification': {
                    'type': classification.evidence_type,
                    'confidence': classification.confidence
                }
            }

        # Step 3: Get document content
        doc_content = self._get_document_content(doc_id)
        if not doc_content:
            return {
                'success': False,
                'error': 'No content found for document',
                'doc_id': doc_id
            }

        # Step 3b: Subtype classification — auto-assign tag if confident
        subtype_result = self.subtype_classifier.classify(
            content=doc_content['text'][:50000],
            property_type=self.property_type,
        )
        if subtype_result.subtype_code and subtype_result.confidence >= 0.6:
            self._assign_subtype_tag(doc_id, subtype_result.subtype_code, workspace_id=1)
            logger.info(
                f"Subtype detected for doc {doc_id}: {subtype_result.subtype_name} "
                f"(confidence={subtype_result.confidence:.2f}, patterns={subtype_result.matched_patterns})"
            )

        # Step 4: Build extraction prompt from registry
        prompt = self._build_registry_extraction_prompt(fields, doc_content['text'])

        # Step 5: Call Claude for extraction
        try:
            client = _get_anthropic_client()
            response = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=4096,
                system="""You are a precise data extraction assistant for real estate documents.
Extract values ONLY for the fields specified in the schema.
Return a JSON object with field_key as keys and extracted values.
For each value, include:
- "value": The extracted value (properly typed: string, number, or boolean)
- "confidence": Your confidence level (high, medium, low)
- "source_quote": Brief quote from document supporting the extraction

If a field cannot be found in the document, omit it from the response.
Never guess or make up values - only extract what is explicitly stated.""",
                messages=[{"role": "user", "content": prompt}]
            )

            response_text = response.content[0].text
            extracted = self._parse_extraction_response(response_text)

            if extracted is None:
                return {
                    'success': False,
                    'error': 'Failed to parse extraction response',
                    'raw_response': response_text[:500],
                    'doc_id': doc_id
                }

            # Step 6: Stage the extractions
            staging_results = self._stage_registry_extractions(
                doc_id=doc_id,
                doc_name=doc_content['doc_name'],
                extractions=extracted,
                fields=fields,
                classification=classification
            )

            return {
                'success': True,
                'doc_id': doc_id,
                'doc_name': doc_content['doc_name'],
                'classification': {
                    'type': classification.evidence_type,
                    'confidence': classification.confidence,
                    'matched_patterns': classification.matched_patterns[:5]
                },
                'extracted_count': len(staging_results['staged']),
                'staging_ids': [s['extraction_id'] for s in staging_results['staged']],
                'extractions': staging_results['staged'],
                'skipped': staging_results['skipped']
            }

        except Exception as e:
            logger.error(f"Registry extraction error for doc {doc_id}: {e}")
            return {
                'success': False,
                'error': str(e),
                'doc_id': doc_id
            }

    def extract_from_project(
        self,
        doc_ids: Optional[List[int]] = None,
        high_extractability_only: bool = True
    ) -> Dict[str, Any]:
        """
        Extract from all (or specified) project documents.

        Args:
            doc_ids: Optional list of specific doc IDs; if None, process all project docs
            high_extractability_only: Focus on high-confidence fields only

        Returns:
            Dict with results per document
        """
        if doc_ids is None:
            doc_ids = self._get_project_document_ids()

        results = {
            'project_id': self.project_id,
            'documents_processed': 0,
            'total_extractions': 0,
            'documents': []
        }

        for doc_id in doc_ids:
            doc_result = self.extract_from_document(
                doc_id,
                high_extractability_only=high_extractability_only
            )
            results['documents'].append(doc_result)
            results['documents_processed'] += 1
            if doc_result.get('success'):
                results['total_extractions'] += doc_result.get('extracted_count', 0)

        return results

    def _assign_subtype_tag(self, doc_id: int, subtype_code: str, workspace_id: int = 1):
        """Auto-assign the tag corresponding to a detected subtype."""
        try:
            with connection.cursor() as cursor:
                # Find the tag linked to this subtype
                cursor.execute("""
                    SELECT tag_id FROM landscape.dms_doc_tags
                    WHERE subtype_code = %s AND workspace_id = %s
                    LIMIT 1
                """, [subtype_code, workspace_id])
                row = cursor.fetchone()

                if row:
                    tag_id = row[0]
                    # Assign to document (ignore if already assigned)
                    cursor.execute("""
                        INSERT INTO landscape.dms_doc_tag_assignments (doc_id, tag_id, assigned_at)
                        VALUES (%s, %s, NOW())
                        ON CONFLICT (doc_id, tag_id) DO NOTHING
                    """, [doc_id, tag_id])

                    # Increment usage count
                    cursor.execute("""
                        UPDATE landscape.dms_doc_tags
                        SET usage_count = (
                            SELECT COUNT(*) FROM landscape.dms_doc_tag_assignments WHERE tag_id = %s
                        )
                        WHERE tag_id = %s
                    """, [tag_id, tag_id])

                    logger.info(f"Auto-assigned subtype tag '{subtype_code}' to doc {doc_id}")
        except Exception as e:
            logger.warning(f"Failed to assign subtype tag: {e}")
            # Non-fatal — don't break extraction over a tag assignment failure

    def classify_document(self, doc_id: int) -> ClassificationResult:
        """Classify a document and return the result."""
        return self.classifier.classify_document(doc_id)

    def get_extractable_fields(self, doc_type: str) -> List[Dict[str, Any]]:
        """Get extractable fields for a document type as JSON-serializable dicts."""
        fields = self.registry.get_extraction_fields_for_doc_type(
            doc_type, self.property_type
        )
        return [
            {
                'field_key': f.field_key,
                'label': f.label,
                'field_type': f.field_type,
                'required': f.required,
                'extractability': f.extractability,
                'scope': f.scope,
                'db_target': f.db_target
            }
            for f in fields
        ]

    def _get_document_content(self, doc_id: int) -> Optional[Dict]:
        """Get document info and content text."""
        with connection.cursor() as cursor:
            # Get doc info
            cursor.execute("""
                SELECT doc_id, doc_name, project_id
                FROM landscape.core_doc
                WHERE doc_id = %s
            """, [doc_id])
            row = cursor.fetchone()
            if not row:
                return None

            doc_info = {
                'doc_id': row[0],
                'doc_name': row[1],
                'project_id': row[2],
            }

            # Get content from embeddings (ordered chunks)
            cursor.execute("""
                SELECT content_text
                FROM landscape.knowledge_embeddings
                WHERE source_id = %s AND source_type = 'document_chunk'
                ORDER BY embedding_id
                LIMIT 100
            """, [doc_id])

            chunks = cursor.fetchall()
            content = "\n".join(chunk[0] for chunk in chunks if chunk[0])

            # Fallback to extracted_text if no embeddings
            if not content:
                cursor.execute("""
                    SELECT COALESCE(extracted_text, '')
                    FROM landscape.core_doc
                    WHERE doc_id = %s
                """, [doc_id])
                row = cursor.fetchone()
                content = row[0] if row and row[0] else ''

            doc_info['text'] = content
            return doc_info

    def _get_project_document_ids(self) -> List[int]:
        """Get all document IDs for the project."""
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT doc_id
                FROM landscape.core_doc
                WHERE project_id = %s
                ORDER BY doc_id
            """, [self.project_id])
            return [row[0] for row in cursor.fetchall()]

    def _build_registry_extraction_prompt(
        self,
        fields: List[FieldMapping],
        document_text: str
    ) -> str:
        """Build extraction prompt from registry field definitions."""
        # Group fields by scope
        array_scopes = {'unit_type', 'unit', 'sales_comp', 'rent_comp', 'opex', 'income'}
        scope_fields = {}
        field_lines = []

        for f in fields:
            type_hint = f.field_type
            if type_hint == 'currency':
                type_hint = 'number (dollars, no currency symbol)'
            elif type_hint == 'percentage':
                type_hint = 'decimal (e.g., 0.05 for 5%)'

            req = "[REQUIRED]" if f.required else "[optional]"
            scope_tag = f" [SCOPE: {f.scope}]" if f.scope in array_scopes else ""
            field_line = f"- {f.field_key}: {f.label} ({type_hint}) {req}{scope_tag}"
            field_lines.append(field_line)

            if f.scope not in scope_fields:
                scope_fields[f.scope] = []
            scope_fields[f.scope].append(f.field_key)

        fields_schema = "\n".join(field_lines)

        # Build array extraction instructions
        array_instructions = ""

        if 'unit_type' in scope_fields:
            array_instructions += f"""
## UNIT TYPE FIELDS (scope=unit_type)
Fields: {', '.join(scope_fields['unit_type'])}

Extract as array - one entry per unit type found in unit mix or rent roll summaries:
{{
  "market_rent": {{
    "value": [
      {{"unit_type_name": "1BR/1BA", "value": 2400}},
      {{"unit_type_name": "2BR/2BA", "value": 3150}}
    ],
    "confidence": "high",
    "source_quote": "Unit Mix Summary"
  }}
}}
"""

        if 'unit' in scope_fields:
            array_instructions += f"""
## INDIVIDUAL UNIT FIELDS (scope=unit) - RENT ROLL
Fields: {', '.join(scope_fields['unit'][:8])}...

Extract ALL units from rent roll as array:
{{
  "unit_number": {{
    "value": [
      {{"unit_number": "101", "unit_type": "1BR/1BA", "bedrooms": 1, "current_rent": 1595, "market_rent": 1650, "occupancy_status": "Occupied"}},
      {{"unit_number": "102", "unit_type": "2BR/2BA", "bedrooms": 2, "current_rent": 2150, "market_rent": 2250, "occupancy_status": "Vacant"}}
    ],
    "confidence": "high",
    "source_quote": "Rent Roll, pages 30-34"
  }}
}}
IMPORTANT: Extract ALL units in the rent roll, not just a sample.
"""

        if 'sales_comp' in scope_fields:
            array_instructions += f"""
## SALES COMPARABLES (scope=sales_comp)
Fields: {', '.join(scope_fields['sales_comp'][:5])}...

Extract ALL sales comps as array:
{{
  "sales_comp_name": {{
    "value": [
      {{"property_name": "Park Place", "address": "123 Main St", "sale_price": 45000000, "units": 120, "cap_rate": 0.045}},
      {{"property_name": "Harbor View", "address": "456 Oak Ave", "sale_price": 38000000, "units": 98, "cap_rate": 0.048}}
    ],
    "confidence": "high",
    "source_quote": "Sales Comparables, page 22"
  }}
}}
"""

        if 'rent_comp' in scope_fields:
            array_instructions += f"""
## RENT COMPARABLES (scope=rent_comp)
Fields: {', '.join(scope_fields['rent_comp'][:5])}...

Extract ALL rent comps as array:
{{
  "rent_comp_name": {{
    "value": [
      {{"property_name": "Airo at South Bay", "year_built": 2021, "total_units": 230, "asking_rent": 2650}},
      {{"property_name": "The Mark", "year_built": 2019, "total_units": 180, "asking_rent": 2450}}
    ],
    "confidence": "high",
    "source_quote": "Rent Comparables Survey"
  }}
}}
"""

        if 'opex' in scope_fields:
            array_instructions += f"""
## OPERATING EXPENSES (scope=opex)
Fields: {', '.join(scope_fields['opex'][:8])}...

Extract each expense category found:
{{
  "opex_real_estate_taxes": {{"value": 285000, "confidence": "high", "source_quote": "Real Estate Taxes: $285,000"}},
  "opex_property_insurance": {{"value": 42000, "confidence": "high", "source_quote": "Insurance: $42,000"}}
}}
"""

        # Truncate document text if too long
        max_text_length = 60000
        if len(document_text) > max_text_length:
            document_text = document_text[:max_text_length] + "\n... [truncated]"

        prompt = f"""Extract the following fields from this document.

EXTRACTION SCHEMA:
{fields_schema}
{array_instructions}
## RESPONSE FORMAT

Return a JSON object where each key is the field_key and the value is an object with:
- "value": The extracted value (single value for project-scoped, array for array-scoped)
- "confidence": "high", "medium", or "low"
- "source_quote": Brief supporting quote (max 100 chars)

Example for project-scoped fields:
{{
  "property_name": {{"value": "Sunset Apartments", "confidence": "high", "source_quote": "Property: Sunset Apartments"}},
  "total_units": {{"value": 120, "confidence": "high", "source_quote": "120 total units"}}
}}

DOCUMENT CONTENT:
{document_text}

Extract all available fields. Return ONLY valid JSON, no other text."""

        return prompt

    def _parse_extraction_response(self, response_text: str) -> Optional[Dict]:
        """Parse JSON from Claude's response."""
        # Try direct parse
        try:
            return json.loads(response_text)
        except json.JSONDecodeError:
            pass

        # Try to find JSON object in response
        import re
        obj_match = re.search(r'\{[\s\S]*\}', response_text)
        if obj_match:
            try:
                return json.loads(obj_match.group())
            except json.JSONDecodeError:
                pass

        return None

    def _find_unit_type_id(self, unit_type_name: str) -> Optional[int]:
        """
        Match unit type name to existing tbl_multifamily_unit_type row.

        Uses scored matching to prefer exact matches over partial matches.
        "1 Bed 1 Bath XL Patio" should match "1BR/1BA XL Patio" not "1BR/1BA".
        """
        import re

        if not unit_type_name:
            return None

        # Normalize for matching - converts various formats to standard form
        # "1 Bed 1 Bath" and "1BR/1BA" both become "1br1ba"
        def normalize(name: str) -> str:
            n = name.lower()
            # Replace "bed" with "br"
            n = re.sub(r'\s*bed(room)?s?\s*', 'br', n)
            # Replace "bath" with "ba"
            n = re.sub(r'\s*bath(room)?s?\s*', 'ba', n)
            # Remove spaces, slashes, dashes
            n = n.replace(' ', '').replace('/', '').replace('-', '')
            return n

        normalized = normalize(unit_type_name)

        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT unit_type_id, unit_type_code, unit_type_name
                FROM landscape.tbl_multifamily_unit_type
                WHERE project_id = %s
            """, [self.project_id])
            rows = cursor.fetchall()

        # Score all candidates, pick the best match
        best_match = None
        best_score = 0

        for row in rows:
            unit_type_id = row[0]
            unit_type_code = row[1] or ''
            existing_name = row[2] or ''

            norm_code = normalize(unit_type_code)
            norm_name = normalize(existing_name)

            score = 0

            # Exact match on code or name = highest score
            if normalized == norm_code:
                score = 100
            elif normalized == norm_name:
                score = 100
            else:
                # Check for variant suffix matching
                has_xlpatio = 'xlpatio' in normalized
                has_tower = 'tower' in normalized
                code_has_xlpatio = 'xlpatio' in norm_code
                code_has_tower = 'tower' in norm_code

                # Strip suffixes to get base
                base_normalized = re.sub(r'(xlpatio|tower)$', '', normalized)
                base_code = re.sub(r'(xlpatio|tower)$', '', norm_code)

                if base_normalized == base_code:
                    # Base matches - now check suffixes
                    if has_xlpatio and code_has_xlpatio:
                        score = 90  # Exact variant match
                    elif has_tower and code_has_tower:
                        score = 90  # Exact variant match
                    elif not has_xlpatio and not has_tower and not code_has_xlpatio and not code_has_tower:
                        score = 90  # Both are base units
                    elif has_xlpatio and not code_has_xlpatio and not code_has_tower:
                        score = 50  # Input has variant, DB has base only
                    elif has_tower and not code_has_xlpatio and not code_has_tower:
                        score = 50  # Input has variant, DB has base only
                    elif not has_xlpatio and not has_tower:
                        score = 40  # Input is base, DB has variant

            if score > best_score:
                best_score = score
                best_match = unit_type_id

        # Only return if we have a reasonable match (score >= 50)
        if best_score >= 50:
            return best_match

        return None  # No match - will need manual mapping

    def _generate_unit_type_code(self, name: str) -> str:
        """
        Generate short code from unit type name.

        Examples:
        - "1 Bed 1 Bath" -> "1B1B"
        - "1 Bed 1 Bath XL Patio" -> "1B1B-XLP"
        - "2 Bed 2 Bath Tower" -> "2B2B-TWR"
        - "Studio" -> "STU"
        """
        name_lower = name.lower()

        # Extract bed/bath counts
        beds = '0'
        baths = '0'

        # Check for bed count
        if 'studio' in name_lower:
            beds = 'STU'
        elif '1 bed' in name_lower or '1bed' in name_lower:
            beds = '1'
        elif '2 bed' in name_lower or '2bed' in name_lower:
            beds = '2'
        elif '3 bed' in name_lower or '3bed' in name_lower:
            beds = '3'
        elif '4 bed' in name_lower or '4bed' in name_lower:
            beds = '4'

        # Check for bath count
        if '1 bath' in name_lower or '1bath' in name_lower:
            baths = '1'
        elif '2 bath' in name_lower or '2bath' in name_lower:
            baths = '2'
        elif '1.5 bath' in name_lower:
            baths = '1H'
        elif '2.5 bath' in name_lower:
            baths = '2H'

        # Build base code
        if beds == 'STU':
            base = 'STU'
        else:
            base = f"{beds}BR{baths}BA"

        # Add suffix for variants
        suffix = ''
        if 'xl' in name_lower and 'patio' in name_lower:
            suffix = '-XLP'
        elif 'patio' in name_lower:
            suffix = '-PAT'
        elif 'tower' in name_lower:
            suffix = '-TWR'
        elif 'penthouse' in name_lower or ' ph' in name_lower:
            suffix = '-PH'
        elif 'garden' in name_lower:
            suffix = '-GDN'
        elif 'loft' in name_lower:
            suffix = '-LFT'

        return base + suffix

    def _create_unit_type(self, unit_type_name: str) -> int:
        """
        Create new unit type row and return its ID.

        Args:
            unit_type_name: Full name like "1 Bed 1 Bath XL Patio"

        Returns:
            The new unit_type_id
        """
        code = self._generate_unit_type_code(unit_type_name)

        with connection.cursor() as cursor:
            cursor.execute("""
                INSERT INTO landscape.tbl_multifamily_unit_type
                (project_id, unit_type_code, unit_type_name, created_at)
                VALUES (%s, %s, %s, NOW())
                RETURNING unit_type_id
            """, [self.project_id, code, unit_type_name])
            new_id = cursor.fetchone()[0]

        logger.info(f"Created new unit type '{unit_type_name}' (code={code}, id={new_id}) for project {self.project_id}")
        return new_id

    def _find_or_create_unit_type_id(self, unit_type_name: str) -> int:
        """
        Match unit type name to existing row, or create new one.

        Args:
            unit_type_name: Name like "1 Bed 1 Bath XL Patio"

        Returns:
            The unit_type_id (existing or newly created)
        """
        # First try to find existing
        existing_id = self._find_unit_type_id(unit_type_name)
        if existing_id:
            return existing_id

        # No match - create new unit type
        return self._create_unit_type(unit_type_name)

    def _stage_registry_extractions(
        self,
        doc_id: int,
        doc_name: str,
        extractions: Dict[str, Any],
        fields: List[FieldMapping],
        classification: ClassificationResult
    ) -> Dict[str, Any]:
        """Stage extracted values with registry field_key linkage."""
        results = {'staged': [], 'skipped': []}

        # Build field lookup
        field_map = {f.field_key: f for f in fields}

        # Array scopes that need special handling
        array_scopes = {'unit_type', 'unit', 'sales_comp', 'rent_comp'}

        with connection.cursor() as cursor:
            for field_key, extraction in extractions.items():
                if field_key not in field_map:
                    results['skipped'].append({
                        'field_key': field_key,
                        'reason': 'Not in registry for this document type'
                    })
                    continue

                field = field_map[field_key]

                # Extract value details
                if isinstance(extraction, dict):
                    value = extraction.get('value')
                    confidence_str = extraction.get('confidence', 'medium')
                    source_quote = extraction.get('source_quote', '')
                else:
                    value = extraction
                    confidence_str = 'medium'
                    source_quote = ''

                # Map confidence string to score
                confidence_map = {'high': 0.9, 'medium': 0.7, 'low': 0.5}
                confidence_score = confidence_map.get(confidence_str, 0.7)

                # Handle array-scoped fields
                if field.scope in array_scopes and isinstance(value, list):
                    for array_index, item in enumerate(value):
                        if not isinstance(item, dict):
                            continue

                        # Get scope label and scope_id based on scope type
                        scope_label = None
                        scope_id = None

                        if field.scope == 'unit_type':
                            scope_label = item.get('unit_type_name', '')
                            item_value = item.get('value')
                            if scope_label:
                                scope_id = self._find_or_create_unit_type_id(scope_label)

                        elif field.scope == 'unit':
                            scope_label = item.get('unit_number', '')
                            # For units, we stage all fields from the item
                            item_value = item  # Store entire item as JSON

                        elif field.scope == 'sales_comp':
                            scope_label = item.get('property_name', f'Comp {array_index + 1}')
                            item_value = item

                        elif field.scope == 'rent_comp':
                            scope_label = item.get('property_name', f'Rent Comp {array_index + 1}')
                            item_value = item

                        else:
                            item_value = item.get('value', item)

                        if item_value is None:
                            continue

                        cursor.execute("""
                            INSERT INTO landscape.ai_extraction_staging (
                                project_id,
                                doc_id,
                                field_key,
                                target_table,
                                target_field,
                                extracted_value,
                                extraction_type,
                                source_text,
                                confidence_score,
                                scope,
                                scope_id,
                                scope_label,
                                array_index,
                                status
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'pending')
                            RETURNING extraction_id
                        """, [
                            self.project_id,
                            doc_id,
                            field_key,
                            field.table_name,
                            field.column_name,
                            json.dumps(item_value),
                            classification.evidence_type,
                            source_quote[:500] if source_quote else f"Extracted from {doc_name}",
                            confidence_score,
                            field.scope,
                            scope_id,
                            scope_label[:100] if scope_label else None,
                            array_index
                        ])

                        extraction_id = cursor.fetchone()[0]

                        results['staged'].append({
                            'extraction_id': extraction_id,
                            'field_key': field_key,
                            'label': field.label,
                            'value': item_value,
                            'scope_label': scope_label,
                            'scope_id': scope_id,
                            'array_index': array_index,
                            'confidence': confidence_str,
                            'db_target': field.db_target,
                            'scope': field.scope
                        })
                else:
                    # Standard single-value insertion (project, mf_property, opex, assumption, etc.)
                    # Get selector info for row-based writes
                    selector_json = field.selector_json
                    scope_label = None
                    if selector_json and 'category' in selector_json:
                        scope_label = selector_json.get('category')
                    elif selector_json and 'key' in selector_json:
                        scope_label = selector_json.get('key')

                    cursor.execute("""
                        INSERT INTO landscape.ai_extraction_staging (
                            project_id,
                            doc_id,
                            field_key,
                            target_table,
                            target_field,
                            extracted_value,
                            extraction_type,
                            source_text,
                            confidence_score,
                            scope,
                            scope_label,
                            status
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'pending')
                        RETURNING extraction_id
                    """, [
                        self.project_id,
                        doc_id,
                        field_key,
                        field.table_name,
                        field.column_name,
                        json.dumps(value),
                        classification.evidence_type,
                        source_quote[:500] if source_quote else f"Extracted from {doc_name}",
                        confidence_score,
                        field.scope,
                        scope_label
                    ])

                    extraction_id = cursor.fetchone()[0]

                    results['staged'].append({
                        'extraction_id': extraction_id,
                        'field_key': field_key,
                        'label': field.label,
                        'value': value,
                        'confidence': confidence_str,
                        'db_target': field.db_target,
                        'scope': field.scope
                    })

        return results


# Convenience functions for registry-based extraction
def extract_document_v3(
    project_id: int,
    doc_id: int,
    property_type: str = 'multifamily',
    field_keys: Optional[List[str]] = None
) -> Dict[str, Any]:
    """Extract from a document using registry-based extractor."""
    extractor = RegistryBasedExtractor(project_id, property_type)
    return extractor.extract_from_document(doc_id, field_keys)


def extract_project_v3(
    project_id: int,
    property_type: str = 'multifamily',
    doc_ids: Optional[List[int]] = None
) -> Dict[str, Any]:
    """Extract from all project documents using registry-based extractor."""
    extractor = RegistryBasedExtractor(project_id, property_type)
    return extractor.extract_from_project(doc_ids)


def classify_and_preview(project_id: int, doc_id: int) -> Dict[str, Any]:
    """Classify a document and preview extractable fields."""
    extractor = RegistryBasedExtractor(project_id)

    # Classify
    classification = extractor.classify_document(doc_id)

    # Get extractable fields
    if classification.doc_type != DocumentType.UNKNOWN:
        fields = extractor.get_extractable_fields(classification.evidence_type)
    else:
        fields = []

    return {
        'doc_id': doc_id,
        'classification': {
            'type': classification.evidence_type,
            'confidence': classification.confidence,
            'matched_patterns': classification.matched_patterns
        },
        'extractable_fields': fields,
        'field_count': len(fields)
    }


# ============================================================
# BATCHED EXTRACTION SERVICE
# ============================================================

EXTRACTION_BATCHES = [
    {
        'name': 'core_property',
        'scopes': ['project', 'mf_property'],
        'description': 'Core property identification and physical characteristics',
    },
    {
        'name': 'financials',
        'scopes': ['opex', 'income', 'assumption'],
        'description': 'Operating expenses, other income, and underwriting assumptions',
    },
    {
        'name': 'unit_types',
        'scopes': ['unit_type'],
        'description': 'Unit type summary (bed/bath, SF, counts, rents by type)',
    },
    {
        'name': 'comparables',
        'scopes': ['sales_comp', 'rent_comp'],
        'description': 'Sales and rent comparables (array extraction)',
    },
    {
        'name': 'deal_market',
        'scopes': ['acquisition', 'market'],
        'description': 'Acquisition terms and market demographics',
    },
    {
        'name': 'rent_roll',
        'scopes': ['unit'],
        'description': 'Full unit-level rent roll (array extraction)',
    },
]


class BatchedExtractionService:
    """
    Extracts document data in batches to avoid prompt/response size limits.
    Each batch targets specific scopes and runs as a separate Claude API call.
    """

    def __init__(self, project_id: int, property_type: str = 'multifamily'):
        self.project_id = project_id
        self.property_type = property_type
        self.registry = get_registry()
        self.classifier = DocumentClassifier()
        # Import here to avoid circular imports at module load
        from .document_processor import processor as doc_processor
        self.doc_processor = doc_processor

    def extract_document_batched(
        self,
        doc_id: int,
        batches: Optional[List[str]] = None,
        user_id: Optional[int] = None,
    ) -> Dict[str, Any]:
        """
        Run batched extraction on a document.

        Args:
            doc_id: Source document
            batches: Optional list of batch names to run (default: all)
            user_id: For audit trail

        Returns:
            {
                'success': bool,
                'batches_run': int,
                'batches_succeeded': int,
                'total_staged': int,
                'results': [...],
                'errors': []
            }
        """
        results = []
        errors = []
        total_staged = 0

        # Determine which batches to run
        batches_to_run = EXTRACTION_BATCHES
        if batches:
            batches_to_run = [b for b in EXTRACTION_BATCHES if b['name'] in batches]

        # Get document content once (reuse across batches)
        doc_content = self._get_document_content(doc_id)
        if not doc_content:
            # Attempt to process the document inline to populate embeddings/text
            try:
                process_result = self.doc_processor.process_document(doc_id)
                if process_result.get('success'):
                    doc_content = self._get_document_content(doc_id)
                else:
                    logger.warning(
                        f"Inline processing failed for doc {doc_id}: {process_result.get('error')}"
                    )
            except Exception as e:
                logger.exception(f"Inline processing exception for doc {doc_id}: {e}")

        if not doc_content:
            return {
                'success': False,
                'error': f'No content found for document {doc_id}',
                'batches_run': 0,
            }

        # Get document info for staging
        doc_info = self._get_document_info(doc_id)
        if not doc_info:
            return {
                'success': False,
                'error': f'Document {doc_id} not found',
                'batches_run': 0,
            }

        logger.info(f"Starting batched extraction for doc {doc_id} ({len(batches_to_run)} batches)")

        for batch in batches_to_run:
            batch_name = batch['name']
            scopes = batch['scopes']

            try:
                # Collect fields for this batch's scopes
                fields = []
                for scope in scopes:
                    scope_fields = self.registry.get_fields_by_scope(scope, self.property_type)
                    fields.extend(scope_fields)

                if not fields:
                    results.append({
                        'batch': batch_name,
                        'success': True,
                        'staged': 0,
                        'fields': 0,
                        'note': 'No fields found for scopes',
                    })
                    continue

                logger.info(f"Batch '{batch_name}': {len(fields)} fields from scopes {scopes}")

                # Run extraction for this batch
                batch_result = self._extract_batch(
                    doc_id=doc_id,
                    doc_info=doc_info,
                    doc_content=doc_content,
                    fields=fields,
                    batch_name=batch_name,
                    batch_description=batch['description'],
                    scopes=scopes,
                )

                results.append(batch_result)
                if batch_result.get('success'):
                    total_staged += batch_result.get('staged', 0)
                else:
                    errors.append(f"{batch_name}: {batch_result.get('error', 'Unknown error')}")

            except Exception as e:
                logger.exception(f"Batch {batch_name} failed: {e}")
                errors.append(f"{batch_name}: {str(e)}")
                results.append({
                    'batch': batch_name,
                    'success': False,
                    'error': str(e),
                })

        return {
            'success': len(errors) == 0,
            'doc_id': doc_id,
            'project_id': self.project_id,
            'batches_run': len(results),
            'batches_succeeded': sum(1 for r in results if r.get('success')),
            'total_staged': total_staged,
            'results': results,
            'errors': errors,
        }

    def _extract_batch(
        self,
        doc_id: int,
        doc_info: Dict[str, Any],
        doc_content: str,
        fields: List[Any],
        batch_name: str,
        batch_description: str,
        scopes: List[str],
    ) -> Dict[str, Any]:
        """Extract a single batch of fields."""
        logger.info(f"Extracting batch '{batch_name}' with {len(fields)} fields")

        # Build batch-specific prompt
        prompt = self._build_batch_prompt(
            doc_content=doc_content,
            fields=fields,
            batch_name=batch_name,
            batch_description=batch_description,
            scopes=scopes,
        )

        # Call Claude API
        try:
            client = _get_anthropic_client()
            response = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=8000,
                system="You are a real estate document analyst. Extract data precisely and completely. Return ONLY valid JSON.",
                messages=[{"role": "user", "content": prompt}]
            )

            response_text = response.content[0].text
            extractions = self._parse_batch_response(response_text, fields, scopes)

            if extractions is None:
                return {
                    'batch': batch_name,
                    'success': False,
                    'error': 'Failed to parse JSON response',
                    'fields': len(fields),
                    'raw_response': response_text[:500] if response_text else None,
                }

            # ISSUE 1 FIX: Deduplicate within batch before staging
            extractions = deduplicate_extractions(extractions)

            if not extractions:
                return {
                    'batch': batch_name,
                    'success': True,
                    'staged': 0,
                    'note': 'No extractable fields found in response',
                    'fields': len(fields),
                    'raw_response': response_text[:500] if response_text else None,
                }

            # Stage extractions
            staged_count = self._stage_batch_extractions(
                doc_id=doc_id,
                doc_info=doc_info,
                extractions=extractions,
                fields=fields,
                scopes=scopes,
            )

            return {
                'batch': batch_name,
                'success': True,
                'staged': staged_count,
                'fields': len(fields),
                'extracted': len(extractions),
            }

        except Exception as e:
            logger.exception(f"Batch extraction failed: {e}")
            return {
                'batch': batch_name,
                'success': False,
                'error': str(e),
                'fields': len(fields),
            }

    def _build_batch_prompt(
        self,
        doc_content: str,
        fields: List[Any],
        batch_name: str,
        batch_description: str,
        scopes: List[str],
    ) -> str:
        """Build extraction prompt for a specific batch."""

        # Build field list with extraction hints
        field_list = []
        for f in fields:
            field_key = f.field_key
            label = f.label
            data_type = f.field_type
            scope = f.scope

            field_list.append(f"- **{field_key}** ({data_type}, scope={scope}): {label}")

        field_block = "\n".join(field_list)

        # Check if batch includes array scopes
        array_scopes = {'unit', 'unit_type', 'sales_comp', 'rent_comp'}
        has_array = any(s in array_scopes for s in scopes)

        array_instructions = ""
        if has_array:
            array_instructions = self._get_array_instructions(scopes, fields)

        # Truncate content if needed
        max_chars = 60000
        if len(doc_content) > max_chars:
            half = max_chars // 2
            doc_content = doc_content[:half] + "\n\n[...content truncated...]\n\n" + doc_content[-half:]

        return f"""You are extracting structured data from a real estate document.

## Extraction Batch: {batch_name}
{batch_description}

## Fields to Extract

{field_block}

{array_instructions}

## Document Content

<document>
{doc_content}
</document>

## Instructions

1. For each field listed, search the document for relevant information
2. Extract the value if found with high precision
3. Include a brief source snippet as evidence
4. Rate confidence (high, medium, low) based on clarity of evidence
5. If not found, omit from response

## Output Format

Return a JSON object with field_key as keys:
```json
{{
  "field_key_1": {{
    "value": <extracted value>,
    "confidence": "high",
    "source_quote": "brief quote from document"
  }},
  "field_key_2": {{
    "value": <extracted value>,
    "confidence": "medium",
    "source_quote": "another quote"
  }}
}}
```

Rules:
- Numeric fields: return numbers (150000 not "150,000")
- Dates: ISO format "2024-01-15"
- Percentages: return decimal (0.05 for 5%)
- Booleans: true/false
- Only include fields you found in the document

Return ONLY the JSON object, no other text."""

    def _get_array_instructions(self, scopes: List[str], fields: List[Any]) -> str:
        """Get array extraction instructions based on scopes in batch."""
        instructions = ["## Array Extraction (IMPORTANT)\n"]

        if 'unit' in scopes:
            unit_field_keys = [f.field_key for f in fields if f.scope == 'unit'][:8]
            instructions.append(f"""
### Rent Roll (scope=unit)
Fields: {', '.join(unit_field_keys)}...

Extract ALL units from rent roll tables. Return array under "unit_number":
```json
{{
  "unit_number": {{
    "value": [
      {{"unit_number": "100", "unit_type": "1BR/1BA", "bedrooms": 1, "current_rent": 1595, "market_rent": 1650, "occupancy_status": "Occupied"}},
      {{"unit_number": "101", "unit_type": "2BR/2BA", "bedrooms": 2, "current_rent": 2150}}
    ],
    "confidence": "high",
    "source_quote": "Rent Roll pages 30-35"
  }}
}}
```
IMPORTANT: Extract ALL units, not just a sample.""")

        if 'unit_type' in scopes:
            ut_field_keys = [f.field_key for f in fields if f.scope == 'unit_type'][:6]
            instructions.append(f"""
### Unit Types (scope=unit_type)
Fields: {', '.join(ut_field_keys)}...

Extract unit type summary as array under "market_rent":
```json
{{
  "market_rent": {{
    "value": [
      {{"unit_type_name": "1BR/1BA", "unit_count": 24, "bedrooms": 1, "bathrooms": 1, "avg_square_feet": 650, "value": 1650}},
      {{"unit_type_name": "2BR/2BA", "unit_count": 48, "bedrooms": 2, "bathrooms": 2, "avg_square_feet": 950, "value": 2150}}
    ],
    "confidence": "high",
    "source_quote": "Unit Mix Summary"
  }}
}}
```""")

        if 'sales_comp' in scopes:
            sc_field_keys = [f.field_key for f in fields if f.scope == 'sales_comp'][:5]
            instructions.append(f"""
### Sales Comparables (scope=sales_comp)
Fields: {', '.join(sc_field_keys)}...

Extract ALL sales comps as array under "sales_comp_name":
```json
{{
  "sales_comp_name": {{
    "value": [
      {{"property_name": "Park Place", "address": "123 Main St", "sale_price": 45000000, "units": 120, "cap_rate": 0.045}},
      {{"property_name": "Harbor View", "address": "456 Oak Ave", "sale_price": 38000000, "units": 98}}
    ],
    "confidence": "high",
    "source_quote": "Sales Comparables page 22"
  }}
}}
```""")

        if 'rent_comp' in scopes:
            rc_field_keys = [f.field_key for f in fields if f.scope == 'rent_comp'][:5]
            instructions.append(f"""
### Rent Comparables (scope=rent_comp)
Fields: {', '.join(rc_field_keys)}...

Extract ALL rent comps as array under "rent_comp_name":
```json
{{
  "rent_comp_name": {{
    "value": [
      {{"property_name": "Airo South Bay", "year_built": 2021, "total_units": 230, "asking_rent": 2650}},
      {{"property_name": "The Mark", "year_built": 2019, "total_units": 180, "asking_rent": 2450}}
    ],
    "confidence": "high",
    "source_quote": "Rent Comparables Survey"
  }}
}}
```""")

        return "\n".join(instructions)

    def _get_document_content(self, doc_id: int) -> Optional[str]:
        """Get document text content from embeddings or direct extraction."""
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT content_text
                FROM landscape.knowledge_embeddings
                WHERE source_id = %s AND source_type = 'document_chunk'
                ORDER BY embedding_id
                LIMIT 100
            """, [doc_id])
            rows = cursor.fetchall()

        if rows:
            return "\n\n".join(row[0] for row in rows if row[0])

        # Fallback: fetch document storage_uri and extract text directly
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT storage_uri, mime_type, doc_name
                FROM landscape.core_doc
                WHERE doc_id = %s
            """, [doc_id])
            doc_row = cursor.fetchone()

        if not doc_row:
            return None

        storage_uri, mime_type, _doc_name = doc_row
        try:
            from .text_extraction import extract_text_from_url
            extracted_text, extract_error = extract_text_from_url(storage_uri, mime_type)
            if extracted_text:
                return extracted_text
            if extract_error:
                logger.warning(f"[doc_id={doc_id}] Direct text extraction failed: {extract_error}")
        except Exception as e:
            logger.exception(f"[doc_id={doc_id}] Direct text extraction exception: {e}")

        return None

    def _get_document_info(self, doc_id: int) -> Optional[Dict[str, Any]]:
        """Get document info."""
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT doc_id, doc_name, project_id, doc_type
                FROM landscape.core_doc
                WHERE doc_id = %s
            """, [doc_id])
            row = cursor.fetchone()

        if row:
            return {
                'doc_id': row[0],
                'doc_name': row[1],
                'project_id': row[2],
                'doc_type': row[3],
            }
        return None

    def _parse_batch_response(
        self,
        response_text: str,
        fields: List[Any],
        scopes: List[str]
    ) -> Optional[Dict[str, Any]]:
        """Parse Claude's JSON response."""
        # Find JSON object in response
        start = response_text.find('{')
        end = response_text.rfind('}') + 1

        if start == -1 or end == 0:
            logger.error("No JSON object found in response")
            return None

        try:
            parsed = json.loads(response_text[start:end])
            return parsed if isinstance(parsed, dict) else None
        except json.JSONDecodeError as e:
            logger.error(f"JSON parse error: {e}")
            # Try to extract partial JSON
            try:
                # Sometimes response has markdown code blocks
                clean = response_text.replace('```json', '').replace('```', '').strip()
                start = clean.find('{')
                end = clean.rfind('}') + 1
                if start >= 0 and end > start:
                    parsed = json.loads(clean[start:end])
                    return parsed if isinstance(parsed, dict) else None
            except Exception:
                pass
            return None

    def _stage_batch_extractions(
        self,
        doc_id: int,
        doc_info: Dict[str, Any],
        extractions: Dict[str, Any],
        fields: List[Any],
        scopes: List[str],
    ) -> int:
        """Stage batch extractions to ai_extraction_staging table."""
        # Build field lookup
        field_map = {f.field_key: f for f in fields}
        array_scopes = {'unit', 'unit_type', 'sales_comp', 'rent_comp'}
        staged_count = 0

        with connection.cursor() as cursor:
            for field_key, extraction in extractions.items():
                if not isinstance(extraction, dict):
                    continue

                value = extraction.get('value')
                confidence = extraction.get('confidence', 'medium')
                source_quote = extraction.get('source_quote', '')

                # Map confidence string to score
                conf_map = {'high': 0.9, 'medium': 0.7, 'low': 0.5}
                conf_score = conf_map.get(confidence, 0.7)

                # Get field mapping
                field = field_map.get(field_key)
                if not field:
                    # Check if this is an array-scope field
                    for f in fields:
                        if f.scope in array_scopes and field_key in [
                            'unit_number', 'market_rent', 'sales_comp_name', 'rent_comp_name'
                        ]:
                            field = f
                            break

                if not field:
                    logger.warning(f"No field mapping for {field_key}")
                    continue

                # Handle array-scoped fields
                if field.scope in array_scopes and isinstance(value, list):
                    for array_idx, item in enumerate(value):
                        # Determine scope_label
                        if field.scope == 'unit_type':
                            scope_label = item.get('unit_type_name', f'Type {array_idx + 1}')
                        elif field.scope == 'unit':
                            scope_label = item.get('unit_number', f'Unit {array_idx + 1}')
                        elif field.scope == 'sales_comp':
                            scope_label = item.get('property_name', f'Sales Comp {array_idx + 1}')
                        elif field.scope == 'rent_comp':
                            scope_label = item.get('property_name', f'Rent Comp {array_idx + 1}')
                        else:
                            scope_label = f'Item {array_idx + 1}'

                        # Get extraction_type from doc_info
                        extraction_type = doc_info.get('doc_type', 'unknown')

                        # Check for conflict with existing accepted value (for array items)
                        conflict = check_for_conflict(
                            project_id=self.project_id,
                            field_key=field_key,
                            new_value=item,
                            new_doc_id=doc_id,
                            new_confidence=conf_score,
                            scope_label=scope_label,
                            array_index=array_idx
                        )

                        # Determine status and conflict reference
                        status = 'conflict' if conflict else 'pending'
                        conflict_extraction_id = conflict['existing_extraction_id'] if conflict else None

                        # Use ON CONFLICT to dedupe array items: update if new confidence is higher
                        cursor.execute("""
                            INSERT INTO landscape.ai_extraction_staging (
                                project_id, doc_id, field_key, extracted_value,
                                confidence_score, source_snippet, status, scope,
                                scope_label, array_index, target_table, target_field,
                                db_write_type, selector_json, property_type, extraction_type,
                                conflict_with_extraction_id, created_at
                            ) VALUES (
                                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW()
                            )
                            ON CONFLICT (project_id, field_key, COALESCE(scope_label, ''), COALESCE(array_index, 0))
                            WHERE status = 'pending'
                            DO UPDATE SET
                                extracted_value = CASE
                                    WHEN EXCLUDED.confidence_score > landscape.ai_extraction_staging.confidence_score
                                    THEN EXCLUDED.extracted_value
                                    ELSE landscape.ai_extraction_staging.extracted_value
                                END,
                                confidence_score = GREATEST(EXCLUDED.confidence_score, landscape.ai_extraction_staging.confidence_score),
                                source_snippet = CASE
                                    WHEN EXCLUDED.confidence_score > landscape.ai_extraction_staging.confidence_score
                                    THEN EXCLUDED.source_snippet
                                    ELSE landscape.ai_extraction_staging.source_snippet
                                END,
                                doc_id = CASE
                                    WHEN EXCLUDED.confidence_score > landscape.ai_extraction_staging.confidence_score
                                    THEN EXCLUDED.doc_id
                                    ELSE landscape.ai_extraction_staging.doc_id
                                END,
                                conflict_with_extraction_id = COALESCE(EXCLUDED.conflict_with_extraction_id, landscape.ai_extraction_staging.conflict_with_extraction_id),
                                status = CASE
                                    WHEN EXCLUDED.conflict_with_extraction_id IS NOT NULL THEN 'conflict'
                                    ELSE landscape.ai_extraction_staging.status
                                END,
                                created_at = NOW()
                        """, [
                            self.project_id,
                            doc_id,
                            field_key,
                            json.dumps(item),
                            conf_score,
                            source_quote[:500] if source_quote else None,
                            status,
                            field.scope,
                            scope_label,
                            array_idx,
                            field.table_name,
                            field.column_name,
                            field.db_write_type,
                            json.dumps(field.selector_json) if field.selector_json else None,
                            self.property_type,
                            extraction_type,
                            conflict_extraction_id,
                        ])
                        staged_count += 1
                else:
                    # Single value extraction
                    extraction_type = doc_info.get('doc_type', 'unknown')

                    # Check for conflict with existing accepted value
                    json_value = json.dumps(value) if not isinstance(value, (str, int, float, bool)) else json.dumps(value)
                    conflict = check_for_conflict(
                        project_id=self.project_id,
                        field_key=field_key,
                        new_value=value,
                        new_doc_id=doc_id,
                        new_confidence=conf_score,
                        scope_label=None,
                        array_index=None
                    )

                    # Determine status and conflict reference
                    status = 'conflict' if conflict else 'pending'
                    conflict_extraction_id = conflict['existing_extraction_id'] if conflict else None

                    # Use ON CONFLICT to dedupe: update if new confidence is higher
                    cursor.execute("""
                        INSERT INTO landscape.ai_extraction_staging (
                            project_id, doc_id, field_key, extracted_value,
                            confidence_score, source_snippet, status, scope,
                            scope_label, array_index,
                            target_table, target_field, db_write_type, selector_json,
                            property_type, extraction_type, conflict_with_extraction_id, created_at
                        ) VALUES (
                            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW()
                        )
                        ON CONFLICT (project_id, field_key, COALESCE(scope_label, ''), COALESCE(array_index, 0))
                        WHERE status = 'pending'
                        DO UPDATE SET
                            extracted_value = CASE
                                WHEN EXCLUDED.confidence_score > landscape.ai_extraction_staging.confidence_score
                                THEN EXCLUDED.extracted_value
                                ELSE landscape.ai_extraction_staging.extracted_value
                            END,
                            confidence_score = GREATEST(EXCLUDED.confidence_score, landscape.ai_extraction_staging.confidence_score),
                            source_snippet = CASE
                                WHEN EXCLUDED.confidence_score > landscape.ai_extraction_staging.confidence_score
                                THEN EXCLUDED.source_snippet
                                ELSE landscape.ai_extraction_staging.source_snippet
                            END,
                            doc_id = CASE
                                WHEN EXCLUDED.confidence_score > landscape.ai_extraction_staging.confidence_score
                                THEN EXCLUDED.doc_id
                                ELSE landscape.ai_extraction_staging.doc_id
                            END,
                            conflict_with_extraction_id = COALESCE(EXCLUDED.conflict_with_extraction_id, landscape.ai_extraction_staging.conflict_with_extraction_id),
                            status = CASE
                                WHEN EXCLUDED.conflict_with_extraction_id IS NOT NULL THEN 'conflict'
                                ELSE landscape.ai_extraction_staging.status
                            END,
                            created_at = NOW()
                    """, [
                        self.project_id,
                        doc_id,
                        field_key,
                        json_value,
                        conf_score,
                        source_quote[:500] if source_quote else None,
                        status,
                        field.scope,
                        None,  # scope_label for single values
                        None,  # array_index for single values
                        field.table_name,
                        field.column_name,
                        field.db_write_type,
                        json.dumps(field.selector_json) if field.selector_json else None,
                        self.property_type,
                        extraction_type,
                        conflict_extraction_id,
                    ])
                    staged_count += 1

        logger.info(f"Staged {staged_count} extractions for doc {doc_id}")
        return staged_count


def extract_document_batched(
    project_id: int,
    doc_id: int,
    property_type: str = 'multifamily',
    batches: Optional[List[str]] = None,
    user_id: Optional[int] = None,
) -> Dict[str, Any]:
    """Convenience function for batched extraction."""
    print(f"=== EXTRACT_DOCUMENT_BATCHED SERVICE ENTRY ===")
    print(f"PROJECT_ID: {project_id}, DOC_ID: {doc_id}, PROPERTY_TYPE: {property_type}")
    print(f"BATCHES: {batches}")

    service = BatchedExtractionService(project_id, property_type)
    return service.extract_document_batched(doc_id, batches, user_id)


# ============================================================
# CHUNKED RENT ROLL EXTRACTION
# ============================================================

RENT_ROLL_CONFIG = {
    'chunk_size': 35,  # Units per chunk
    'max_chunks': 10,  # Safety limit
    'fields_per_unit': [
        'unit_number',
        'unit_type',
        'bedrooms',
        'bathrooms',
        'square_feet',
        'current_rent',
        'market_rent',
        'occupancy_status',
        'lease_start',
        'lease_end',
        'tenant_name',
        'move_in_date',
        'rent_effective_date',
    ],
}


class ChunkedRentRollExtractor:
    """
    Extracts rent roll data in chunks to handle 100+ unit properties.

    The problem: Claude's response truncates when trying to return 113+ units
    in one call. Solution: Extract rent roll in chunks of ~35 units per call.
    """

    def __init__(self, project_id: int, property_type: str = 'multifamily'):
        self.project_id = project_id
        self.property_type = property_type
        self.registry = get_registry()

    def extract_rent_roll_chunked(
        self,
        doc_id: int,
        user_id: Optional[int] = None,
    ) -> Dict[str, Any]:
        """
        Extract rent roll in chunks.

        1. First, get total unit count estimate
        2. Split into chunks of ~35 units
        3. Extract each chunk with specific unit range
        4. Combine and stage all units

        Args:
            doc_id: Document ID containing rent roll
            user_id: For audit trail

        Returns:
            {
                'success': bool,
                'total_units': int,
                'chunks_processed': int,
                'units_extracted': int,
                'staged_count': int,
                'errors': []
            }
        """
        errors = []
        all_units = []

        # Get document content
        doc_content = self._get_document_content(doc_id)
        if not doc_content:
            return {
                'success': False,
                'error': f'No content found for document {doc_id}',
            }

        doc_info = self._get_document_info(doc_id)
        if not doc_info:
            return {
                'success': False,
                'error': f'Document {doc_id} not found',
            }

        logger.info(f"Starting chunked rent roll extraction for doc {doc_id}")

        # Step 1: Estimate total unit count
        estimated_units = self._estimate_unit_count(doc_content)
        logger.info(f"Estimated unit count: {estimated_units}")

        # Calculate number of chunks needed
        chunk_size = RENT_ROLL_CONFIG['chunk_size']
        max_chunks = RENT_ROLL_CONFIG['max_chunks']
        num_chunks = min((estimated_units + chunk_size - 1) // chunk_size, max_chunks)

        if num_chunks == 0:
            num_chunks = 1  # At least one chunk

        logger.info(f"Will extract in {num_chunks} chunks of {chunk_size} units each")

        # Step 2: Extract each chunk
        for chunk_idx in range(num_chunks):
            start_unit = chunk_idx * chunk_size + 1
            end_unit = (chunk_idx + 1) * chunk_size

            logger.info(f"Extracting chunk {chunk_idx + 1}/{num_chunks}: units {start_unit}-{end_unit}")

            try:
                chunk_result = self._extract_rent_roll_chunk(
                    doc_content=doc_content,
                    chunk_idx=chunk_idx,
                    start_unit=start_unit,
                    end_unit=end_unit,
                    total_chunks=num_chunks,
                )

                if chunk_result.get('success'):
                    units = chunk_result.get('units', [])
                    all_units.extend(units)
                    logger.info(f"Chunk {chunk_idx + 1}: extracted {len(units)} units")
                else:
                    error_msg = f"Chunk {chunk_idx + 1}: {chunk_result.get('error', 'Unknown error')}"
                    errors.append(error_msg)
                    logger.warning(error_msg)

            except Exception as e:
                error_msg = f"Chunk {chunk_idx + 1}: {str(e)}"
                errors.append(error_msg)
                logger.exception(error_msg)

        # Step 3: Deduplicate units by unit_number
        unique_units = {}
        for unit in all_units:
            unit_num = unit.get('unit_number')
            if unit_num and unit_num not in unique_units:
                unique_units[unit_num] = unit

        deduped_units = list(unique_units.values())
        logger.info(f"After dedup: {len(deduped_units)} unique units (from {len(all_units)} total)")

        # Step 4: Stage all units
        staged_count = 0
        if deduped_units:
            staged_count = self._stage_rent_roll_units(
                doc_id=doc_id,
                doc_info=doc_info,
                units=deduped_units,
            )

        return {
            'success': len(errors) == 0 or len(deduped_units) > 0,
            'doc_id': doc_id,
            'project_id': self.project_id,
            'estimated_units': estimated_units,
            'chunks_processed': num_chunks,
            'units_extracted': len(deduped_units),
            'staged_count': staged_count,
            'errors': errors,
        }

    def _estimate_unit_count(self, doc_content: str) -> int:
        """
        Estimate total unit count from document.

        Looks for patterns like:
        - "113 units"
        - "Total Units: 113"
        - Count of unit number patterns
        """
        import re

        # Collect all potential unit counts from various patterns
        candidates = []
        doc_lower = doc_content.lower()

        # Pattern: "X units" or "X total units"
        matches = re.findall(r'(\d+)\s*(?:total\s+)?units', doc_lower)
        for m in matches:
            count = int(m)
            if 10 <= count <= 500:
                candidates.append(count)

        # Pattern: "total units: X"
        matches = re.findall(r'total\s+units[:\s]+(\d+)', doc_lower)
        for m in matches:
            count = int(m)
            if 10 <= count <= 500:
                candidates.append(count)

        # Pattern: "X-unit" property
        matches = re.findall(r'(\d+)-unit', doc_lower)
        for m in matches:
            count = int(m)
            if 10 <= count <= 500:
                candidates.append(count)

        # If we found candidates, pick the most common or the one in typical MF range (50-200)
        if candidates:
            # Prefer counts in typical multifamily range
            typical_range = [c for c in candidates if 50 <= c <= 200]
            if typical_range:
                # Pick the most frequent in typical range, or the highest
                from collections import Counter
                counter = Counter(typical_range)
                return counter.most_common(1)[0][0]
            # Otherwise pick the most frequent candidate
            from collections import Counter
            counter = Counter(candidates)
            return counter.most_common(1)[0][0]

        # Fallback: count unit number patterns more aggressively
        # Look for apartment-style unit numbers (100-999)
        unit_numbers = re.findall(r'\b(\d{3})\b', doc_content)

        # Filter to likely unit numbers
        likely_units = set()
        for u in unit_numbers:
            num = int(u)
            # Apartment-style unit numbers typically 100-999
            # Exclude common non-unit numbers like years (2020-2025), amounts
            if 100 <= num <= 999 and num < 2000:
                likely_units.add(u)

        if len(likely_units) >= 5:
            # Return the count, but pad by 50% to ensure we get all units
            # (better to have extra chunks than miss units)
            return int(len(likely_units) * 1.5)

        # Look for rent roll table rows - each row is typically a unit
        # Count rows that look like unit entries
        status_count = len(re.findall(r'(?:occupied|vacant|notice)', doc_lower))
        if status_count >= 10:
            return int(status_count * 1.1)  # Pad slightly

        # Default - assume a medium-sized property if we can't estimate
        return 120

    def _extract_rent_roll_chunk(
        self,
        doc_content: str,
        chunk_idx: int,
        start_unit: int,
        end_unit: int,
        total_chunks: int,
    ) -> Dict[str, Any]:
        """Extract a single chunk of rent roll units."""

        prompt = self._build_rent_roll_chunk_prompt(
            doc_content=doc_content,
            chunk_idx=chunk_idx,
            start_unit=start_unit,
            end_unit=end_unit,
            total_chunks=total_chunks,
        )

        try:
            client = _get_anthropic_client()
            response = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=8000,
                system="""You are a real estate rent roll data extractor.
Extract unit-level data with high precision. Return ONLY valid JSON array.
IMPORTANT: Extract exactly the ROW RANGE requested - count rows from the first data row after headers.
The row number is NOT the same as unit number (e.g., row 1 might be unit 100, row 36 might be unit 201).
CRITICAL: For tenant_name, extract the EXACT name from the source. NEVER use placeholders like "Current Tenant" or "Tenant" - always use the actual name or omit the field.""",
                messages=[{"role": "user", "content": prompt}]
            )

            response_text = response.content[0].text
            units = self._parse_rent_roll_response(response_text)

            if units is None:
                return {
                    'success': False,
                    'error': 'Failed to parse response',
                    'raw_response': response_text[:500],
                }

            return {
                'success': True,
                'units': units,
                'count': len(units),
            }

        except Exception as e:
            logger.exception(f"Chunk extraction failed: {e}")
            return {
                'success': False,
                'error': str(e),
            }

    def _build_rent_roll_chunk_prompt(
        self,
        doc_content: str,
        chunk_idx: int,
        start_unit: int,
        end_unit: int,
        total_chunks: int,
    ) -> str:
        """Build prompt for extracting a specific chunk of units."""

        fields = RENT_ROLL_CONFIG['fields_per_unit']
        field_list = "\n".join([f"- {f}" for f in fields])

        # Truncate content if very long
        max_chars = 80000
        if len(doc_content) > max_chars:
            # Keep more of the middle where rent roll likely is
            quarter = max_chars // 4
            half = max_chars // 2
            doc_content = (
                doc_content[:quarter] +
                "\n\n[...content trimmed...]\n\n" +
                doc_content[len(doc_content)//2 - half//2:len(doc_content)//2 + half//2] +
                "\n\n[...content trimmed...]\n\n" +
                doc_content[-quarter:]
            )

        return f"""Extract rent roll data from this document.

## CHUNK INSTRUCTIONS
This is chunk {chunk_idx + 1} of {total_chunks}. The rent roll has approximately {total_chunks * 35} units total.

IMPORTANT: Extract units based on their ROW POSITION in the rent roll table (not their unit number):
- Chunk 1: Extract rows 1-35 (the first 35 units in the table)
- Chunk 2: Extract rows 36-70 (the next 35 units)
- Chunk 3: Extract rows 71-105
- And so on...

For THIS chunk ({chunk_idx + 1}), extract rows {start_unit} through {end_unit} from the rent roll table.
Count from the first data row (after headers). Include ALL units in this row range regardless of their unit numbers.

## FIELDS TO EXTRACT (per unit)
{field_list}

## DOCUMENT CONTENT
<document>
{doc_content}
</document>

## OUTPUT FORMAT
Return a JSON array of unit objects:
```json
[
  {{
    "unit_number": "100",
    "unit_type": "1BR/1BA",
    "bedrooms": 1,
    "bathrooms": 1,
    "square_feet": 650,
    "current_rent": 1595,
    "market_rent": 1650,
    "occupancy_status": "Occupied",
    "lease_start": "2024-01-15",
    "lease_end": "2025-01-14",
    "tenant_name": "John Smith",
    "move_in_date": "2024-01-15",
    "rent_effective_date": "2024-01-15"
  }},
  {{
    "unit_number": "101",
    "unit_type": "2BR/2BA",
    "bedrooms": 2,
    "bathrooms": 2,
    "square_feet": 950,
    "current_rent": 2150,
    "market_rent": 2250,
    "occupancy_status": "Vacant"
  }}
]
```

RULES:
- Extract ONLY units in the requested range for this chunk
- Numeric fields: numbers only (1595 not "$1,595")
- Dates: ISO format "YYYY-MM-DD"
- Include all available fields; omit if not found
- tenant_name: Extract the EXACT name from the source document. NEVER substitute "Current Tenant", "Tenant", or any placeholder - use the actual name or omit the field entirely if not present
- Return ONLY the JSON array, no other text"""

    def _parse_rent_roll_response(self, response_text: str) -> Optional[List[Dict]]:
        """Parse JSON array from response."""
        # Try direct parse
        try:
            result = json.loads(response_text)
            if isinstance(result, list):
                return result
        except json.JSONDecodeError:
            pass

        # Find JSON array in response
        import re

        # Clean markdown code blocks
        clean = response_text.replace('```json', '').replace('```', '').strip()

        # Find array
        start = clean.find('[')
        end = clean.rfind(']') + 1

        if start >= 0 and end > start:
            try:
                result = json.loads(clean[start:end])
                if isinstance(result, list):
                    return result
            except json.JSONDecodeError:
                pass

        return None

    def _stage_rent_roll_units(
        self,
        doc_id: int,
        doc_info: Dict[str, Any],
        units: List[Dict],
    ) -> int:
        """Stage rent roll units to extraction staging table."""

        staged_count = 0

        # Clear any existing pending extractions for this project's rent roll
        # This prevents IntegrityError on re-upload of same file
        with connection.cursor() as cursor:
            cursor.execute("""
                DELETE FROM landscape.ai_extraction_staging
                WHERE project_id = %s
                AND scope = 'unit'
                AND status = 'pending'
            """, [self.project_id])
            deleted_count = cursor.rowcount
            if deleted_count > 0:
                logger.info(f"Cleared {deleted_count} existing pending unit extractions for project {self.project_id}")

        # Get unit scope fields from registry
        unit_fields = self.registry.get_fields_by_scope('unit', self.property_type)

        # Find primary field for unit scope (unit_number)
        primary_field = None
        for f in unit_fields:
            if f.field_key == 'unit_number':
                primary_field = f
                break

        if not primary_field:
            # Fallback
            primary_field = unit_fields[0] if unit_fields else None

        with connection.cursor() as cursor:
            for array_idx, unit in enumerate(units):
                unit_number = unit.get('unit_number', f'Unit {array_idx + 1}')

                # Clean up tenant_name - remove any placeholders that slipped through
                raw_tenant = unit.get('tenant_name')
                clean_tenant = extract_tenant_name(raw_tenant) if raw_tenant else None
                if clean_tenant != raw_tenant:
                    if clean_tenant:
                        unit['tenant_name'] = clean_tenant
                    else:
                        # Remove placeholder tenant_name entirely
                        unit.pop('tenant_name', None)

                # Get extraction_type from doc_info
                extraction_type = doc_info.get('doc_type', 'rent_roll')

                # Stage the entire unit as a JSON object
                cursor.execute("""
                    INSERT INTO landscape.ai_extraction_staging (
                        project_id, doc_id, field_key, extracted_value,
                        confidence_score, source_snippet, status, scope,
                        scope_label, array_index, target_table, target_field,
                        db_write_type, property_type, extraction_type, created_at
                    ) VALUES (
                        %s, %s, %s, %s, %s, %s, 'pending', %s, %s, %s, %s, %s, %s, %s, %s, NOW()
                    )
                """, [
                    self.project_id,
                    doc_id,
                    'unit_number',  # Primary field key
                    json.dumps(unit),
                    0.85,  # Default confidence for chunked extraction
                    f"Rent roll unit {unit_number}",
                    'unit',  # scope
                    unit_number,  # scope_label
                    array_idx,
                    primary_field.table_name if primary_field else 'tbl_mf_unit',
                    primary_field.column_name if primary_field else 'unit_number',
                    primary_field.db_write_type if primary_field else 'insert',
                    self.property_type,
                    extraction_type,
                ])
                staged_count += 1

        logger.info(f"Staged {staged_count} rent roll units for doc {doc_id}")
        return staged_count

    def _get_document_content(self, doc_id: int) -> Optional[str]:
        """
        Get document text content.

        For Excel/CSV files: Download and parse directly for complete data.
        For other files: Use embeddings.
        """
        import tempfile
        import requests

        # Get document info including storage URI and mime type
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT storage_uri, mime_type, doc_name
                FROM landscape.core_doc
                WHERE doc_id = %s
            """, [doc_id])
            row = cursor.fetchone()

        if not row:
            return None

        storage_uri, mime_type, doc_name = row

        # For Excel/CSV files, read directly from file for complete data
        excel_types = [
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'application/vnd.ms-excel',
            'text/csv',
            'application/csv',
        ]

        if mime_type in excel_types and storage_uri:
            try:
                logger.info(f"Reading Excel/CSV file directly for doc {doc_id}")

                # Download file
                response = requests.get(storage_uri, timeout=60)
                response.raise_for_status()

                # Save to temp file
                suffix = '.xlsx' if 'spreadsheet' in mime_type else '.csv'
                with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                    tmp.write(response.content)
                    tmp_path = tmp.name

                try:
                    if suffix == '.xlsx':
                        content = self._parse_excel_to_text(tmp_path)
                    else:
                        content = self._parse_csv_to_text(tmp_path)

                    if content:
                        logger.info(f"Parsed {len(content)} chars from Excel/CSV")
                        return content
                finally:
                    import os
                    try:
                        os.unlink(tmp_path)
                    except Exception:
                        pass

            except Exception as e:
                logger.warning(f"Failed to read Excel/CSV directly: {e}, falling back to embeddings")

        # Fallback: Use embeddings
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT content_text
                FROM landscape.knowledge_embeddings
                WHERE source_id = %s AND source_type = 'document_chunk'
                ORDER BY embedding_id
                LIMIT 100
            """, [doc_id])
            rows = cursor.fetchall()

        if rows:
            return "\n\n".join(row[0] for row in rows if row[0])
        return None

    def _parse_excel_to_text(self, file_path: str) -> Optional[str]:
        """Parse Excel file to text format for extraction."""
        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path, read_only=True, data_only=True)
            sheet = wb.active

            lines = []
            for row in sheet.iter_rows(values_only=True):
                # Convert row to tab-separated string
                row_vals = [str(cell) if cell is not None else '' for cell in row]
                if any(v.strip() for v in row_vals):  # Skip empty rows
                    lines.append('\t'.join(row_vals))

            wb.close()
            return '\n'.join(lines)

        except Exception as e:
            logger.error(f"Error parsing Excel: {e}")
            return None

    def _parse_csv_to_text(self, file_path: str) -> Optional[str]:
        """Parse CSV file to text format for extraction."""
        try:
            import csv

            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                reader = csv.reader(f)
                lines = ['\t'.join(row) for row in reader if any(cell.strip() for cell in row)]

            return '\n'.join(lines)

        except Exception as e:
            logger.error(f"Error parsing CSV: {e}")
            return None

    def _get_document_info(self, doc_id: int) -> Optional[Dict[str, Any]]:
        """Get document info."""
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT doc_id, doc_name, project_id, doc_type
                FROM landscape.core_doc
                WHERE doc_id = %s
            """, [doc_id])
            row = cursor.fetchone()

        if row:
            return {
                'doc_id': row[0],
                'doc_name': row[1],
                'project_id': row[2],
                'doc_type': row[3],
            }
        return None


def extract_rent_roll_chunked(
    project_id: int,
    doc_id: int,
    property_type: str = 'multifamily',
    user_id: Optional[int] = None,
) -> Dict[str, Any]:
    """Convenience function for chunked rent roll extraction."""
    extractor = ChunkedRentRollExtractor(project_id, property_type)
    return extractor.extract_rent_roll_chunked(doc_id, user_id)
