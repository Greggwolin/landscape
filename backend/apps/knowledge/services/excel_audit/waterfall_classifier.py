"""
Phase 4 — Waterfall Classifier.

Reads the Waterfall sheet of a financial model and produces a structured
classification of the distribution structure: tier count, tier types,
hurdle rates, split ratios, pref rate, compounding convention, sponsor
co-invest %, and a `source_cells` map keyed to Sheet!Cell refs for every
recovered value.

Output feeds Phase 5 (Python replication). The classifier identifies WHAT
to replicate; Phase 5 replicates it. Anything the classifier can't parse
gets returned as `waterfall_type="custom"` with findings — never silently
defaulted to a template.

Design notes:
  - Heuristic, not exhaustive. CRE waterfalls are bespoke; we identify
    structure by tier-label scan + adjacent numeric read, not by parsing
    arbitrary formula trees.
  - Sheet name prefix stripping (i./m./r.) matches the Phase 0 classifier.
  - Tier labels are pattern-matched against TIER_LABEL_PATTERNS.
  - Splits detected as complementary numeric pairs that sum to ~1.0.
  - Hurdle classification: numeric in 0.05-0.25 range adjacent to a hurdle
    label = IRR; numeric in 1.2-3.0 range = EM (equity multiple).
  - Compounding inferred from formula text in the same row (presence of
    `^(1/12)`, `/12`, etc.).

Returns a dict (not a dataclass instance) so tool wrappers can serialize
directly. Every numeric value carries a Sheet!Cell ref in `source_cells`.
"""

import re
from dataclasses import dataclass, field, asdict
from typing import Dict, List, Optional, Tuple

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────

# Matches "i.", "m.", "r." at the start of a sheet name. Mirrors the
# Phase 0 classifier's pattern exactly — REQUIRES the dot. Bare-letter
# variants like "iProject" are NOT stripped here (matches existing
# pipeline behavior; CLAUDE.md note about iProject → Project is
# aspirational and not yet implemented anywhere in the audit pipeline).
PREFIX_RE = re.compile(r"^([imr])\.\s*", re.IGNORECASE)

# Sheet-name tokens that mark a candidate waterfall sheet. Mirrors the
# Phase 0 classifier's FULL_MODEL_TOKENS but extended per SKILL.md §Phase 4
# Step 1 to include partnership / allocation / splits.
WATERFALL_SHEET_TOKENS = {
    "waterfall", "wfall", "promote", "distribution",
    "returns split", "jv split", "equity split",
    "partner", "partnership", "prtnership",
    "allocation", "splits",
    "cash flow split", "cf split",
}

# Tier label patterns. Order matters: more specific patterns first so that
# "pref catch-up" doesn't match the bare "catch-up" pattern before being
# classified as catch_up.
TIER_LABEL_PATTERNS = [
    ("return_of_capital", [
        r"\breturn of capital\b",
        r"\broc\b",
        r"\bcapital return\b",
        r"\breturn.*equity\b",
    ]),
    ("pref_accrual", [
        r"\bpref(erred)? return\b",
        r"\bpreferred\b",
        r"\bpref accrual\b",
        r"\bpref\b",
    ]),
    ("catch_up", [
        r"\bcatch[- ]?up\b",
        r"\bcatchup\b",
        r"\bgp catch\b",
    ]),
    ("hurdle_split", [
        r"\bhurdle\b",
        r"\birr hurdle\b",
        r"\bem hurdle\b",
        r"\bequity multiple hurdle\b",
        r"\btier\s*\d\b",
    ]),
    ("residual_split", [
        r"\bresidual\b",
        r"\bremaining\b",
        r"\bthereafter\b",
        r"\babove\b",
        r"\bfinal split\b",
    ]),
]

# Sponsor co-invest detection
COINVEST_PATTERNS = [
    r"\bsponsor co[- ]?invest",
    r"\bgp co[- ]?invest",
    r"\bsponsor equity",
    r"\bgp equity",
    r"\bgp contribution",
    r"\bsp co[- ]?invest",
]

# Secondary labeled-value scan patterns. When the primary tier scan finds
# a tier (e.g., pref_accrual) but no value adjacent to its label row, run
# this scan over the WHOLE sheet looking for cells whose OWN text label
# matches one of these patterns AND have a numeric value adjacent. This
# is reading a labeled cell — NOT inferring a value from arbitrary numerics
# nearby. The label is the authority.
PREF_RATE_LABEL_PATTERNS = [
    re.compile(r"\bpref(erred)? return rate\b", re.IGNORECASE),
    re.compile(r"\bpref(erred)? rate\b", re.IGNORECASE),
    re.compile(r"\bpref(erred)? return\b(?!\s+is)", re.IGNORECASE),
    re.compile(r"\bpref accrual rate\b", re.IGNORECASE),
    re.compile(r"\bhurdle\s+1\b|\btier\s*1\s+irr\b", re.IGNORECASE),  # often pref = tier 1 hurdle
]

LP_SPLIT_LABEL_PATTERNS = [
    re.compile(r"\b(lp|investor|limited partner)\s*(%|share|split|allocation)?\b", re.IGNORECASE),
    re.compile(r"\b(lp|investor)\s*$", re.IGNORECASE),
]

GP_SPLIT_LABEL_PATTERNS = [
    re.compile(r"\b(gp|sponsor|general partner|promote)\s*(%|share|split|allocation)?\b", re.IGNORECASE),
    re.compile(r"\b(gp|sponsor)\s*$", re.IGNORECASE),
]

# Maximum scan windows
MAX_LABEL_SCAN_ROWS = 100   # how deep into the sheet we look for tier labels
MAX_LABEL_SCAN_COLS = 4     # leftmost columns to treat as label columns
MAX_VALUE_SCAN_COLS = 8     # how many cols right of a label to read for values
MAX_COINVEST_SCAN_ROWS = 80
# NOTE: no widened scan. Per the no-inference directive (see
# feedback_no_value_inference.md), if narrow scan doesn't find a value
# adjacent to the label, we report the gap and ask the user — we do NOT
# go fishing in distant cells for a "plausible" candidate.

# Label sanity thresholds — reject prose-as-tier captures (Brownstone fix)
LABEL_MAX_WORDS = 6         # real tier labels are short ("Tier 1 IRR Hurdle Split")
LABEL_MAX_CHARS = 50        # title rows / explanatory sentences exceed this
LABEL_REJECT_END_PUNCT = (".", ":", ";", "?", "!")
LABEL_REJECT_SUBSTRINGS = (  # narrative giveaways
    " first,", " second,", " third,", " fourth,", " fifth,",
    " then ", " above ", " below ", " however,", " note:", " e.g.",
)

# Heuristic ranges
IRR_MIN, IRR_MAX = 0.04, 0.20      # plausible IRR HURDLE band (definitional)
IRR_HIGH_BAND = 0.25               # 0.20–0.25 = accepted but flagged as suspect
IRR_HARD_CEILING = 0.30            # > 0.30 = almost certainly an IRR result, reject
EM_MIN, EM_MAX = 1.10, 4.00        # plausible equity multiple hurdle
PREF_RATE_MIN, PREF_RATE_MAX = 0.04, 0.15
COINVEST_MAX = 0.50                # sponsor coinvest rarely above 50%

# Tier dedupe — extract "Tier N" number from labels so accrual / distribution /
# narrative variants of the same logical tier collapse to one entry
TIER_NUMBER_RE = re.compile(r"\btier\s*(\d+)\b", re.IGNORECASE)

# Strong tier-definition keywords. When present in a prose-shaped row, they
# OVERRIDE the prose filter — the row is treated as a real tier definition
# even though it's sentence-shaped. Brownstone calibration: pref tier defined
# as "First, Preferred Return is distributed to Investor and Sponsor Pari-Passu"
# is sentence-shaped but unambiguously a tier definition.
#
# The override is keyword-strict: bare "waterfall", "tier", or "hurdle" do NOT
# qualify (those are common in narrative description). Only multi-word noun
# phrases that name a SPECIFIC tier type qualify.
STRONG_DEFINITION_PATTERNS = [
    re.compile(r"\bpreferred return\b", re.IGNORECASE),
    re.compile(r"\bpref(erred)? accrual\b", re.IGNORECASE),
    re.compile(r"\breturn of capital\b", re.IGNORECASE),
    re.compile(r"\b(?:gp|sponsor|investor)\s+catch[- ]?up\b", re.IGNORECASE),
    re.compile(r"\bcatch[- ]?up (?:tier|provision|distribution)\b", re.IGNORECASE),
    re.compile(r"\btier\s*\d+\s+(?:promote|distribution|allocation|hurdle|split)\b", re.IGNORECASE),
    re.compile(r"\bremaining cash\b", re.IGNORECASE),
    re.compile(r"\bresidual (?:split|cash|distribution)\b", re.IGNORECASE),
]


# ─────────────────────────────────────────────────────────────────────────────
# Dataclasses
# ─────────────────────────────────────────────────────────────────────────────


@dataclass
class WaterfallTier:
    index: int                              # 1-indexed position
    tier_type: str                          # return_of_capital | pref_accrual | catch_up | hurdle_split | residual_split | unknown
    label: str                              # raw label text as found
    label_cell: Optional[str] = None        # Sheet!Cell of the label
    hurdle_type: Optional[str] = None       # IRR | EM | pref_rate | None
    hurdle_value: Optional[float] = None
    split_lp_pct: Optional[float] = None
    split_gp_pct: Optional[float] = None
    pref_rate: Optional[float] = None
    pref_compounding: Optional[str] = None  # monthly | simple | quarterly | annual
    source_cells: Dict[str, str] = field(default_factory=dict)


@dataclass
class WaterfallClassification:
    waterfall_type: str                     # tiered_irr_hurdle | pref_then_split | pref_catchup_split | em_hurdle | hybrid | custom | none
    tier_count: int
    sheet_name: Optional[str]
    tiers: List[WaterfallTier] = field(default_factory=list)
    pref_rate: Optional[float] = None
    pref_compounding: Optional[str] = None
    hurdle_type: Optional[str] = None       # IRR | EM | IRR_EM
    sponsor_coinvest_pct: Optional[float] = None
    source_cells: Dict[str, str] = field(default_factory=dict)
    findings: List[Dict] = field(default_factory=list)
    rationale: str = ""

    def to_dict(self) -> Dict:
        d = asdict(self)
        # Tiers are dataclasses inside a list — asdict already recurses, but
        # be explicit so the JSON shape is stable.
        return d


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────


def _strip_prefix(name: str) -> str:
    return PREFIX_RE.sub("", name or "").strip().lower()


def _find_waterfall_sheets(wb: Workbook) -> List[str]:
    """Return sheet names that match any waterfall keyword (prefix-stripped)."""
    hits: List[str] = []
    for name in wb.sheetnames:
        clean = _strip_prefix(name)
        for tok in WATERFALL_SHEET_TOKENS:
            if tok in clean:
                hits.append(name)
                break
    return hits


def _is_prose_label(text: str) -> bool:
    """
    True if `text` reads like prose / narrative rather than a structural tier
    label. Catches workbook title rows, explanatory sentences, and section
    notes — the Brownstone classifier's biggest false-positive source.

    Reject rules (any one trips it):
      - longer than LABEL_MAX_CHARS
      - more than LABEL_MAX_WORDS whitespace-delimited tokens
      - ends with sentence-style punctuation
      - contains commas AND len > 30 (multi-clause narrative)
      - contains narrative giveaway substrings
    """
    if not text:
        return True
    s = text.strip()
    if len(s) > LABEL_MAX_CHARS:
        return True
    if len(s.split()) > LABEL_MAX_WORDS:
        return True
    if s.endswith(LABEL_REJECT_END_PUNCT):
        return True
    if "," in s and len(s) > 30:
        return True
    lower = " " + s.lower() + " "
    for needle in LABEL_REJECT_SUBSTRINGS:
        if needle in lower:
            return True
    return False


def _has_strong_definition(text: str) -> bool:
    """
    True if `text` contains a strong tier-defining noun phrase that should
    OVERRIDE the prose filter. Brownstone-driven: real tier definitions can
    be sentence-shaped ("First, Preferred Return is distributed...").
    """
    if not text:
        return False
    return any(pat.search(text) for pat in STRONG_DEFINITION_PATTERNS)


def _match_tier_type(text: str) -> Optional[str]:
    """
    Pure pattern match — returns the tier_type matched by `text`, or None.
    Does NOT apply prose / strong-def filtering. The two-pass scan in
    `_scan_sheet_for_tiers` decides which labels are eligible.
    """
    if not text:
        return None
    lower = text.lower().strip()
    for tier_type, patterns in TIER_LABEL_PATTERNS:
        for pat in patterns:
            if re.search(pat, lower):
                return tier_type
    return None


def _tier_number_key(label: str) -> Optional[str]:
    """
    Extract a 'Tier N' dedupe key from a label, e.g.
      "Tier 1"               -> "tier_1"
      "Tier 2 Accrual"       -> "tier_2"
      "Tier 1 Distribution"  -> "tier_1"
    Returns None if the label doesn't reference a numbered tier.
    """
    m = TIER_NUMBER_RE.search(label or "")
    if not m:
        return None
    return f"tier_{m.group(1)}"


def _coord(ws: Worksheet, row: int, col: int) -> str:
    """Sheet!Cell reference, e.g. 'Waterfall!C16'."""
    return f"{ws.title}!{ws.cell(row=row, column=col).coordinate}"


def _read_numeric_neighbors(
    ws_values: Worksheet,
    row: int,
    col_start: int,
    width: int = MAX_VALUE_SCAN_COLS,
    stop_on_text: bool = True,
) -> List[Tuple[int, float, str]]:
    """
    Return (col, value, ref) for non-zero numeric cells right of col_start.
    Default stops at first text cell or after `width` cols (narrow scan).
    Pass stop_on_text=False to ignore text and scan the full width.
    """
    out: List[Tuple[int, float, str]] = []
    last_col = min(col_start + width, ws_values.max_column or col_start)
    for c in range(col_start + 1, last_col + 1):
        v = ws_values.cell(row=row, column=c).value
        if isinstance(v, bool):
            continue
        if isinstance(v, (int, float)) and v != 0:
            out.append((c, float(v), _coord(ws_values, row, c)))
        elif stop_on_text and isinstance(v, str) and v.strip():
            # Hit a text label — stop scanning right
            break
    return out


def _detect_compounding(ws_formulas: Worksheet, row: int) -> Optional[str]:
    """
    Inspect formulas in the same row for compounding convention markers.
    Looks at up to first 50 columns.
    """
    last_col = min(50, ws_formulas.max_column or 0)
    for c in range(1, last_col + 1):
        v = ws_formulas.cell(row=row, column=c).value
        if not isinstance(v, str) or not v.startswith("="):
            continue
        f = v.lower()
        if "(1/12)" in f or "^(1/12)" in f:
            return "monthly"
        if "(1/4)" in f or "^(1/4)" in f:
            return "quarterly"
        if "/12" in f and "rate" in f:
            return "monthly"
        if re.search(r"\bmonthly\b", f):
            return "monthly"
        if re.search(r"\bquarterly\b", f):
            return "quarterly"
        if re.search(r"\bannual\b", f):
            return "annual"
    return None


def _split_pair(values: List[float]) -> Optional[Tuple[float, float]]:
    """
    Find a complementary pair in `values` that sums to ~1.0.
    Returns (lp_share, gp_share) with LP > GP by convention, or None.
    """
    for i, a in enumerate(values):
        if not (0 < a < 1):
            continue
        for b in values[i + 1:]:
            if not (0 < b < 1):
                continue
            if abs((a + b) - 1.0) < 0.01:
                lp = max(a, b)
                gp = min(a, b)
                return (lp, gp)
    return None


def _build_tier(
    ws_values: Worksheet,
    ws_formulas: Worksheet,
    row: int,
    col: int,
    tier_type: str,
    label_text: str,
    widen: bool = False,
) -> WaterfallTier:
    """
    Read adjacent cells and build a structured tier object.

    Always uses NARROW scan — values must sit in cells right of the label
    (stopping at first text cell). If they don't, the tier is captured with
    None values and a finding fires asking the user to provide them. We do
    NOT scan distant cells looking for "plausible" values.

    The `widen` arg is preserved for signal but ignored — kept so the
    two-pass scan can still pass intent through, and so we can re-introduce
    a controlled wider scan later if real workbook patterns require it.
    """
    tier = WaterfallTier(
        index=0,  # set by caller
        tier_type=tier_type,
        label=label_text,
        label_cell=_coord(ws_values, row, col),
    )

    candidates = _read_numeric_neighbors(ws_values, row, col)

    if not candidates:
        # Pref tier still records compounding from row formulas (formula
        # reads aren't inference — they're reading the model's spec)
        if tier_type == "pref_accrual":
            tier.pref_compounding = _detect_compounding(ws_formulas, row)
            if tier.pref_compounding:
                tier.source_cells["pref_compounding"] = "(read from row formulas)"
        return tier

    nums = [v for _, v, _ in candidates]
    refs_by_value = {v: ref for _, v, ref in candidates}

    if tier_type == "pref_accrual":
        # First numeric in plausible pref range = pref rate
        for v in nums:
            if PREF_RATE_MIN <= v <= PREF_RATE_MAX:
                tier.pref_rate = v
                tier.hurdle_type = "pref_rate"
                tier.source_cells["pref_rate"] = refs_by_value[v]
                break
        tier.pref_compounding = _detect_compounding(ws_formulas, row)
        if tier.pref_compounding:
            tier.source_cells["pref_compounding"] = "(read from row formulas)"

    elif tier_type == "hurdle_split":
        # Find the hurdle (IRR or EM). Apply band sanity:
        #   IRR_MIN .. IRR_MAX        : clean accept
        #   IRR_MAX .. IRR_HARD_CEILING : accept but flag (likely-result territory)
        #   > IRR_HARD_CEILING         : reject (almost certainly an IRR result cell)
        for v in nums:
            if IRR_MIN <= v <= IRR_HARD_CEILING:
                tier.hurdle_value = v
                tier.hurdle_type = "IRR"
                tier.source_cells["hurdle_value"] = refs_by_value[v]
                break
            if EM_MIN <= v <= EM_MAX:
                tier.hurdle_value = v
                tier.hurdle_type = "EM"
                tier.source_cells["hurdle_value"] = refs_by_value[v]
                break
        # Find LP/GP split
        pair = _split_pair(nums)
        if pair:
            lp, gp = pair
            tier.split_lp_pct = lp
            tier.split_gp_pct = gp
            tier.source_cells["split_lp_pct"] = refs_by_value[lp]
            tier.source_cells["split_gp_pct"] = refs_by_value[gp]

    elif tier_type == "catch_up":
        # 100% / 0% catch-up to GP, OR a partial catch-up split
        for v in nums:
            if abs(v - 1.0) < 0.01:
                tier.split_gp_pct = 1.0
                tier.split_lp_pct = 0.0
                tier.source_cells["catch_up"] = refs_by_value[v]
                break
        if tier.split_gp_pct is None:
            # Try partial catch-up — pair that sums to 1.0 with GP > LP
            pair = _split_pair(nums)
            if pair:
                # In catch-up, GP typically takes the larger share
                lp, gp = pair
                tier.split_lp_pct = min(lp, gp)
                tier.split_gp_pct = max(lp, gp)
                tier.source_cells["split_gp_pct"] = refs_by_value[max(lp, gp)]
                tier.source_cells["split_lp_pct"] = refs_by_value[min(lp, gp)]

    elif tier_type == "residual_split":
        pair = _split_pair(nums)
        if pair:
            lp, gp = pair
            tier.split_lp_pct = lp
            tier.split_gp_pct = gp
            tier.source_cells["split_lp_pct"] = refs_by_value[lp]
            tier.source_cells["split_gp_pct"] = refs_by_value[gp]

    elif tier_type == "return_of_capital":
        for v in nums:
            if abs(v - 1.0) < 0.01:
                tier.split_lp_pct = 1.0
                tier.split_gp_pct = 0.0
                tier.source_cells["roc"] = refs_by_value[v]
                break

    return tier


def _scan_sheet_for_tiers(
    ws_values: Worksheet, ws_formulas: Worksheet
) -> List[WaterfallTier]:
    """
    Walk the sheet looking for tier labels and build structured tier objects.

    Two-pass design (Brownstone calibration v3):

      Pass 1 — clean labels only.
        Skips prose-shaped rows entirely. Captures plain labels like
        "TIER 1", "Pref", "Catch-up". Uses NARROW value scan — accurate
        because the value sits adjacent to the label.

      Pass 2 — prose labels with strong-definition keywords.
        Picks up sentence-shaped tier definitions ("First, Preferred Return
        is distributed...") that pass 1 skipped. Only fires for tier types
        NOT already captured in pass 1, so a clean "Pref" label always
        wins over its narrative twin. Uses WIDENED value scan because
        values won't be in the cell next to the prose label.

    Then:
      - Single-instance tier types (ROC / pref / catch-up / residual) dedupe
        within and across passes.
      - hurdle_split tiers dedupe by extracted "Tier N" number across passes.
      - Post-pass drops hurdle_split tiers that yielded NO hurdle AND NO
        split (no useful data == not a real tier).

    NO autonomous value inference at any stage. Missing values stay missing
    and surface as findings asking the user to provide them.
    """
    tiers: List[WaterfallTier] = []
    seen_unique: set = set()           # tier_types other than hurdle_split
    seen_tier_numbers: set = set()     # hurdle_split dedupe by "Tier N"
    last_row = min(MAX_LABEL_SCAN_ROWS, ws_values.max_row or 0)
    last_col = min(MAX_LABEL_SCAN_COLS, ws_values.max_column or 0)

    def _try_capture(row: int, col: int, label_text: str, widen: bool) -> bool:
        """Attempt to add a tier from this cell. Returns True if added."""
        cleaned_label = label_text.strip()
        tier_type = _match_tier_type(cleaned_label)
        if not tier_type:
            return False

        if tier_type != "hurdle_split" and tier_type in seen_unique:
            return False

        if tier_type == "hurdle_split":
            key = _tier_number_key(cleaned_label)
            if key and key in seen_tier_numbers:
                return False
            if key:
                seen_tier_numbers.add(key)

        tier = _build_tier(
            ws_values, ws_formulas, row, col, tier_type, cleaned_label,
            widen=widen,
        )
        tiers.append(tier)
        if tier_type != "hurdle_split":
            seen_unique.add(tier_type)
        return True

    # ── Pass 1: clean (non-prose) labels, narrow scan ─────────────────────
    for row in range(1, last_row + 1):
        for col in range(1, last_col + 1):
            label_text = ws_values.cell(row=row, column=col).value
            if not isinstance(label_text, str):
                continue
            if _is_prose_label(label_text):
                continue
            if _try_capture(row, col, label_text, widen=False):
                break  # one label per row

    # ── Pass 2: prose labels with strong-definition override, widened scan ─
    for row in range(1, last_row + 1):
        for col in range(1, last_col + 1):
            label_text = ws_values.cell(row=row, column=col).value
            if not isinstance(label_text, str):
                continue
            if not _is_prose_label(label_text):
                continue
            if not _has_strong_definition(label_text):
                continue
            if _try_capture(row, col, label_text, widen=True):
                break  # one label per row

    # NO post-pass filter on empty tiers. A tier whose number is in the
    # workbook but whose hurdle / split values can't be read is itself a
    # finding — dropping it would erase the user's only signal that the
    # tier exists. (Brownstone calibration v3: TIER 2 was being dropped
    # because the model has no extractable hurdle on it; user couldn't
    # see that the middle tier even existed.)

    # Structural sort: tier order in the output matches canonical waterfall
    # order, not capture order. Phase 5 will iterate tiers in array order;
    # we don't want pref sitting at index 4 because pass 2 captured it last.
    canonical_order = {
        "return_of_capital": 1,
        "pref_accrual":      2,
        "catch_up":          3,
        "hurdle_split":      4,   # sub-sorted by Tier N number when present
        "residual_split":    5,
    }
    def _sort_key(t: WaterfallTier) -> Tuple[int, int]:
        primary = canonical_order.get(t.tier_type, 99)
        if t.tier_type == "hurdle_split":
            n = _tier_number_key(t.label)
            secondary = int(n.split("_")[1]) if n else 0
        else:
            secondary = 0
        return (primary, secondary)

    tiers.sort(key=_sort_key)

    # Re-index 1-based AFTER sort
    for i, t in enumerate(tiers, start=1):
        t.index = i
    return tiers


def _detect_overall_type(tiers: List[WaterfallTier]) -> str:
    """
    Roll up tier list to a waterfall_type enum.

    Decision tree (SKILL.md §Phase 4 Step 4 enum):
      - none           : no tiers
      - pref_catchup_split : pref + catch_up + split
      - pref_then_split: pref + split, no catch_up
      - tiered_irr_hurdle : multiple hurdle_split tiers with IRR hurdles, no pref
      - em_hurdle      : EM hurdle present, no IRR
      - hybrid         : both IRR and EM hurdles present
      - custom         : tiers found but pattern unrecognized
    """
    if not tiers:
        return "none"

    types = {t.tier_type for t in tiers}
    has_pref = "pref_accrual" in types
    has_catchup = "catch_up" in types
    has_split = "residual_split" in types or any(
        t.tier_type == "hurdle_split" and t.split_lp_pct is not None for t in tiers
    )
    irr_tiers = [t for t in tiers if t.hurdle_type == "IRR"]
    em_tiers = [t for t in tiers if t.hurdle_type == "EM"]

    if has_pref and has_catchup and has_split:
        return "pref_catchup_split"
    if has_pref and has_split and not has_catchup:
        return "pref_then_split"
    if irr_tiers and em_tiers:
        return "hybrid"
    if irr_tiers and not has_pref:
        return "tiered_irr_hurdle"
    if em_tiers and not irr_tiers and not has_pref:
        return "em_hurdle"
    return "custom"


def _scan_labeled_value(
    ws_values: Worksheet,
    label_patterns: List,
    value_min: float,
    value_max: float,
) -> Tuple[Optional[float], Optional[str]]:
    """
    Whole-sheet scan for a cell whose TEXT LABEL matches one of `label_patterns`
    and which has a numeric value in (value_min, value_max] in an adjacent cell.

    Reads labeled cells — does NOT infer values from arbitrary numerics. The
    label is the authority for what the value means.

    Returns (value, "Sheet!Cell" of the value cell), or (None, None).

    Adjacent = same row to the right, OR same column below. Both are common
    parameter-table layouts.
    """
    last_row = min(MAX_LABEL_SCAN_ROWS, ws_values.max_row or 0)
    last_col = min(20, ws_values.max_column or 0)  # parameter blocks rarely past col 20
    for row in range(1, last_row + 1):
        for col in range(1, last_col + 1):
            v = ws_values.cell(row=row, column=col).value
            if not isinstance(v, str):
                continue
            if not any(p.search(v) for p in label_patterns):
                continue
            # Try same row to the right (up to 5 cells)
            for c in range(col + 1, min(col + 6, last_col + 1)):
                val = ws_values.cell(row=row, column=c).value
                if isinstance(val, (int, float)) and not isinstance(val, bool):
                    if value_min < val <= value_max:
                        return float(val), _coord(ws_values, row, c)
            # Try same column below (up to 3 cells)
            for r in range(row + 1, min(row + 4, last_row + 1)):
                val = ws_values.cell(row=r, column=col).value
                if isinstance(val, (int, float)) and not isinstance(val, bool):
                    if value_min < val <= value_max:
                        return float(val), _coord(ws_values, r, col)
    return None, None


def _scan_labeled_split_pair(
    ws_values: Worksheet,
) -> Tuple[Optional[float], Optional[float], Dict[str, str]]:
    """
    Scan for an LP/GP split pair where labels are explicit. Returns
    (lp_share, gp_share, source_cells_dict) where source_cells maps
    'lp' / 'gp' to Sheet!Cell refs. Returns (None, None, {}) if no
    plausible labeled pair found.

    Strategy: find a cell labeled LP-like with adjacent numeric, find another
    cell labeled GP-like with adjacent numeric, verify the pair sums to ~1.0.
    """
    lp_val, lp_ref = _scan_labeled_value(ws_values, LP_SPLIT_LABEL_PATTERNS, 0.0, 1.0)
    gp_val, gp_ref = _scan_labeled_value(ws_values, GP_SPLIT_LABEL_PATTERNS, 0.0, 1.0)
    if lp_val is None or gp_val is None:
        return None, None, {}
    if abs((lp_val + gp_val) - 1.0) > 0.01:
        return None, None, {}
    return lp_val, gp_val, {"lp": lp_ref or "", "gp": gp_ref or ""}


def _detect_sponsor_coinvest(
    ws_values: Worksheet,
) -> Tuple[Optional[float], Optional[str]]:
    """Scan first MAX_COINVEST_SCAN_ROWS for a sponsor co-invest % label."""
    last_row = min(MAX_COINVEST_SCAN_ROWS, ws_values.max_row or 0)
    last_col = min(4, ws_values.max_column or 0)
    for row in range(1, last_row + 1):
        for col in range(1, last_col + 1):
            v = ws_values.cell(row=row, column=col).value
            if not isinstance(v, str):
                continue
            lower = v.lower()
            if not any(re.search(p, lower) for p in COINVEST_PATTERNS):
                continue
            for c in range(col + 1, col + MAX_VALUE_SCAN_COLS + 1):
                val = ws_values.cell(row=row, column=c).value
                if isinstance(val, (int, float)) and not isinstance(val, bool):
                    if 0 < val < COINVEST_MAX:
                        return float(val), _coord(ws_values, row, c)
    return None, None


def _build_findings(
    waterfall_type: str,
    tiers: List[WaterfallTier],
    pref_rate: Optional[float],
    pref_compounding: Optional[str],
    sheet_name: Optional[str],
) -> List[Dict]:
    """
    Surface anything Phase 5 should know about before replicating.

    Diagnostic categories (Brownstone calibration):
      - classification : structure doesn't match a known pattern
      - data_quality   : a value was extracted but looks suspect
      - extraction_gap : a tier label was found but no value recovered
      - high_band_hurdle : a hurdle was extracted in the suspect 0.20-0.30
                           band (likely an IRR result cell, not a definition)
    """
    findings: List[Dict] = []
    cell = sheet_name or ""

    if waterfall_type == "custom":
        findings.append({
            "severity": "medium",
            "category": "classification",
            "sheet_cell": cell,
            "message": (
                "Waterfall structure does not match a standard pattern. "
                "Phase 5 replication will need bespoke logic — review tier "
                "list against the source sheet before replication."
            ),
        })

    if pref_rate is not None and not (PREF_RATE_MIN <= pref_rate <= PREF_RATE_MAX):
        findings.append({
            "severity": "low",
            "category": "data_quality",
            "sheet_cell": cell,
            "message": (
                f"Pref rate {pref_rate:.2%} falls outside typical "
                f"{PREF_RATE_MIN:.0%}–{PREF_RATE_MAX:.0%} range. Verify "
                "label match — could be a different rate field."
            ),
        })

    # Tier count sanity
    if 0 < len(tiers) < 2 and waterfall_type != "none":
        findings.append({
            "severity": "low",
            "category": "classification",
            "sheet_cell": cell,
            "message": (
                f"Only {len(tiers)} tier detected. Most CRE waterfalls "
                "have 2-4 tiers. Review whether tier labels in the sheet "
                "use non-standard naming."
            ),
        })

    # Pref tier present but no rate extracted
    pref_tier = next((t for t in tiers if t.tier_type == "pref_accrual"), None)
    if pref_tier and pref_tier.pref_rate is None:
        findings.append({
            "severity": "high",
            "category": "extraction_gap",
            "sheet_cell": pref_tier.label_cell,
            "message": (
                "Preferred return tier identified but no pref rate value "
                "could be read from adjacent or downstream cells. Phase 5 "
                "cannot accrue pref without it — user direction required: "
                "what is the pref rate, and which cell holds it?"
            ),
        })

    # Compounding gap on pref tier
    if pref_tier and pref_tier.pref_rate and not pref_compounding:
        findings.append({
            "severity": "medium",
            "category": "extraction_gap",
            "sheet_cell": pref_tier.label_cell,
            "message": (
                "Pref rate detected but compounding convention could not be "
                "read from row formulas. User direction required: monthly, "
                "quarterly, or annual? Phase 5 will not run until specified."
            ),
        })

    # Hurdles in the high band (0.20-0.30) are likely IRR result cells, not defs
    for t in tiers:
        if t.hurdle_type == "IRR" and t.hurdle_value is not None:
            if IRR_MAX < t.hurdle_value <= IRR_HIGH_BAND:
                findings.append({
                    "severity": "low",
                    "category": "high_band_hurdle",
                    "sheet_cell": t.source_cells.get("hurdle_value", t.label_cell),
                    "message": (
                        f"Tier {t.index} hurdle of {t.hurdle_value:.2%} sits "
                        f"in the {IRR_MAX:.0%}–{IRR_HIGH_BAND:.0%} band — "
                        "uncommonly high for a definitional hurdle. Verify "
                        "this isn't an IRR result cell mislabeled as input."
                    ),
                })
            elif IRR_HIGH_BAND < t.hurdle_value <= IRR_HARD_CEILING:
                findings.append({
                    "severity": "medium",
                    "category": "high_band_hurdle",
                    "sheet_cell": t.source_cells.get("hurdle_value", t.label_cell),
                    "message": (
                        f"Tier {t.index} hurdle of {t.hurdle_value:.2%} is "
                        "almost certainly an IRR result cell, not a hurdle "
                        "definition. Manual review required before Phase 5."
                    ),
                })

    # Inflated tier count (after dedupe + post-filter, > 6 is suspicious)
    if len(tiers) > 6:
        findings.append({
            "severity": "medium",
            "category": "classification",
            "sheet_cell": cell,
            "message": (
                f"{len(tiers)} tiers detected — unusually high. Review "
                "whether some captured rows are narrative description, "
                "section restatements, or result cells rather than real "
                "tier definitions."
            ),
        })

    # hurdle_split tier with no hurdle and no split — kept (presence is signal)
    # but flagged HIGH because Phase 5 can't replicate the tier without values.
    for t in tiers:
        if t.tier_type == "hurdle_split" and t.hurdle_value is None and t.split_lp_pct is None:
            findings.append({
                "severity": "high",
                "category": "extraction_gap",
                "sheet_cell": t.label_cell,
                "message": (
                    f"Tier {t.index} ('{t.label}') captured by label but no "
                    "hurdle value or LP/GP split could be read from adjacent "
                    "cells. User direction required: what is the hurdle (or "
                    "is it split-only above the prior hurdle?), and what are "
                    "the LP/GP split percentages? Which cells hold them?"
                ),
            })

    return findings


# ─────────────────────────────────────────────────────────────────────────────
# Public entrypoint
# ─────────────────────────────────────────────────────────────────────────────


def classify(values_wb: Workbook, formulas_wb: Workbook) -> Dict:
    """
    Phase 4 entry point. Returns a WaterfallClassification as a dict.

    Strategy:
      1. Find candidate waterfall sheet(s) by name.
      2. If none -> waterfall_type = "none".
      3. For each candidate, scan first MAX_LABEL_SCAN_ROWS rows for tier
         labels. Pick the candidate that yields the most tiers.
      4. For each tier, read adjacent cells for rate / split / hurdle data.
      5. Roll up tier list to a waterfall_type enum.
      6. Detect sponsor co-invest %.
      7. Build findings list (data quality, classification gaps).
      8. Return structured classification.
    """
    candidates = _find_waterfall_sheets(values_wb)
    if not candidates:
        return WaterfallClassification(
            waterfall_type="none",
            tier_count=0,
            sheet_name=None,
            rationale=(
                "No sheet matches waterfall keywords "
                "(waterfall, promote, distribution, partnership, splits, etc.). "
                "Either this is not a financial model or the sheet uses non-standard naming."
            ),
        ).to_dict()

    # Pick the candidate that yields the most tiers
    best_sheet_name: Optional[str] = None
    best_tiers: List[WaterfallTier] = []
    for sheet_name in candidates:
        if sheet_name not in formulas_wb.sheetnames:
            # Defensive: should not happen since both wbs come from the same file
            continue
        ws_values = values_wb[sheet_name]
        ws_formulas = formulas_wb[sheet_name]
        tiers = _scan_sheet_for_tiers(ws_values, ws_formulas)
        if len(tiers) > len(best_tiers):
            best_sheet_name = sheet_name
            best_tiers = tiers

    if not best_tiers:
        # Found a candidate sheet but couldn't parse tier structure
        first = candidates[0]
        return WaterfallClassification(
            waterfall_type="custom",
            tier_count=0,
            sheet_name=first,
            rationale=(
                f"Found waterfall sheet '{first}' but no recognizable tier "
                "labels (Pref / Catch-up / Hurdle / Split / Residual). "
                "Manual review required before Phase 5 replication."
            ),
            findings=[{
                "severity": "high",
                "category": "classification",
                "message": (
                    "Waterfall sheet identified but tier structure could not "
                    "be parsed automatically."
                ),
            }],
        ).to_dict()

    ws_values = values_wb[best_sheet_name]
    overall_type = _detect_overall_type(best_tiers)
    coinvest, coinvest_ref = _detect_sponsor_coinvest(ws_values)

    # Pull pref data from pref_accrual tier (if present).
    # NO autonomous inference — values are reported as found, or reported as
    # missing via a finding asking the user to provide them.
    pref_rate = None
    pref_compounding = None
    pref_tier = next((t for t in best_tiers if t.tier_type == "pref_accrual"), None)
    if pref_tier:
        pref_rate = pref_tier.pref_rate
        pref_compounding = pref_tier.pref_compounding

    # Secondary labeled-value scan: if pref_tier exists with no rate, search
    # the WHOLE sheet for cells labeled "Pref Rate" / "Preferred Return Rate"
    # / etc. with a numeric value adjacent. Reading a labeled cell — NOT
    # inference. The label is the authority for what the value means.
    if pref_tier and pref_rate is None:
        scanned_rate, scanned_ref = _scan_labeled_value(
            ws_values, PREF_RATE_LABEL_PATTERNS, PREF_RATE_MIN - 0.001, PREF_RATE_MAX
        )
        if scanned_rate is not None:
            pref_rate = scanned_rate
            pref_tier.pref_rate = scanned_rate
            pref_tier.source_cells["pref_rate"] = scanned_ref or "(labeled scan)"

    # Same secondary scan for hurdle_split tiers missing LP/GP splits.
    # Looks for explicitly-labeled LP / GP / Sponsor / Investor split cells
    # somewhere on the sheet that sum to ~1.0.
    for t in best_tiers:
        if t.tier_type != "hurdle_split":
            continue
        if t.split_lp_pct is not None:
            continue
        lp, gp, refs = _scan_labeled_split_pair(ws_values)
        if lp is not None and gp is not None:
            t.split_lp_pct = lp
            t.split_gp_pct = gp
            if refs.get("lp"):
                t.source_cells["split_lp_pct"] = refs["lp"]
            if refs.get("gp"):
                t.source_cells["split_gp_pct"] = refs["gp"]

    # Hurdle type rollup
    hurdle_types = {t.hurdle_type for t in best_tiers if t.hurdle_type}
    hurdle_types.discard("pref_rate")  # pref handled separately
    if hurdle_types == {"IRR"}:
        rollup_hurdle = "IRR"
    elif hurdle_types == {"EM"}:
        rollup_hurdle = "EM"
    elif "IRR" in hurdle_types and "EM" in hurdle_types:
        rollup_hurdle = "IRR_EM"
    else:
        rollup_hurdle = next(iter(hurdle_types), None)

    findings = _build_findings(
        waterfall_type=overall_type,
        tiers=best_tiers,
        pref_rate=pref_rate,
        pref_compounding=pref_compounding,
        sheet_name=best_sheet_name,
    )

    source_cells: Dict[str, str] = {}
    if coinvest_ref:
        source_cells["sponsor_coinvest"] = coinvest_ref

    return WaterfallClassification(
        waterfall_type=overall_type,
        tier_count=len(best_tiers),
        sheet_name=best_sheet_name,
        tiers=best_tiers,
        pref_rate=pref_rate,
        pref_compounding=pref_compounding,
        hurdle_type=rollup_hurdle,
        sponsor_coinvest_pct=coinvest,
        source_cells=source_cells,
        findings=findings,
        rationale=(
            f"Detected {len(best_tiers)}-tier '{overall_type}' waterfall on "
            f"sheet '{best_sheet_name}'."
        ),
    ).to_dict()
