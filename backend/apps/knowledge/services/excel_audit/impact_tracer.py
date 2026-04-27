"""
Phase 2f — Downstream Impact Tracer.

After formula_integrity identifies error cells and broken refs,
this module answers: "Do any of these errors reach headline output
cells (IRR, equity multiple, net cash flow, DSCR, etc.)?"

Approach (option 2d, label-proximity + future LLM sanity check):
  1. detect_sinks()  — scan workbook for cells whose adjacent label
                        matches a known headline-output keyword set.
  2. trace()         — build a forward dependency graph from formulas,
                        BFS from each error cell, report which sinks
                        are reachable.
  3. annotate()      — enrich formula_integrity findings with
                        reaches_headline / sink_hits metadata.

The result lets Landscaper say "51 errors detected, 0 affect headline
outputs" vs "51 errors, 3 cascade into IRR calculation."

Wide sink set (option 1d): IRR, equity multiple, NPV, net CF,
interim period CFs, DSCR, coverage, CoC, distributions, promote.
"""

import re
from collections import deque
from typing import Dict, List, Optional, Set, Tuple

from openpyxl.utils.cell import (
    column_index_from_string,
    coordinate_from_string,
    get_column_letter,
)
from openpyxl.workbook.workbook import Workbook


# ─────────────────────────────────────────────────────────────────────────────
# Headline output label keywords (case-insensitive match)
# ─────────────────────────────────────────────────────────────────────────────

SINK_LABELS = {
    # Return metrics
    "irr", "xirr", "levered irr", "unlevered irr", "deal irr",
    "investor irr", "sponsor irr", "lp irr", "gp irr",
    "equity multiple", "equity mult", "moic",
    "npv", "net present value",
    # Cash flow headline
    "net cash flow", "net cf", "total cash flow", "total cf",
    "levered cash flow", "levered cf", "unlevered cash flow", "unlevered cf",
    "free cash flow", "fcf",
    "cash-on-cash", "cash on cash", "coc", "coc return",
    "total return", "total profit", "net profit", "gain on sale",
    # Debt / coverage
    "dscr", "debt service coverage", "coverage ratio",
    "debt yield", "ltv", "loan to value",
    # Distribution / promote
    "total distributions", "total distribution",
    "promote", "carried interest", "carry",
    "pref return", "preferred return",
    # Interim period CFs (wide net per 1d)
    "annual cash flow", "monthly cash flow", "quarterly cash flow",
    "year 1 cf", "year 2 cf", "year 3 cf", "year 4 cf", "year 5 cf",
    "period cash flow", "periodic cf",
    # NOI
    "net operating income", "noi", "stabilized noi",
}

# Compiled for fast substring check
_SINK_LABEL_LIST = sorted(SINK_LABELS, key=len, reverse=True)  # longest first


def _normalize(text: str) -> str:
    """Lowercase, collapse whitespace, strip punctuation edges."""
    return re.sub(r"\s+", " ", str(text).strip().lower().rstrip(":"))


# Short labels that need word-boundary matching to avoid false positives
_SHORT_LABELS = {"noi", "coc", "fcf", "ltv"}


def _is_sink_label(text: str) -> Optional[str]:
    """Return matched sink label if text contains one, else None."""
    normed = _normalize(text)
    if not normed or len(normed) > 80:
        return None
    for label in _SINK_LABEL_LIST:
        if label not in normed:
            continue
        # Short labels (<=3 chars) require word boundary to avoid
        # matching "item", "estimate", "document" etc.
        if len(label) <= 3 or label in _SHORT_LABELS:
            if not re.search(r"(?:^|[\s/\-_(])" + re.escape(label) + r"(?:$|[\s/\-_):])", normed):
                continue
        return label
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Step 1: Detect headline output sinks via label proximity
# ─────────────────────────────────────────────────────────────────────────────

def detect_sinks(values_wb: Workbook, formulas_wb: Workbook) -> List[Dict]:
    """
    Scan for cells that are likely headline outputs.

    Strategy: for each cell containing a numeric value or formula,
    check the cell immediately to its left (same row) and the cell
    immediately above (same column). If either contains a string
    matching a sink label, this cell is a candidate sink.

    Returns list of {"sheet", "cell", "label", "value"}.
    """
    sinks = []
    seen: Set[str] = set()

    for vi, sheet_v in enumerate(values_wb.worksheets):
        sheet_name = sheet_v.title
        # Also get the formula sheet to check if it's a formula cell
        sheet_f = None
        if vi < len(formulas_wb.worksheets):
            sheet_f = formulas_wb.worksheets[vi]

        max_row = sheet_v.max_row or 0
        max_col = sheet_v.max_column or 0
        if max_row < 1 or max_col < 1:
            continue

        for row in sheet_v.iter_rows(min_row=1, max_row=max_row,
                                      min_col=1, max_col=max_col,
                                      values_only=False):
            for cell in row:
                val = cell.value
                # Only consider numeric values or formula results
                if not isinstance(val, (int, float)):
                    continue
                r, c = cell.row, cell.column

                # Check left neighbor for label
                matched_label = None
                if c > 1:
                    left = sheet_v.cell(row=r, column=c - 1).value
                    if isinstance(left, str):
                        matched_label = _is_sink_label(left)

                # Check above neighbor for label
                if not matched_label and r > 1:
                    above = sheet_v.cell(row=r - 1, column=c).value
                    if isinstance(above, str):
                        matched_label = _is_sink_label(above)

                # Check two-left (label | blank | value pattern)
                if not matched_label and c > 2:
                    mid = sheet_v.cell(row=r, column=c - 1).value
                    if mid is None or (isinstance(mid, str) and mid.strip() == ""):
                        two_left = sheet_v.cell(row=r, column=c - 2).value
                        if isinstance(two_left, str):
                            matched_label = _is_sink_label(two_left)

                if matched_label:
                    ref = f"{sheet_name}!{cell.coordinate}"
                    if ref not in seen:
                        seen.add(ref)
                        # Check if this is a formula cell (stronger signal)
                        is_formula = False
                        if sheet_f:
                            fc = sheet_f.cell(row=r, column=c).value
                            is_formula = isinstance(fc, str) and fc.startswith("=")
                        sinks.append({
                            "sheet": sheet_name,
                            "cell": cell.coordinate,
                            "ref": ref,
                            "label": matched_label,
                            "value": val,
                            "is_formula": is_formula,
                        })
    return _filter_dead_sinks(sinks)


def _filter_dead_sinks(sinks: List[Dict]) -> List[Dict]:
    """
    Drop sinks whose cached value is #N/A, #REF!, #VALUE!, etc. These are
    "dead outputs" — the workbook itself shows the metric is broken, which
    almost always means the user has abandoned that calculation path (e.g.,
    refinancing IRR cells in a deal that doesn't contemplate a refi). Such
    sinks should NOT count as "headline outputs errors propagate to" — that
    creates false-positive ALL_REACH_OUTPUTS verdicts for errors that only
    feed the dormant branch.

    Also drops sinks whose value is None (cell never computed) for the same
    reason — if the user hasn't activated the calculation, errors reaching
    it are not affecting reported numbers.

    Numeric-zero sinks are KEPT — zero is a legitimate value (e.g., year-1
    cash flow can legitimately be zero). Only error-string and None values
    indicate dead state.
    """
    error_strings = {"#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#NUM!", "#NULL!"}
    kept = []
    for s in sinks:
        v = s.get("value")
        if v is None:
            continue
        if isinstance(v, str) and v.strip().upper() in error_strings:
            continue
        kept.append(s)
    return kept


# ─────────────────────────────────────────────────────────────────────────────
# Step 2: Build dependency graph + BFS trace
# ─────────────────────────────────────────────────────────────────────────────

# Parse cell references from formula text
_SINGLE_CELL_RE = re.compile(
    r"(?:(?P<sheet>'[^']+'|[A-Za-z0-9_]+)!)?\$?(?P<col>[A-Z]+)\$?(?P<row>\d+)"
    r"(?![\d:])"  # negative lookahead: not part of a range
)
_RANGE_RE = re.compile(
    r"(?:(?P<sheet>'[^']+'|[A-Za-z0-9_]+)!)?"
    r"\$?(?P<col1>[A-Z]+)\$?(?P<row1>\d+)"
    r":"
    r"\$?(?P<col2>[A-Z]+)\$?(?P<row2>\d+)"
)


def _parse_refs_from_formula(formula: str, current_sheet: str) -> Set[str]:
    """
    Extract all cell references from a formula string.
    Returns set of "Sheet!A1" normalized refs.
    """
    refs: Set[str] = set()
    if not formula or not formula.startswith("="):
        return refs

    # First extract ranges (expand to individual cells, capped)
    for m in _RANGE_RE.finditer(formula):
        sheet = (m.group("sheet") or current_sheet).strip("'")
        try:
            c1 = column_index_from_string(m.group("col1"))
            c2 = column_index_from_string(m.group("col2"))
            r1 = int(m.group("row1"))
            r2 = int(m.group("row2"))
        except (ValueError, AttributeError):
            continue
        # Cap expansion to avoid blowing up on huge ranges
        num_cells = (abs(c2 - c1) + 1) * (abs(r2 - r1) + 1)
        if num_cells > 500:
            # For large ranges, just record corners
            refs.add(f"{sheet}!{m.group('col1')}{m.group('row1')}")
            refs.add(f"{sheet}!{m.group('col2')}{m.group('row2')}")
            continue
        for ri in range(min(r1, r2), max(r1, r2) + 1):
            for ci in range(min(c1, c2), max(c1, c2) + 1):
                refs.add(f"{sheet}!{get_column_letter(ci)}{ri}")

    # Then single cell refs (excluding those already captured as part of ranges)
    formula_no_ranges = _RANGE_RE.sub("", formula)
    for m in _SINGLE_CELL_RE.finditer(formula_no_ranges):
        sheet = (m.group("sheet") or current_sheet).strip("'")
        col = m.group("col")
        row = m.group("row")
        refs.add(f"{sheet}!{col}{row}")

    return refs


def _build_dependency_graph(formulas_wb: Workbook) -> Dict[str, Set[str]]:
    """
    Build a forward dependency graph: precedent → set of dependents.

    If cell B2 contains =A1+A2, then:
      graph["Sheet!A1"] contains "Sheet!B2"
      graph["Sheet!A2"] contains "Sheet!B2"

    This lets us BFS forward from an error cell to find what it feeds.
    """
    # precedent → set of dependents
    graph: Dict[str, Set[str]] = {}

    for sheet in formulas_wb.worksheets:
        sheet_name = sheet.title
        for row in sheet.iter_rows(values_only=False):
            for cell in row:
                val = cell.value
                if not (isinstance(val, str) and val.startswith("=")):
                    continue
                dependent = f"{sheet_name}!{cell.coordinate}"
                precedents = _parse_refs_from_formula(val, sheet_name)
                for prec in precedents:
                    if prec not in graph:
                        graph[prec] = set()
                    graph[prec].add(dependent)

    return graph


def trace(
    formulas_wb: Workbook,
    error_refs: List[str],
    sink_refs: Set[str],
    max_depth: int = 50,
) -> Dict[str, Dict]:
    """
    BFS forward from each error cell through the dependency graph.
    Returns {error_ref: {"reaches": bool, "sinks_hit": [...], "depth": n}}.

    max_depth caps traversal to prevent runaway on deeply nested models.
    """
    graph = _build_dependency_graph(formulas_wb)
    results: Dict[str, Dict] = {}

    for err_ref in error_refs:
        visited: Set[str] = set()
        queue = deque([(err_ref, 0)])
        sinks_hit: List[str] = []
        max_reached = 0

        while queue:
            current, depth = queue.popleft()
            if depth > max_depth:
                continue
            if current in visited:
                continue
            visited.add(current)
            max_reached = max(max_reached, depth)

            if current in sink_refs and current != err_ref:
                sinks_hit.append(current)
                # Don't stop — keep tracing to find all reachable sinks

            for dependent in graph.get(current, set()):
                if dependent not in visited:
                    queue.append((dependent, depth + 1))

        results[err_ref] = {
            "reaches": len(sinks_hit) > 0,
            "sinks_hit": sinks_hit,
            "depth": max_reached,
            "cells_traversed": len(visited),
        }

    return results


# ─────────────────────────────────────────────────────────────────────────────
# Step 3: Annotate findings with impact data
# ─────────────────────────────────────────────────────────────────────────────

def annotate(findings: List[Dict], impact_map: Dict[str, Dict]) -> List[Dict]:
    """
    Enrich formula_integrity findings with downstream impact info.
    Mutates findings in-place and returns them.

    Severity downgrade rule:
      - When trace is conclusive AND nothing downstream is hit, the finding
        is cosmetic (range/value error doesn't reach any headline output).
        Severity is downgraded to 'low' and `is_cosmetic` is set True.
        Original severity preserved as `severity_raw` for audit trail.
      - When trace is conclusive AND a sink IS hit, severity is preserved
        and `is_cosmetic` is False — the error feeds a headline output.
      - When trace is inconclusive (reaches_headline=None), severity is
        preserved and `is_cosmetic` is None — no claim either way.

    The `effective_severity` post-trace lets the report differentiate
    "this 2e truncation feeds IRR" (high, real) from "this 2e truncation
    is in a quarantined cosmetic block" (low, no action needed).
    """
    for f in findings:
        ref = f.get("ref", "")
        impact = impact_map.get(ref)
        original_severity = f.get("severity")
        if impact:
            f["reaches_headline"] = impact["reaches"]
            f["sinks_hit"] = impact["sinks_hit"]
            if not impact["reaches"]:
                f["is_cosmetic"] = True
                if original_severity and original_severity != "low":
                    f["severity_raw"] = original_severity
                    f["severity"] = "low"
            else:
                f["is_cosmetic"] = False
        else:
            # Finding wasn't traced (e.g., no ref, or check not in TRACEABLE_CHECKS)
            f["reaches_headline"] = None
            f["sinks_hit"] = []
            f["is_cosmetic"] = None
    return findings


# ─────────────────────────────────────────────────────────────────────────────
# Convenience: run full impact analysis
# ─────────────────────────────────────────────────────────────────────────────

def run_impact_analysis(
    values_wb: Workbook,
    formulas_wb: Workbook,
    findings: List[Dict],
) -> Dict:
    """
    End-to-end: detect sinks, trace errors, annotate findings.

    Returns {
        "sinks_detected": [...],
        "impact_map": {ref: {reaches, sinks_hit, depth}},
        "summary": {
            "errors_reaching_headline": n,
            "errors_quarantined": m,
            "total_traced": t,
        },
    }
    """
    sinks = detect_sinks(values_wb, formulas_wb)
    sink_ref_set = {s["ref"] for s in sinks}

    # Trace ALL formula-integrity findings that have a cell ref. Previously
    # this filter only included 2a (error cells) and 2b (broken refs), which
    # left 2c (hardcoded overrides) and 2e (range-consistency / truncated
    # SUMs) untagged — every finding looked equally critical even when it
    # didn't feed any headline output. Wider trace lets the report mark
    # cosmetic findings explicitly.
    TRACEABLE_CHECKS = (
        "2a_error_cell",
        "2b_broken_ref",
        "2c_hardcoded_override",
        "2e_range_consistency",
    )
    error_refs = [
        f["ref"] for f in findings
        if f.get("check") in TRACEABLE_CHECKS and f.get("ref")
    ]

    if not error_refs or not sink_ref_set:
        # Nothing to trace — annotate everything as unknown
        for f in findings:
            f["reaches_headline"] = None
            f["sinks_hit"] = []
        return {
            "sinks_detected": sinks,
            "impact_map": {},
            "summary": {
                "errors_reaching_headline": 0,
                "errors_quarantined": 0 if not error_refs else len(error_refs),
                "total_traced": 0,
                "note": "no sinks detected" if not sink_ref_set else "no error cells to trace",
                "verdict": "UNTRACED",
                "verdict_text": (
                    "Could not trace impact — no headline-output cells detected in workbook. "
                    "Findings cannot be classified as cosmetic vs. impactful without sinks to trace to."
                ),
            },
        }

    impact_map = trace(formulas_wb, error_refs, sink_ref_set)
    annotate(findings, impact_map)

    reaching = sum(1 for v in impact_map.values() if v["reaches"])
    quarantined = sum(1 for v in impact_map.values() if not v["reaches"])

    # Verdict — prominent + impossible-to-misread summary so the LLM (and any
    # downstream summarizer) leads with impact status rather than raw error
    # counts. Brownstone calibration: a 41-cell quarantined error block was
    # being reported as "critical, cascades to 41 cells" because the model
    # was reading raw `cells_traversed` instead of the impact-summary fields.
    if reaching == 0 and quarantined > 0:
        verdict = "ALL_QUARANTINED"
        verdict_text = (
            f"All {quarantined} errors are in quarantined cells that do NOT feed any "
            f"headline output (IRR, equity multiple, DSCR, net cash flow, NOI, "
            f"distributions). The errors are cosmetic from a results-integrity standpoint. "
            f"They should still be fixed for cleanliness, but the model's reported "
            f"returns and cash flows are not affected by them."
        )
    elif reaching > 0 and quarantined == 0:
        verdict = "ALL_REACH_OUTPUTS"
        verdict_text = (
            f"All {reaching} errors propagate to headline outputs (IRR, EM, DSCR, "
            f"net CF, etc.). The model's reported numbers cannot be trusted until "
            f"these are fixed."
        )
    elif reaching > 0 and quarantined > 0:
        verdict = "MIXED"
        verdict_text = (
            f"{reaching} of {reaching + quarantined} errors propagate to headline "
            f"outputs and require fixes before trusting reported numbers. The other "
            f"{quarantined} are quarantined / cosmetic and don't affect results."
        )
    else:
        verdict = "NONE"
        verdict_text = "No errors found."

    return {
        "sinks_detected": sinks,
        "impact_map": impact_map,
        "summary": {
            "errors_reaching_headline": reaching,
            "errors_quarantined": quarantined,
            "total_traced": len(error_refs),
            "verdict": verdict,
            "verdict_text": verdict_text,
        },
    }
