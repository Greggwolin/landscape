"""
Phase 2 — Formula Integrity checks.

Checks:
  2a. Error cells        - any cell whose computed value is an Excel error
                           (#REF!, #VALUE!, #DIV/0!, #N/A, #NAME?, #NUM!, #NULL!)
  2b. Broken references  - formulas containing #REF! in the formula text
  2c. Hardcoded overrides - numeric literals in cells that look like they should
                            be formulas (heuristic: row has formulas on both sides)
  2e. Range consistency   - SUM / XIRR / NPV / AVERAGE formulas across adjacent
                            columns that reference *different* row spans.
                            Catches truncated SUMs that miss newly-added rows.

Circular dependency detection (2d) deferred — openpyxl does not resolve the
dependency graph natively, and the workbook's recomputed values already
reflect circular behavior when data_only=True was loaded.

Findings returned as a list of dicts with Sheet!Cell refs. Severity:
  high   - 2a error cells, 2b broken refs, 2e inconsistent ranges
  medium - 2c hardcoded overrides
"""

import re
from typing import Dict, List

from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter
from openpyxl.workbook.workbook import Workbook


EXCEL_ERRORS = {"#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NAME?", "#NUM!", "#NULL!"}

# Capture =FUNC(range) — range may be A1:A99, Sheet1!A1:A99, $A$1:$A$99, etc.
RANGE_FUNC_RE = re.compile(
    r"(?P<func>SUM|SUMPRODUCT|AVERAGE|AVG|XIRR|XNPV|NPV|IRR|COUNT|COUNTA|MAX|MIN)\s*\(",
    re.IGNORECASE,
)
CELL_RANGE_RE = re.compile(
    r"(?:(?P<sheet>'[^']+'|[A-Za-z0-9_]+)!)?\$?(?P<col1>[A-Z]+)\$?(?P<row1>\d+):\$?(?P<col2>[A-Z]+)\$?(?P<row2>\d+)"
)


def _cell_ref(sheet_name: str, coordinate: str) -> str:
    # Always quote sheet names that contain spaces or leading digits
    needs_quote = any(c in sheet_name for c in " -") or (sheet_name and sheet_name[0].isdigit())
    q = f"'{sheet_name}'" if needs_quote else sheet_name
    return f"{q}!{coordinate}"


def _check_errors(values_wb: Workbook) -> List[Dict]:
    findings = []
    for sheet in values_wb.worksheets:
        for row in sheet.iter_rows(values_only=False):
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val in EXCEL_ERRORS:
                    findings.append(
                        {
                            "check": "2a_error_cell",
                            "severity": "high",
                            "ref": _cell_ref(sheet.title, cell.coordinate),
                            "detail": val,
                        }
                    )
    return findings


def _check_broken_refs(formulas_wb: Workbook) -> List[Dict]:
    findings = []
    for sheet in formulas_wb.worksheets:
        for row in sheet.iter_rows(values_only=False):
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val.startswith("=") and "#REF!" in val:
                    findings.append(
                        {
                            "check": "2b_broken_ref",
                            "severity": "high",
                            "ref": _cell_ref(sheet.title, cell.coordinate),
                            "detail": val[:120],
                        }
                    )
    return findings


def _check_hardcoded_overrides(formulas_wb: Workbook) -> List[Dict]:
    """
    Heuristic: within a row that already contains formulas, any literal
    numeric cell in a column position where adjacent rows have formulas
    is suspect. Limited to cells flanked by formulas on left AND right OR
    above AND below.
    """
    findings = []
    for sheet in formulas_wb.worksheets:
        max_row = sheet.max_row or 0
        max_col = sheet.max_column or 0
        if max_row < 3 or max_col < 3:
            continue
        # Pre-compute formula map
        is_formula = {}
        for row in sheet.iter_rows(values_only=False):
            for cell in row:
                v = cell.value
                is_formula[(cell.row, cell.column)] = isinstance(v, str) and v.startswith("=")

        for (r, c), is_f in is_formula.items():
            if is_f:
                continue
            cell = sheet.cell(row=r, column=c)
            if not isinstance(cell.value, (int, float)):
                continue
            left = is_formula.get((r, c - 1), False)
            right = is_formula.get((r, c + 1), False)
            up = is_formula.get((r - 1, c), False)
            down = is_formula.get((r + 1, c), False)
            if (left and right) or (up and down):
                findings.append(
                    {
                        "check": "2c_hardcoded_override",
                        "severity": "medium",
                        "ref": _cell_ref(sheet.title, cell.coordinate),
                        "detail": f"value={cell.value} surrounded by formulas",
                    }
                )
                if len(findings) > 100:
                    return findings  # cap to keep audit bounded
    return findings


def _check_range_consistency(formulas_wb: Workbook) -> List[Dict]:
    """
    Check 2e — the killer check.

    For each row of cells containing range-based aggregate formulas
    (SUM/XIRR/NPV/etc.) in adjacent columns, the referenced row spans
    should match. If column B sums A2:A99 and column C sums A2:A50,
    column C is probably truncated.

    We group formulas by (sheet, row) and compare row-span of the first
    range reference in each. Any mismatch in the same row → flag.
    """
    findings = []
    for sheet in formulas_wb.worksheets:
        # Gather formulas by row
        by_row: Dict[int, List[Dict]] = {}
        for row in sheet.iter_rows(values_only=False):
            for cell in row:
                v = cell.value
                if not (isinstance(v, str) and v.startswith("=")):
                    continue
                # Only track formulas that contain a range-based function
                if not RANGE_FUNC_RE.search(v):
                    continue
                m = CELL_RANGE_RE.search(v)
                if not m:
                    continue
                r1 = int(m.group("row1"))
                r2 = int(m.group("row2"))
                by_row.setdefault(cell.row, []).append(
                    {
                        "coord": cell.coordinate,
                        "formula": v,
                        "span": (r1, r2),
                    }
                )

        for row_num, items in by_row.items():
            if len(items) < 2:
                continue
            spans = {it["span"] for it in items}
            if len(spans) > 1:
                # Inconsistent row-spans in the same row — flag each cell
                for it in items:
                    findings.append(
                        {
                            "check": "2e_range_consistency",
                            "severity": "high",
                            "ref": _cell_ref(sheet.title, it["coord"]),
                            "detail": (
                                f"row-span {it['span']} differs from siblings in row "
                                f"{row_num}: {sorted(spans)}"
                            ),
                        }
                    )
    return findings


def check(values_wb: Workbook, formulas_wb: Workbook) -> Dict:
    findings: List[Dict] = []
    findings.extend(_check_errors(values_wb))
    findings.extend(_check_broken_refs(formulas_wb))
    findings.extend(_check_hardcoded_overrides(formulas_wb))
    findings.extend(_check_range_consistency(formulas_wb))

    by_check: Dict[str, int] = {}
    by_severity: Dict[str, int] = {"high": 0, "medium": 0, "low": 0}
    for f in findings:
        by_check[f["check"]] = by_check.get(f["check"], 0) + 1
        by_severity[f["severity"]] = by_severity.get(f["severity"], 0) + 1

    return {
        "findings": findings,
        "finding_count": len(findings),
        "by_check": by_check,
        "by_severity": by_severity,
    }
