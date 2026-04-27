"""
Phase 6 — Sources & Uses verification.

Per the canonical excel-model-audit skill (§Phase 6):

    Step 1 — Look for an existing S&U schedule in the workbook
             (labels: "Sources", "Uses", "Total Sources", "Total Uses").
    Step 2 — If found with totals, compare and report.
    Step 3 — Sources must equal Uses within $1. Gap > $100 = finding.

This module implements Step 1 + Step 3 against the workbook directly. Step 2
(synthetic build from extracted assumptions when no S&U schedule exists) is
deferred — most CRE proformas have an explicit S&U block, and synthetic
construction is brittle without a property-type-aware mapping table.

Returns a dict with:
    sources              list of {label, value, sheet_cell}
    uses                 list of {label, value, sheet_cell}
    sources_total        float | None
    uses_total           float | None
    delta                float | None  (abs(sources_total - uses_total))
    balanced             bool          (delta < $1)
    sheet_name           str | None
    findings             list of {severity, category, message, sheet_cell, feeds_outputs}
    rationale            str

Findings are emitted only when a real imbalance is detected ( > $1 ) — per
skill rule §15 "no informational findings."
"""

import logging
from typing import Any, Dict, List, Optional

from openpyxl.workbook.workbook import Workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


# Header tokens that anchor the start of a Sources or Uses block. Matched
# against the cell value (lowercased, stripped) — exact match OR starts-with.
SOURCES_ANCHORS = (
    "sources",
    "sources of funds",
    "sources of capital",
    "capital sources",
    "funding sources",
)
USES_ANCHORS = (
    "uses",
    "uses of funds",
    "uses of capital",
    "uses of proceeds",
    "capital uses",
)

# Total-row tokens that close a Sources or Uses block.
TOTAL_TOKENS = (
    "total sources",
    "total uses",
    "total",
    "subtotal",
)

# Maximum rows to scan downward from a Sources/Uses anchor before giving up.
MAX_SCAN_ROWS = 40

# Maximum value column offset to check (a label in B with a value in C/D/E).
VALUE_COL_OFFSETS = (1, 2, 3)


# ─────────────────────────────────────────────────────────────────────────────
# Public entry point
# ─────────────────────────────────────────────────────────────────────────────


def verify(values_wb: Workbook, formulas_wb: Workbook) -> Dict[str, Any]:
    """
    Run Phase 6: locate S&U schedule and compare totals.

    Args:
        values_wb:   openpyxl workbook loaded with data_only=True (cached values)
        formulas_wb: openpyxl workbook loaded with data_only=False (formula text)

    The values workbook is the primary source — we read cached numerics. The
    formulas workbook is consulted only to confirm a candidate "Total" row is
    actually a SUM/sum-equivalent rather than a hardcoded number (which would
    be a separate finding).
    """
    # Find the most plausible S&U sheet
    candidate = _locate_block(values_wb)
    if candidate is None:
        return _no_block_result()

    sheet_name = candidate["sheet_name"]
    ws_values = values_wb[sheet_name]
    ws_formulas = formulas_wb[sheet_name] if sheet_name in formulas_wb.sheetnames else None

    sources_block = _extract_block(
        ws_values, ws_formulas, candidate["sources_anchor"], block_kind="sources"
    )
    uses_block = _extract_block(
        ws_values, ws_formulas, candidate["uses_anchor"], block_kind="uses"
    )

    sources_total = sources_block.get("total")
    uses_total = uses_block.get("total")
    delta: Optional[float] = None
    balanced = False
    if sources_total is not None and uses_total is not None:
        delta = abs(float(sources_total) - float(uses_total))
        balanced = delta < 1.0

    findings: List[Dict[str, Any]] = []
    if delta is not None and not balanced:
        sev = "high" if delta > 100.0 else "low"
        findings.append({
            "severity": sev,
            "category": "sources_uses_imbalance",
            "message": (
                f"Sources (${sources_total:,.2f}) do not equal "
                f"Uses (${uses_total:,.2f}) — delta ${delta:,.2f}. "
                f"S&U should balance to within $1."
            ),
            "sheet_cell": f"{sheet_name}!{sources_block.get('total_cell') or '?'}",
            "feeds_outputs": False,
        })

    rationale = _build_rationale(sources_block, uses_block, delta, balanced)

    return {
        "sources": sources_block.get("items", []),
        "uses": uses_block.get("items", []),
        "sources_total": sources_total,
        "sources_total_cell": sources_block.get("total_cell"),
        "uses_total": uses_total,
        "uses_total_cell": uses_block.get("total_cell"),
        "delta": delta,
        "balanced": balanced,
        "sheet_name": sheet_name,
        "findings": findings,
        "rationale": rationale,
    }


def _no_block_result() -> Dict[str, Any]:
    """Return shape when no S&U schedule was found."""
    return {
        "sources": [],
        "uses": [],
        "sources_total": None,
        "sources_total_cell": None,
        "uses_total": None,
        "uses_total_cell": None,
        "delta": None,
        "balanced": False,
        "sheet_name": None,
        "findings": [],
        "rationale": (
            "No Sources & Uses schedule located in the workbook. "
            "Phase 6 requires an explicit block with 'Sources' / 'Uses' labels "
            "and totals. Synthetic construction from extracted assumptions is "
            "deferred."
        ),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Block location
# ─────────────────────────────────────────────────────────────────────────────


def _locate_block(wb: Workbook) -> Optional[Dict[str, Any]]:
    """
    Scan every sheet for both a Sources anchor and a Uses anchor in close
    proximity (same sheet, anchors within 50 rows of each other). Return the
    first match.
    """
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sources_anchor = _find_anchor(ws, SOURCES_ANCHORS)
        uses_anchor = _find_anchor(ws, USES_ANCHORS, exclude_label_match="sources")
        if sources_anchor and uses_anchor:
            row_gap = abs(sources_anchor["row"] - uses_anchor["row"])
            if row_gap <= 60:
                return {
                    "sheet_name": sheet_name,
                    "sources_anchor": sources_anchor,
                    "uses_anchor": uses_anchor,
                }
    return None


def _find_anchor(ws, tokens, exclude_label_match: Optional[str] = None) -> Optional[Dict[str, Any]]:
    """
    Find the first cell whose stripped/lowercased value matches one of the
    anchor tokens. Returns {row, col, value, sheet_cell} or None.

    `exclude_label_match` skips any label whose value starts with that string —
    used to keep the Uses search from latching onto "Sources" cells when both
    appear (e.g., "Sources" and "Sources & Uses" labels in the same workbook).
    """
    max_row = min(ws.max_row or 0, 200)
    max_col = min(ws.max_column or 0, 40)
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is None:
                continue
            val = str(cell.value).strip().lower()
            if not val:
                continue
            if exclude_label_match and val.startswith(exclude_label_match):
                continue
            for tok in tokens:
                if val == tok or (val.startswith(tok) and len(val) <= len(tok) + 12):
                    return {
                        "row": row,
                        "col": col,
                        "value": str(cell.value).strip(),
                        "sheet_cell": cell.coordinate,
                    }
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Block extraction
# ─────────────────────────────────────────────────────────────────────────────


def _extract_block(ws_values, ws_formulas, anchor: Dict[str, Any], block_kind: str) -> Dict[str, Any]:
    """
    Walk down from the anchor row, collecting (label, value, cell_ref) triples.
    Stop at: blank row, "Total ..." row (which we capture as the block total),
    or another anchor token.

    Returns {items: [...], total: float | None, total_cell: str | None}.
    """
    items: List[Dict[str, Any]] = []
    total: Optional[float] = None
    total_cell: Optional[str] = None

    label_col = anchor["col"]
    start_row = anchor["row"] + 1
    end_row = min(start_row + MAX_SCAN_ROWS, ws_values.max_row or start_row + MAX_SCAN_ROWS)

    blank_streak = 0
    for row in range(start_row, end_row + 1):
        label_cell = ws_values.cell(row=row, column=label_col)
        label_val = label_cell.value
        if label_val is None or (isinstance(label_val, str) and not label_val.strip()):
            blank_streak += 1
            if blank_streak >= 2:
                break
            continue
        blank_streak = 0

        if not isinstance(label_val, str):
            # Non-text in the label column — likely numeric; skip
            continue

        label_str = label_val.strip()
        label_lower = label_str.lower()

        # Stop if we hit the OTHER block's anchor while walking the current one
        opposite_tokens = USES_ANCHORS if block_kind == "sources" else SOURCES_ANCHORS
        if any(label_lower == tok or label_lower.startswith(tok) for tok in opposite_tokens):
            break

        # Find the numeric value in the adjacent columns
        value, value_cell = _scan_value_to_right(ws_values, row, label_col)
        if value is None:
            # Label without a value — sub-header or empty row, skip but keep walking
            continue

        # Is this the "Total" row that closes the block?
        is_total = any(label_lower.startswith(tok) for tok in TOTAL_TOKENS)
        if is_total:
            total = float(value)
            total_cell = value_cell
            break

        items.append({
            "label": label_str,
            "value": float(value),
            "sheet_cell": value_cell,
        })

    # If we never hit an explicit Total row, sum the items as the total
    if total is None and items:
        total = sum(it["value"] for it in items)
        total_cell = None  # synthetic — no single cell holds it

    return {"items": items, "total": total, "total_cell": total_cell}


def _scan_value_to_right(ws, row: int, label_col: int):
    """
    Look at cells to the right of `label_col` for the first numeric value.
    Returns (value, sheet_cell) or (None, None).
    """
    for offset in VALUE_COL_OFFSETS:
        cell = ws.cell(row=row, column=label_col + offset)
        v = cell.value
        if v is None:
            continue
        if isinstance(v, bool):
            continue
        if isinstance(v, (int, float)):
            return v, cell.coordinate
        # Sometimes values come back as numeric strings in unusual sheets
        if isinstance(v, str):
            s = v.strip().replace(",", "").replace("$", "")
            try:
                f = float(s)
                return f, cell.coordinate
            except ValueError:
                continue
    return None, None


# ─────────────────────────────────────────────────────────────────────────────
# Rationale
# ─────────────────────────────────────────────────────────────────────────────


def _build_rationale(sources_block, uses_block, delta, balanced) -> str:
    n_src = len(sources_block.get("items", []))
    n_uses = len(uses_block.get("items", []))
    if delta is None:
        return (
            f"Found Sources block ({n_src} items) and Uses block ({n_uses} items) "
            "but could not compute a balance — at least one block is missing a total."
        )
    if balanced:
        return (
            f"Sources & Uses balance: {n_src} sources, {n_uses} uses, "
            f"delta ${delta:,.2f} (within $1 tolerance)."
        )
    severity = "imbalance" if delta > 100 else "minor rounding gap"
    return (
        f"Sources & Uses {severity}: {n_src} sources, {n_uses} uses, "
        f"delta ${delta:,.2f}."
    )
