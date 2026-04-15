"""
Phase 0 — Three-tier classifier.

Decides how much audit machinery to run on an uploaded workbook.

Tiers:
  FLAT             - single tabular block; rent roll, parcel list, comp grid.
                     No formulas worth auditing. Extraction layer handles it.
  ASSUMPTION_HEAVY - multiple sheets of labeled inputs, but no interdependent
                     model (no waterfall, no debt schedule, no CF build).
                     Audit = structural + formula integrity + assumption pull.
  FULL_MODEL       - waterfall / loan / cash flow sheets present with
                     interdependent formulas. Run all 7 phases.

Sheet-name prefix convention (Gregg): strip leading `i.` `m.` `r.` before
classification (i=inputs, m=model, r=reports).
"""

import re
from dataclasses import dataclass, asdict
from typing import Dict, List

from openpyxl.workbook.workbook import Workbook


class Tier:
    FLAT = "flat"
    ASSUMPTION_HEAVY = "assumption_heavy"
    FULL_MODEL = "full_model"


PREFIX_RE = re.compile(r"^([imr])\.\s*", re.IGNORECASE)

# Tokens that, when present in a sheet name, strongly signal full-model
FULL_MODEL_TOKENS = {
    "waterfall", "wfall", "promote", "distribution",
    "loan", "debt", "financing",
    "cashflow", "cash flow", "cf", "proforma", "pro forma",
    "irr", "returns", "operating",
}

ASSUMPTION_TOKENS = {
    "assumption", "assumptions", "inputs", "input",
    "rent roll", "rentroll", "rr",
    "absorption", "schedule",
    "sources", "uses", "s&u", "sources & uses",
}


@dataclass
class ClassificationResult:
    tier: str
    sheet_count: int
    formula_count: int
    full_model_hits: List[str]
    assumption_hits: List[str]
    rationale: str

    def to_dict(self) -> Dict:
        return asdict(self)


def _strip_prefix(name: str) -> str:
    return PREFIX_RE.sub("", name or "").strip().lower()


def _count_formulas(formulas_wb: Workbook, cap: int = 500) -> int:
    """
    Count formula cells, short-circuiting once we pass `cap`.

    Flat files have ~0; assumption-heavy ~50-500; full-model often 1000+.
    We only need a bucket, not the exact count.
    """
    count = 0
    for sheet in formulas_wb.worksheets:
        for row in sheet.iter_rows(values_only=False):
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val.startswith("="):
                    count += 1
                    if count >= cap:
                        return count
    return count


def _scan_sheet_tokens(values_wb: Workbook):
    full_hits: List[str] = []
    assumption_hits: List[str] = []
    for name in values_wb.sheetnames:
        clean = _strip_prefix(name)
        for tok in FULL_MODEL_TOKENS:
            if tok in clean:
                full_hits.append(name)
                break
        else:
            for tok in ASSUMPTION_TOKENS:
                if tok in clean:
                    assumption_hits.append(name)
                    break
    return full_hits, assumption_hits


def classify(values_wb: Workbook, formulas_wb: Workbook) -> Dict:
    """
    Returns a ClassificationResult as a dict.

    Rules:
      1. Any full-model token in any sheet name -> FULL_MODEL.
      2. Else if multiple sheets OR formula_count > 0 OR assumption tokens present
         -> ASSUMPTION_HEAVY.
      3. Else -> FLAT.

    Callers use the result to decide which downstream phases to run.
    """
    sheet_count = len(values_wb.sheetnames)
    full_hits, assumption_hits = _scan_sheet_tokens(values_wb)
    formula_count = _count_formulas(formulas_wb)

    if full_hits:
        tier = Tier.FULL_MODEL
        rationale = f"Full-model tokens detected in sheets: {full_hits}"
    elif sheet_count > 1 or formula_count > 0 or assumption_hits:
        tier = Tier.ASSUMPTION_HEAVY
        bits = []
        if sheet_count > 1:
            bits.append(f"{sheet_count} sheets")
        if formula_count > 0:
            bits.append(f"{formula_count}+ formulas" if formula_count >= 500 else f"{formula_count} formulas")
        if assumption_hits:
            bits.append(f"assumption sheets: {assumption_hits}")
        rationale = "Assumption-heavy: " + "; ".join(bits)
    else:
        tier = Tier.FLAT
        rationale = "Single sheet, no formulas, no structural tokens -> flat data"

    return ClassificationResult(
        tier=tier,
        sheet_count=sheet_count,
        formula_count=formula_count,
        full_model_hits=full_hits,
        assumption_hits=assumption_hits,
        rationale=rationale,
    ).to_dict()
