"""
Workbook loader for excel_audit.

Loads a workbook twice:
  - values_wb  (data_only=True)  -> computed values as last saved by Excel
  - formulas_wb (data_only=False) -> raw formula strings

Both are required: values for Python replication comparisons, formulas for
integrity checks (range consistency, broken refs, circular deps).

File resolution:
  core_doc.storage_uri -> django default_storage -> NamedTemporaryFile
"""

import logging
import os
import shutil
import tempfile
from typing import Tuple, Optional

import requests
from django.core.files.storage import default_storage
from django.db import connection
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

logger = logging.getLogger(__name__)

EXCEL_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
    "application/vnd.ms-excel.sheet.macroEnabled.12",                      # .xlsm
    "application/vnd.ms-excel",                                            # .xls (legacy)
}


class UnsupportedFileError(Exception):
    pass


def _fetch_doc_row(doc_id: int) -> Optional[Tuple[str, str, str]]:
    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT storage_uri, mime_type, doc_name
            FROM landscape.core_doc
            WHERE doc_id = %s
            """,
            [doc_id],
        )
        return cursor.fetchone()


def load_workbook_from_doc(doc_id: int) -> Tuple[Workbook, Workbook, str]:
    """
    Materialize a core_doc row to local disk and load twice.

    Returns:
        (values_wb, formulas_wb, tmp_path)

    Caller is responsible for cleaning up tmp_path via `os.unlink`.

    Raises:
        UnsupportedFileError if doc missing or non-Excel mime type.
    """
    row = _fetch_doc_row(doc_id)
    if not row:
        raise UnsupportedFileError(f"core_doc {doc_id} not found")

    storage_uri, mime_type, doc_name = row

    if mime_type not in EXCEL_MIME_TYPES:
        raise UnsupportedFileError(
            f"core_doc {doc_id} mime_type={mime_type!r} not supported by audit"
        )

    if not storage_uri:
        raise UnsupportedFileError(f"core_doc {doc_id} has no storage_uri")

    is_xlsm = "macroEnabled" in (mime_type or "")
    suffix = ".xlsm" if is_xlsm else ".xlsx"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        if storage_uri.startswith(("http://", "https://")):
            # UploadThing / any HTTP(S) URL — fetch directly
            resp = requests.get(storage_uri, stream=True, timeout=30)
            resp.raise_for_status()
            shutil.copyfileobj(resp.raw, tmp)
        else:
            # S3 key or local path — route through Django storage backend
            with default_storage.open(storage_uri, "rb") as src:
                tmp.write(src.read())
        tmp_path = tmp.name

    try:
        values_wb = load_workbook(
            tmp_path, data_only=True, read_only=False, keep_vba=is_xlsm
        )
        formulas_wb = load_workbook(
            tmp_path, data_only=False, read_only=False, keep_vba=is_xlsm
        )
    except Exception:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise

    logger.info(
        "excel_audit.load_workbook_from_doc doc_id=%s sheets=%d path=%s",
        doc_id,
        len(values_wb.sheetnames),
        tmp_path,
    )
    return values_wb, formulas_wb, tmp_path


def cleanup(tmp_path: str) -> None:
    """Best-effort temp file removal."""
    if not tmp_path:
        return
    try:
        os.unlink(tmp_path)
    except OSError:
        pass
