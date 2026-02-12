"""
Column Discovery Service for Rent Roll Extraction

Analyzes uploaded Excel/CSV files and proposes column mappings.
Does NOT extract data - just analyzes structure for user confirmation.

Also checks existing data in the project to inform import strategy,
analyzes Plan field derivation from Bed/Bath, and offers dynamic
column creation for valuable unmapped data.
"""

import re
import logging
import tempfile
from typing import List, Dict, Any, Optional, Tuple, Set
from dataclasses import dataclass, asdict
from enum import Enum
from decimal import Decimal
from datetime import datetime

import requests
from django.db import models as db_models

logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import csv
    HAS_CSV = True
except ImportError:
    HAS_CSV = False


class MappingConfidence(Enum):
    """Confidence level for column mapping."""
    HIGH = "high"      # 90%+ certain (exact header match)
    MEDIUM = "medium"  # 60-89% (fuzzy match or pattern)
    LOW = "low"        # <60% (guess based on data)
    NONE = "none"      # No match found


class MappingAction(Enum):
    """Suggested action for column mapping."""
    AUTO = "auto"           # High confidence, will map automatically
    SUGGEST = "suggest"     # Medium confidence, user should confirm
    NEEDS_INPUT = "needs_input"  # Low/no confidence, user must decide
    SKIP = "skip"           # User chose to skip


@dataclass
class ColumnMapping:
    """Column mapping proposal."""
    source_column: str           # Original column name from file
    source_index: int            # Column position in source (0-based)
    sample_values: List[str]     # First 5 non-empty values
    proposed_target: Optional[str]  # Suggested Landscape field
    confidence: str              # MappingConfidence value
    action: str                  # MappingAction value
    data_type_hint: str          # text, number, currency, date, boolean, percent
    notes: Optional[str]         # AI explanation of mapping decision


@dataclass
class DiscoveryResult:
    """Result of column discovery."""
    file_name: str
    total_rows: int
    total_columns: int
    columns: List[ColumnMapping]
    parse_warnings: List[str]    # Any issues during parsing
    is_structured: bool          # True for Excel/CSV, False for PDF


# Standard Landscape rent roll fields with aliases
# Based on tbl_multifamily_unit schema
STANDARD_FIELDS = {
    # Unit identification
    "unit_number": {
        "aliases": ["unit", "unit #", "unit no", "unit number", "apt", "apt #", "apartment", "unit no."],
        "data_type": "text",
        "required": True,
    },
    "building_name": {
        "aliases": ["building", "bldg", "building name", "property", "address"],
        "data_type": "text",
        "required": False,
    },
    "unit_type": {
        "aliases": ["type", "unit type", "floor plan", "floorplan", "plan", "style"],
        "data_type": "text",
        "required": False,
    },

    # Physical attributes
    "bedrooms": {
        "aliases": ["beds", "bed", "br", "bedrooms", "bd", "bedroom", "# beds"],
        "data_type": "number",
        "required": False,
    },
    "bathrooms": {
        "aliases": ["baths", "bath", "ba", "bathrooms", "bathroom", "# baths"],
        "data_type": "number",
        "required": False,
    },
    "square_feet": {
        "aliases": ["sqft", "sf", "sq ft", "square feet", "size", "area", "sq. ft.", "square footage"],
        "data_type": "number",
        "required": False,
    },

    # Occupancy
    "occupancy_status": {
        "aliases": ["status", "occupancy", "occ status", "lease status", "occ", "occupied"],
        "data_type": "text",
        "required": False,
    },

    # Tenant
    "tenant_name": {
        "aliases": ["tenant", "tenant name", "resident", "resident name", "lessee", "occupant", "name"],
        "data_type": "text",
        "required": False,
    },

    # Lease dates
    "lease_start": {
        "aliases": ["lease start", "lease from", "lease begin", "start date", "move in", "move-in", "movein", "commencement", "lease start date"],
        "data_type": "date",
        "required": False,
    },
    "lease_end": {
        "aliases": ["lease end", "lease to", "lease expiration", "expiration", "expiry", "end date", "lease end date", "expires", "lease exp"],
        "data_type": "date",
        "required": False,
    },
    "move_in_date": {
        "aliases": ["move in date", "move-in date", "movein date", "original move in"],
        "data_type": "date",
        "required": False,
    },

    # Financial
    "current_rent": {
        "aliases": ["rent", "contract rent", "actual rent", "current rent", "monthly rent", "base rent", "gross rent"],
        "data_type": "currency",
        "required": False,
    },
    "market_rent": {
        "aliases": ["market", "market rent", "asking rent", "asking", "mkt rent", "pro forma rent"],
        "data_type": "currency",
        "required": False,
    },

    # Renovation
    "renovation_status": {
        "aliases": ["reno status", "renovation", "renovated", "upgraded", "classic", "premium"],
        "data_type": "text",
        "required": False,
    },
    "renovation_date": {
        "aliases": ["reno date", "renovation date", "upgrade date", "renovated date"],
        "data_type": "date",
        "required": False,
    },
    "renovation_cost": {
        "aliases": ["reno cost", "renovation cost", "upgrade cost"],
        "data_type": "currency",
        "required": False,
    },
}


# Patterns that suggest specific data types
TYPE_PATTERNS = {
    "currency": [r"^\$", r"rent", r"income", r"fee", r"charge", r"amount", r"balance", r"delinquent", r"cost", r"price"],
    "date": [r"date", r"from", r"to", r"start", r"end", r"expir", r"move.?in", r"move.?out", r"effective"],
    "boolean": [r"section\s*8", r"voucher", r"subsidized", r"y/n", r"yes/no", r"is_", r"has_"],
    "number": [r"sqft", r"sf", r"beds", r"baths", r"bd", r"ba", r"count", r"#", r"number", r"qty"],
    "percent": [r"percent", r"rate", r"%", r"occupancy.*rate"],
}


def discover_columns(
    storage_uri: str,
    mime_type: Optional[str] = None,
    file_name: Optional[str] = None,
    project_id: Optional[int] = None
) -> DiscoveryResult:
    """
    Analyze an uploaded rent roll file and discover all columns.
    Returns mapping proposals for each column.

    Args:
        storage_uri: URL to the uploaded file
        mime_type: MIME type (optional, will infer from URL)
        file_name: Original filename (optional)
        project_id: Project ID for context

    Returns:
        DiscoveryResult with all columns and proposed mappings
    """
    warnings: List[str] = []

    # Infer mime type from URL if not provided
    if not mime_type:
        mime_type = _infer_mime_type(storage_uri, file_name)

    # Check if file type is supported for column discovery
    is_structured = _is_structured_file(mime_type)
    if not is_structured:
        return DiscoveryResult(
            file_name=file_name or "unknown",
            total_rows=0,
            total_columns=0,
            columns=[],
            parse_warnings=["PDF detected—Landscaper will analyze this document and propose extracted data for your review."],
            is_structured=False,
        )

    # Download file to temp location
    try:
        response = requests.get(storage_uri, timeout=60)
        response.raise_for_status()
    except Exception as e:
        return DiscoveryResult(
            file_name=file_name or "unknown",
            total_rows=0,
            total_columns=0,
            columns=[],
            parse_warnings=[f"Failed to download file: {str(e)}"],
            is_structured=False,
        )

    # Save to temp file
    suffix = _get_extension(mime_type)
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(response.content)
        tmp_path = tmp.name

    try:
        # Parse based on file type
        if mime_type in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'application/vnd.ms-excel']:
            headers, data_rows, parse_warnings = _parse_xlsx(tmp_path)
        elif mime_type in ['text/csv', 'application/csv']:
            headers, data_rows, parse_warnings = _parse_csv(tmp_path)
        else:
            return DiscoveryResult(
                file_name=file_name or "unknown",
                total_rows=0,
                total_columns=0,
                columns=[],
                parse_warnings=[f"Unsupported file type: {mime_type}"],
                is_structured=False,
            )

        warnings.extend(parse_warnings)

        # Build column mappings
        columns = []
        for idx, header in enumerate(headers):
            sample_values = _get_sample_values(data_rows, idx, max_samples=5)
            proposed_target, confidence = _match_column_to_field(header, sample_values)
            data_type = _infer_data_type(header, sample_values)

            # Determine action based on confidence
            if confidence == MappingConfidence.HIGH:
                action = MappingAction.AUTO
            elif confidence == MappingConfidence.MEDIUM:
                action = MappingAction.SUGGEST
            else:
                action = MappingAction.NEEDS_INPUT

            # Generate notes
            notes = _generate_mapping_notes(header, proposed_target, confidence, data_type)

            columns.append(ColumnMapping(
                source_column=header,
                source_index=idx,
                sample_values=sample_values,
                proposed_target=proposed_target,
                confidence=confidence.value,
                action=action.value,
                data_type_hint=data_type,
                notes=notes,
            ))

        return DiscoveryResult(
            file_name=file_name or "unknown",
            total_rows=len(data_rows),
            total_columns=len(headers),
            columns=columns,
            parse_warnings=warnings,
            is_structured=True,
        )

    finally:
        # Clean up temp file
        import os
        try:
            os.unlink(tmp_path)
        except Exception:
            pass


def _parse_xlsx(file_path: str) -> Tuple[List[str], List[List[Any]], List[str]]:
    """
    Parse Excel file and extract headers and data rows.

    Returns:
        Tuple of (headers, data_rows, warnings)
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")

    warnings = []
    wb = load_workbook(file_path, read_only=True, data_only=True)

    # Use first sheet (most rent rolls are single-sheet)
    sheet_name = wb.sheetnames[0]
    if len(wb.sheetnames) > 1:
        warnings.append(f"Multiple sheets detected. Using first sheet: '{sheet_name}'")

    sheet = wb[sheet_name]

    # Read all rows
    all_rows = []
    for row in sheet.iter_rows(values_only=True):
        # Convert to list and handle None
        row_data = [str(cell) if cell is not None else "" for cell in row]
        all_rows.append(row_data)

    wb.close()

    if not all_rows:
        return [], [], ["File appears to be empty"]

    # Detect header row (first row with multiple non-empty cells)
    header_row_idx = _find_header_row(all_rows)
    if header_row_idx == -1:
        warnings.append("Could not detect header row. Using first row.")
        header_row_idx = 0

    headers = all_rows[header_row_idx]
    data_rows = all_rows[header_row_idx + 1:]

    # Clean headers
    headers = [h.strip() if h else f"Column_{i+1}" for i, h in enumerate(headers)]

    # Remove completely empty rows
    data_rows = [row for row in data_rows if any(cell.strip() for cell in row)]

    return headers, data_rows, warnings


def _parse_csv(file_path: str) -> Tuple[List[str], List[List[Any]], List[str]]:
    """
    Parse CSV file and extract headers and data rows.

    Returns:
        Tuple of (headers, data_rows, warnings)
    """
    warnings = []

    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
        # Detect dialect
        sample = f.read(8192)
        f.seek(0)

        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel

        reader = csv.reader(f, dialect)
        all_rows = [row for row in reader]

    if not all_rows:
        return [], [], ["File appears to be empty"]

    # Detect header row
    header_row_idx = _find_header_row(all_rows)
    if header_row_idx == -1:
        warnings.append("Could not detect header row. Using first row.")
        header_row_idx = 0

    headers = all_rows[header_row_idx]
    data_rows = all_rows[header_row_idx + 1:]

    # Clean headers
    headers = [h.strip() if h else f"Column_{i+1}" for i, h in enumerate(headers)]

    # Remove completely empty rows
    data_rows = [row for row in data_rows if any(cell.strip() for cell in row if cell)]

    return headers, data_rows, warnings


def _find_header_row(rows: List[List[str]], max_check: int = 10) -> int:
    """
    Find the header row in the data.
    Headers typically have multiple non-empty text cells without numbers.
    """
    for idx, row in enumerate(rows[:max_check]):
        non_empty = [cell for cell in row if cell and cell.strip()]
        if len(non_empty) >= 3:
            # Check if this looks like headers (mostly text, not numbers)
            text_count = sum(1 for cell in non_empty if not _is_numeric(cell))
            if text_count >= len(non_empty) * 0.7:
                return idx
    return 0


def _get_sample_values(data_rows: List[List[Any]], col_idx: int, max_samples: int = 5) -> List[str]:
    """Get sample non-empty values from a column."""
    samples = []
    for row in data_rows:
        if col_idx < len(row):
            val = str(row[col_idx]).strip()
            if val and val not in samples:
                samples.append(val)
                if len(samples) >= max_samples:
                    break
    return samples


def _match_column_to_field(
    column_name: str,
    sample_values: List[str]
) -> Tuple[Optional[str], MappingConfidence]:
    """
    Attempt to match a source column to a standard Landscape field.
    Returns (field_name, confidence) or (None, NONE).
    """
    if not column_name:
        return None, MappingConfidence.NONE

    normalized = column_name.lower().strip()

    # Check exact matches first (HIGH confidence)
    for field, config in STANDARD_FIELDS.items():
        for alias in config["aliases"]:
            if normalized == alias:
                return field, MappingConfidence.HIGH

    # Check partial matches (MEDIUM confidence)
    for field, config in STANDARD_FIELDS.items():
        for alias in config["aliases"]:
            # Check if alias is contained in column name or vice versa
            if alias in normalized or normalized in alias:
                return field, MappingConfidence.MEDIUM

    # Check word boundaries (MEDIUM confidence)
    words = set(re.split(r'\W+', normalized))
    for field, config in STANDARD_FIELDS.items():
        for alias in config["aliases"]:
            alias_words = set(re.split(r'\W+', alias))
            if alias_words & words:  # Any overlap
                return field, MappingConfidence.MEDIUM

    # No match found
    return None, MappingConfidence.NONE


def _infer_data_type(column_name: str, sample_values: List[str]) -> str:
    """
    Infer the data type from column name and sample values.
    Returns: text, number, currency, date, boolean, percent
    """
    col_lower = column_name.lower()

    # Check column name patterns
    for dtype, patterns in TYPE_PATTERNS.items():
        for pattern in patterns:
            if re.search(pattern, col_lower, re.IGNORECASE):
                return dtype

    # Analyze sample values
    if sample_values:
        # Check for currency
        if any(v.startswith('$') or v.startswith('-$') or v.startswith('($') for v in sample_values if v):
            return "currency"

        # Check for dates
        date_patterns = [r'\d{1,2}/\d{1,2}/\d{2,4}', r'\d{4}-\d{2}-\d{2}', r'\d{1,2}-\d{1,2}-\d{2,4}']
        if any(re.match(p, v) for v in sample_values for p in date_patterns if v):
            return "date"

        # Check for percentages
        if any('%' in v for v in sample_values if v):
            return "percent"

        # Check for boolean-like
        bool_vals = {'yes', 'no', 'y', 'n', 'true', 'false', '1', '0', 'x', ''}
        cleaned_vals = [v.lower().strip() for v in sample_values if v]
        if cleaned_vals and all(v in bool_vals for v in cleaned_vals):
            return "boolean"

        # Check for numbers
        try:
            for v in sample_values:
                if v:
                    clean = v.replace(',', '').replace('$', '').replace('(', '-').replace(')', '').strip()
                    if clean and clean != '-':
                        float(clean)
            return "number"
        except ValueError:
            pass

    return "text"


def _generate_mapping_notes(
    header: str,
    proposed_target: Optional[str],
    confidence: MappingConfidence,
    data_type: str
) -> Optional[str]:
    """Generate human-readable notes about the mapping decision."""
    if confidence == MappingConfidence.HIGH:
        return f"Exact match to '{proposed_target}' field"
    elif confidence == MappingConfidence.MEDIUM:
        return f"Likely matches '{proposed_target}' - please confirm"
    elif proposed_target:
        return f"Possible match to '{proposed_target}' - review recommended"
    else:
        if data_type == "currency":
            return "Contains currency values - may be a financial field"
        elif data_type == "date":
            return "Contains date values"
        elif data_type == "boolean":
            return "Contains yes/no or flag values"
        return "No automatic mapping found - choose a field or create new column"


def _is_structured_file(mime_type: str) -> bool:
    """Check if file type supports column discovery."""
    structured_types = [
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'application/vnd.ms-excel',
        'text/csv',
        'application/csv',
    ]
    return mime_type in structured_types


def _infer_mime_type(url: str, file_name: Optional[str] = None) -> str:
    """Infer MIME type from URL or filename."""
    check_str = (file_name or url).lower()

    if check_str.endswith('.xlsx'):
        return 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    elif check_str.endswith('.xls'):
        return 'application/vnd.ms-excel'
    elif check_str.endswith('.csv'):
        return 'text/csv'
    elif check_str.endswith('.pdf'):
        return 'application/pdf'

    return 'application/octet-stream'


def _get_extension(mime_type: str) -> str:
    """Get file extension for MIME type."""
    extensions = {
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': '.xlsx',
        'application/vnd.ms-excel': '.xls',
        'text/csv': '.csv',
        'application/csv': '.csv',
    }
    return extensions.get(mime_type, '.tmp')


def _is_numeric(value: str) -> bool:
    """Check if a string value is numeric."""
    if not value:
        return False
    try:
        clean = value.replace(',', '').replace('$', '').replace('(', '').replace(')', '').replace('%', '').strip()
        if clean:
            float(clean)
            return True
    except (ValueError, TypeError):
        pass
    return False


def analyze_complex_column(column_name: str, sample_values: List[str]) -> Dict[str, Any]:
    """
    Analyze columns that might contain multiple data types or need splitting.
    Example: "Tags" column with "Residential Unit, Sec. 8, UD Mirage"

    Returns suggestions for splitting or parsing.
    """
    suggestions = []

    # Check for delimiter-separated values
    delimiters = [',', ';', '|', '/']
    for delim in delimiters:
        if any(delim in v for v in sample_values if v):
            # This column might need splitting
            all_parts = set()
            for v in sample_values:
                if v:
                    parts = [p.strip() for p in v.split(delim)]
                    all_parts.update(p for p in parts if p)

            if len(all_parts) > 1:
                suggestions.append({
                    "type": "split",
                    "delimiter": delim,
                    "unique_values": sorted(list(all_parts))[:20],  # Cap at 20
                    "suggestion": f"Contains {len(all_parts)} unique values separated by '{delim}'. Consider splitting into separate columns."
                })

    # Check for Section 8 indicators
    sec8_indicators = ['sec 8', 'sec. 8', 'section 8', 'section8', 'hcv', 'voucher']
    if any(any(ind in v.lower() for ind in sec8_indicators) for v in sample_values if v):
        suggestions.append({
            "type": "extract_boolean",
            "field": "is_section_8",
            "suggestion": "Contains Section 8/voucher indicators. Extract as a Yes/No column?"
        })

    # Check for delinquency patterns
    delinquent_patterns = [r'delinquent', r'past.?due', r'balance', r'owing']
    col_lower = column_name.lower()
    if any(re.search(p, col_lower) for p in delinquent_patterns):
        suggestions.append({
            "type": "extract_currency",
            "field": "delinquent_balance",
            "suggestion": "Contains delinquency/balance data. Extract as currency column?"
        })

    return {
        "column": column_name,
        "suggestions": suggestions,
        "sample_values": sample_values[:10]
    }


def discovery_result_to_dict(result: DiscoveryResult) -> Dict[str, Any]:
    """Convert DiscoveryResult to JSON-serializable dict."""
    return {
        "file_name": result.file_name,
        "total_rows": result.total_rows,
        "total_columns": result.total_columns,
        "columns": [asdict(col) for col in result.columns],
        "parse_warnings": result.parse_warnings,
        "is_structured": result.is_structured,
    }


def apply_column_mapping(
    project_id: int,
    document_id: int,
    mappings: List[Dict[str, Any]],
) -> Dict[str, Any]:
    """
    Apply column mapping decisions and queue an extraction job.

    Shared logic used by both the HTTP endpoint (extraction_views.apply_rent_roll_mapping)
    and the Landscaper tool handler (confirm_column_mapping).

    Args:
        project_id: Project ID
        document_id: Document ID of the rent roll file
        mappings: List of mapping decision dicts, each with:
            - source_column: str (required)
            - target_field: str|null (standard field name)
            - create_dynamic: bool (optional)
            - dynamic_column_name: str (optional, if create_dynamic)
            - data_type: str (optional, if create_dynamic)

    Returns:
        Dict with: success, job_id, job_status, dynamic_columns_created,
        standard_mappings, skipped_columns, error (if any)
    """
    import logging
    import threading
    from django.utils import timezone

    logger = logging.getLogger(__name__)

    # Lazy imports to avoid circular dependencies
    from apps.dynamic.models import DynamicColumnDefinition
    from apps.knowledge.models import ExtractionJob

    # Separate standard field mappings from dynamic column creations
    standard_mappings = {}
    dynamic_columns_to_create = []
    skipped_columns = []

    for mapping in mappings:
        source_col = mapping.get('source_column')
        if not source_col:
            continue

        if mapping.get('create_dynamic'):
            dynamic_columns_to_create.append({
                'source_column': source_col,
                'name': mapping.get('dynamic_column_name', source_col),
                'data_type': mapping.get('data_type', 'text'),
            })
        elif mapping.get('target_field'):
            standard_mappings[source_col] = mapping['target_field']
        else:
            skipped_columns.append(source_col)

    # Create dynamic column definitions (immediately active, not proposed)
    created_columns = []
    for dc in dynamic_columns_to_create:
        col_key = dc['name'].lower().replace(' ', '_').replace('-', '_')

        col_def, created = DynamicColumnDefinition.objects.get_or_create(
            project_id=project_id,
            table_name='multifamily_unit',
            column_key=col_key,
            defaults={
                'display_label': dc['name'],
                'data_type': dc['data_type'],
                'source': 'user',
                'is_proposed': False,
                'proposed_from_document_id': document_id,
            }
        )
        created_columns.append({
            'id': col_def.id,
            'column_key': col_def.column_key,
            'display_label': col_def.display_label,
            'source_column': dc['source_column'],
            'created': created,
        })

    # Check for existing active job for this document
    existing_job = ExtractionJob.objects.filter(
        project_id=project_id,
        document_id=document_id,
        scope='rent_roll',
        status__in=['queued', 'processing']
    ).first()

    if existing_job:
        return {
            'success': False,
            'error': 'An extraction job is already in progress for this document',
            'existing_job_id': existing_job.job_id,
        }

    # Create extraction job with mapping metadata
    job = ExtractionJob.objects.create(
        project_id=project_id,
        document_id=document_id,
        scope='rent_roll',
        status='queued',
        result_summary={
            'standard_mappings': standard_mappings,
            'dynamic_columns': [c for c in created_columns],
            'skipped_columns': skipped_columns,
        }
    )

    def run_extraction_async(job_id: int, proj_id: int, doc_id: int):
        """Background thread to run extraction without blocking."""
        import django
        django.db.connections.close_all()

        try:
            from apps.knowledge.services.extraction_service import ChunkedRentRollExtractor
            from apps.knowledge.models import ExtractionJob as EJ

            job = EJ.objects.get(job_id=job_id)
            logger.info(f"[async_extraction] Starting extraction for job {job_id}, doc {doc_id}")

            job.status = 'processing'
            job.started_at = timezone.now()
            job.save()

            extractor = ChunkedRentRollExtractor(
                project_id=proj_id,
                property_type='multifamily'
            )

            extract_result = extractor.extract_rent_roll_chunked(
                doc_id=doc_id,
                user_id=None
            )

            job.status = 'completed'
            job.completed_at = timezone.now()
            job.result_summary.update({
                'units_extracted': extract_result.get('units_extracted', 0),
                'staged_count': extract_result.get('staged_count', 0),
            })
            job.save()
            logger.info(f"[async_extraction] Completed job {job_id}: {extract_result.get('units_extracted', 0)} units")

        except Exception as e:
            logger.exception(f"[async_extraction] Failed for job {job_id}: {e}")
            try:
                job = EJ.objects.get(job_id=job_id)
                job.status = 'failed'
                job.error_message = str(e)
                job.completed_at = timezone.now()
                job.save()
            except Exception as save_err:
                logger.exception(f"[async_extraction] Failed to save error state: {save_err}")

    # Start background thread
    thread = threading.Thread(
        target=run_extraction_async,
        args=(job.job_id, project_id, document_id),
        daemon=False,
        name=f"extraction-job-{job.job_id}"
    )
    thread.start()

    logger.info(f"[apply_column_mapping] Started async extraction job {job.job_id}")

    return {
        'success': True,
        'job_id': job.job_id,
        'job_status': 'queued',
        'dynamic_columns_created': len([c for c in created_columns if c.get('created')]),
        'standard_mappings': len(standard_mappings),
        'skipped_columns': len(skipped_columns),
        'message': 'Extraction started. Poll job status for progress.',
    }


# ─────────────────────────────────────────────────────────────────────────────
# Existing Data Analysis
# ─────────────────────────────────────────────────────────────────────────────

def analyze_existing_rent_roll(project_id: int) -> dict:
    """
    Analyze existing rent roll data for a project.
    Returns summary of what's populated vs empty.
    """
    from apps.multifamily.models import MultifamilyUnit, MultifamilyLease

    units = MultifamilyUnit.objects.filter(project_id=project_id)

    if not units.exists():
        return {
            'status': 'empty',
            'unit_count': 0,
            'lease_count': 0,
            'existing_unit_numbers': [],
            'field_stats': {},
            'message': 'No existing units. This will be a fresh import.',
        }

    unit_count = units.count()

    # Check field population on tbl_multifamily_unit
    field_stats = {}
    # (field_name, label, is_text_field)
    unit_fields_to_check = [
        ('unit_number', 'Unit Number', True),
        ('bedrooms', 'Bedrooms', False),
        ('bathrooms', 'Bathrooms', False),
        ('square_feet', 'Square Feet', False),
        ('current_rent', 'Current Rent', False),
        ('market_rent', 'Market Rent', False),
        ('occupancy_status', 'Occupancy Status', True),
        ('unit_type', 'Plan/Type', True),
    ]

    for field, label, is_text in unit_fields_to_check:
        qs = units.exclude(**{f'{field}__isnull': True})
        if is_text:
            qs = qs.exclude(**{field: ''})
        populated = qs.count()
        field_stats[field] = {
            'label': label,
            'populated': populated,
            'empty': unit_count - populated,
            'percent_populated': round(populated / unit_count * 100) if unit_count > 0 else 0,
        }

    # Check for leases
    leases = MultifamilyLease.objects.filter(unit__project_id=project_id)
    lease_count = leases.count()

    # Get unit numbers for comparison
    existing_unit_numbers = list(units.values_list('unit_number', flat=True))

    # Determine overall status
    mostly_populated = sum(1 for f in field_stats.values() if f['percent_populated'] > 80)
    if mostly_populated >= 5:
        status = 'fully_populated'
    elif mostly_populated >= 2:
        status = 'partially_populated'
    else:
        status = 'sparse'

    return {
        'status': status,
        'unit_count': unit_count,
        'lease_count': lease_count,
        'existing_unit_numbers': existing_unit_numbers,
        'field_stats': field_stats,
        'message': _generate_existing_data_message(status, unit_count, field_stats),
    }


def _generate_existing_data_message(status: str, unit_count: int, field_stats: dict) -> str:
    """Generate human-readable message about existing data."""
    if status == 'empty':
        return "No existing units found. This will be a fresh import."

    # Find fields with gaps
    fields_with_gaps = [
        f['label'] for f in field_stats.values()
        if f['empty'] > 0 and f['percent_populated'] < 100
    ]

    if status == 'fully_populated':
        return f"Found {unit_count} existing units with most fields populated."
    elif status == 'partially_populated':
        gaps = ', '.join(fields_with_gaps[:3])
        return f"Found {unit_count} existing units. Some fields have gaps: {gaps}."
    else:
        return f"Found {unit_count} existing units but many fields are empty."


def compare_file_to_existing(file_unit_numbers: list, existing_unit_numbers: list) -> dict:
    """Compare units in file vs existing in database."""
    file_set = set(str(u) for u in file_unit_numbers)
    existing_set = set(str(u) for u in existing_unit_numbers)

    return {
        'in_file_only': sorted(list(file_set - existing_set)),
        'in_db_only': sorted(list(existing_set - file_set)),
        'in_both': sorted(list(file_set & existing_set)),
        'file_count': len(file_set),
        'existing_count': len(existing_set),
        'new_units': len(file_set - existing_set),
        'matching_units': len(file_set & existing_set),
    }


def suggest_import_action(existing: dict, comparison: Optional[dict]) -> dict:
    """Suggest import action based on existing data and comparison."""
    if existing['status'] == 'empty':
        return {
            'action': 'create_all',
            'message': 'No existing data. All units from this file will be created.',
            'options': None,
        }

    if comparison is None:
        return {
            'action': 'unknown',
            'message': 'Could not compare file to existing data.',
            'options': ['overwrite_all', 'cancel'],
        }

    new_units = comparison['new_units']
    matching = comparison['matching_units']

    if new_units > 0 and matching > 0:
        return {
            'action': 'needs_decision',
            'message': f"Found {matching} existing units that match the file, plus {new_units} new units.",
            'options': [
                {'key': 'A', 'action': 'add_new_only', 'label': f'Add {new_units} new units only (keep existing unchanged)'},
                {'key': 'B', 'action': 'update_and_add', 'label': f'Update {matching} existing + add {new_units} new'},
                {'key': 'C', 'action': 'fill_blanks', 'label': 'Fill blank fields only (never overwrite existing values)'},
                {'key': 'D', 'action': 'overwrite_all', 'label': 'Replace all data from file'},
            ],
        }
    elif matching > 0 and new_units == 0:
        # Check for field gaps
        fields_with_gaps = [
            k for k, v in existing.get('field_stats', {}).items()
            if v['empty'] > 0
        ]

        if fields_with_gaps:
            return {
                'action': 'needs_decision',
                'message': f"All {matching} units exist. Some fields have gaps that could be filled.",
                'options': [
                    {'key': 'A', 'action': 'fill_blanks', 'label': 'Fill blank fields only'},
                    {'key': 'B', 'action': 'update_changed', 'label': 'Update fields where file value differs'},
                    {'key': 'C', 'action': 'overwrite_all', 'label': 'Overwrite all fields from file'},
                    {'key': 'D', 'action': 'skip', 'label': 'Cancel - keep existing data'},
                ],
            }
        else:
            return {
                'action': 'needs_decision',
                'message': f"All {matching} units exist and are fully populated.",
                'options': [
                    {'key': 'A', 'action': 'update_changed', 'label': 'Update only where values differ'},
                    {'key': 'B', 'action': 'overwrite_all', 'label': 'Overwrite all from file'},
                    {'key': 'C', 'action': 'skip', 'label': 'Cancel - keep existing data'},
                ],
            }
    else:
        return {
            'action': 'create_all',
            'message': f'File contains {new_units} units not in database.',
            'options': None,
        }


def _extract_unit_numbers_from_data(data_rows: List[List[Any]], unit_col_idx: Optional[int]) -> list:
    """Extract unit numbers from parsed file data."""
    if unit_col_idx is None:
        return []

    unit_numbers = []
    for row in data_rows:
        if unit_col_idx < len(row):
            val = str(row[unit_col_idx]).strip()
            if val and val.lower() not in ('', 'none', 'nan'):
                unit_numbers.append(val)
    return unit_numbers


# ─────────────────────────────────────────────────────────────────────────────
# Plan Field Analysis
# ─────────────────────────────────────────────────────────────────────────────

def analyze_plan_field_status(project_id: int) -> dict:
    """
    Analyze Plan/Type field population and whether it can be derived from Bed/Bath.
    """
    from apps.multifamily.models import MultifamilyUnit

    units = MultifamilyUnit.objects.filter(project_id=project_id)

    if not units.exists():
        return {
            'populated': 0,
            'empty': 0,
            'derivable_count': 0,
            'derivable_sample': [],
            'message': 'No existing units to check Plan field.',
        }

    # Count units with Plan populated
    with_plan = units.exclude(unit_type__isnull=True).exclude(unit_type='').exclude(unit_type='Unknown').count()
    without_plan_qs = units.filter(
        db_models.Q(unit_type__isnull=True) | db_models.Q(unit_type='') | db_models.Q(unit_type='Unknown')
    )
    without_plan_count = without_plan_qs.count()

    # Check if we can derive Plan from Bed/Bath
    derivable = []
    for unit in without_plan_qs.exclude(
        bedrooms__isnull=True
    ).exclude(
        bathrooms__isnull=True
    )[:10]:  # Limit query
        br = int(unit.bedrooms)
        ba = int(unit.bathrooms)
        if br == 0 and ba == 0:
            continue
        suggested = f"Studio/{ba}BA" if br == 0 else f"{br}BR/{ba}BA"
        derivable.append({
            'unit_number': unit.unit_number,
            'bedrooms': float(unit.bedrooms),
            'bathrooms': float(unit.bathrooms),
            'suggested_plan': suggested,
        })

    return {
        'populated': with_plan,
        'empty': without_plan_count,
        'derivable_count': len(derivable),
        'derivable_sample': derivable[:5],
        'message': _generate_plan_message(with_plan, without_plan_count, len(derivable)),
    }


def _generate_plan_message(populated: int, empty: int, derivable: int) -> str:
    """Generate message about Plan field status."""
    if empty == 0 and populated > 0:
        return f"All {populated} units have Plan populated."
    elif empty == 0 and populated == 0:
        return "No units to check."
    elif derivable > 0:
        return (
            f"{populated} units have Plan populated, {empty} are blank. "
            f"I can derive {derivable} Plan values from Bed/Bath data "
            f"(e.g., '3BR/2BA')."
        )
    else:
        return (
            f"{populated} units have Plan populated, {empty} are blank "
            f"but cannot be derived (missing Bed/Bath data)."
        )


# ─────────────────────────────────────────────────────────────────────────────
# Dynamic Column Offers for Unmapped Columns
# ─────────────────────────────────────────────────────────────────────────────

def analyze_unmapped_columns_for_dynamic(
    columns: List[ColumnMapping],
    data_rows: List[List[Any]],
    project_id: int,
) -> list:
    """
    Analyze unmapped columns and determine if they should be offered as dynamic columns.

    Returns a list of proposed dynamic column offers for the user to accept/reject.
    """
    from apps.dynamic.models import DynamicColumnDefinition

    # Get existing dynamic columns
    existing_dynamic = set(
        DynamicColumnDefinition.objects.filter(
            project_id=project_id,
            table_name='multifamily_unit'
        ).values_list('column_key', flat=True)
    )

    valuable_unmapped = []

    for col in columns:
        # Only look at unmapped columns
        if col.proposed_target is not None:
            continue

        source_header = col.source_column
        source_lower = source_header.lower()
        col_idx = col.source_index

        # Get full column data for analysis
        col_values = []
        for row in data_rows:
            if col_idx < len(row):
                val = str(row[col_idx]).strip()
                if val and val.lower() not in ('', 'none', 'nan'):
                    col_values.append(val)

        # Skip columns with very few values
        if len(col_values) < 5:
            continue

        column_offer = {
            'source_header': source_header,
            'samples': col_values[:5],
            'non_null_count': len(col_values),
        }

        # --- Tags/Labels columns ---
        if 'tag' in source_lower or 'label' in source_lower or 'type' in source_lower:
            detected_flags = _detect_tag_flags(col_values)
            if detected_flags:
                column_offer['type'] = 'tags'
                column_offer['suggestion'] = 'split'
                column_offer['proposed_columns'] = []

                if detected_flags.get('section_8', 0) > 0:
                    column_offer['proposed_columns'].append({
                        'key': 'is_section_8',
                        'label': 'Section 8',
                        'data_type': 'boolean',
                        'description': f"Yes/No flag ({detected_flags['section_8']} units have Section 8)",
                    })
                if detected_flags.get('payment_plan', 0) > 0:
                    column_offer['proposed_columns'].append({
                        'key': 'has_payment_plan',
                        'label': 'Payment Plan',
                        'data_type': 'boolean',
                        'description': f"Yes/No flag ({detected_flags['payment_plan']} units on payment plans)",
                    })
                if detected_flags.get('unlawful_detainer', 0) > 0:
                    column_offer['proposed_columns'].append({
                        'key': 'has_unlawful_detainer',
                        'label': 'Unlawful Detainer',
                        'data_type': 'boolean',
                        'description': f"Yes/No flag ({detected_flags['unlawful_detainer']} units with UD status)",
                    })

                if column_offer['proposed_columns']:
                    valuable_unmapped.append(column_offer)
                    continue

        # --- Delinquency columns ---
        if 'delinqu' in source_lower or 'past due' in source_lower or 'owed' in source_lower or 'balance' in source_lower:
            currency_stats = _analyze_currency_column(col_values)
            if currency_stats['non_zero_count'] > 0:
                column_offer['type'] = 'currency'
                column_offer['suggestion'] = 'create'
                column_offer['proposed_columns'] = [{
                    'key': 'delinquent_amount',
                    'label': 'Delinquent Rent',
                    'data_type': 'currency',
                    'description': f"${currency_stats['total']:,.0f} total across {currency_stats['non_zero_count']} units",
                }]
                valuable_unmapped.append(column_offer)
                continue

        # --- Payment/Receipt columns ---
        if 'received' in source_lower or 'paid' in source_lower or 'payment' in source_lower:
            currency_stats = _analyze_currency_column(col_values)
            if currency_stats['non_zero_count'] > 0:
                col_key = re.sub(r'[^a-z0-9_]', '_', source_lower).strip('_')
                column_offer['type'] = 'currency'
                column_offer['suggestion'] = 'create'
                column_offer['proposed_columns'] = [{
                    'key': col_key,
                    'label': source_header,
                    'data_type': 'currency',
                    'description': f"${currency_stats['total']:,.0f} total across {currency_stats['non_zero_count']} units",
                }]
                valuable_unmapped.append(column_offer)
                continue

    return valuable_unmapped


def _detect_tag_flags(values: list) -> dict:
    """Detect common flags in a Tags/Labels column."""
    flags = {
        'section_8': 0,
        'payment_plan': 0,
        'unlawful_detainer': 0,
    }

    sec8_patterns = ['sec 8', 'sec. 8', 'section 8', 'section8', 'hcv', 'voucher']
    plan_patterns = ['payment plan', 'pay plan', 'pmt plan']
    ud_patterns = ['unlawful detainer', 'ud ', 'u.d.', 'eviction', 'ud,']

    for val in values:
        lower = val.lower()
        if any(p in lower for p in sec8_patterns):
            flags['section_8'] += 1
        if any(p in lower for p in plan_patterns):
            flags['payment_plan'] += 1
        if any(p in lower for p in ud_patterns):
            flags['unlawful_detainer'] += 1

    # Only return flags that were actually found
    return {k: v for k, v in flags.items() if v > 0}


def _analyze_currency_column(values: list) -> dict:
    """Analyze a currency column for totals and non-zero counts."""
    total = 0.0
    non_zero_count = 0

    for val in values:
        try:
            clean = val.replace('$', '').replace(',', '').replace('(', '-').replace(')', '').strip()
            if clean and clean != '-':
                amount = float(clean)
                if amount != 0:
                    total += abs(amount)
                    non_zero_count += 1
        except (ValueError, TypeError):
            continue

    return {
        'total': total,
        'non_zero_count': non_zero_count,
        'count': len(values),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Enhanced Discovery (wraps discover_columns with existing data + plan + dynamic)
# ─────────────────────────────────────────────────────────────────────────────

def discover_columns_enhanced(
    storage_uri: str,
    project_id: int,
    mime_type: Optional[str] = None,
    file_name: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Full enhanced column discovery: file analysis + existing data check +
    Plan field status + dynamic column offers.

    Returns a comprehensive dict for the Landscaper to present to the user.
    """
    # 1. Analyze existing data in the project
    existing_analysis = analyze_existing_rent_roll(project_id)

    # 2. Run standard column discovery
    result = discover_columns(
        storage_uri=storage_uri,
        mime_type=mime_type,
        file_name=file_name,
        project_id=project_id,
    )

    if not result.is_structured:
        return {
            **discovery_result_to_dict(result),
            'existing_data': existing_analysis,
            'comparison': None,
            'suggested_action': None,
            'plan_analysis': None,
            'dynamic_column_offers': [],
        }

    # 3. Re-parse file to get raw data rows for unit number extraction and dynamic analysis
    data_rows = []
    unit_col_idx = None

    # Find unit_number column index from discovery results
    for col in result.columns:
        if col.proposed_target == 'unit_number':
            unit_col_idx = col.source_index
            break

    # Re-download and parse for raw data (needed for unit numbers + dynamic analysis)
    if not mime_type:
        mime_type = _infer_mime_type(storage_uri, file_name)

    try:
        response = requests.get(storage_uri, timeout=60)
        response.raise_for_status()
        suffix = _get_extension(mime_type)
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(response.content)
            tmp_path = tmp.name

        try:
            if mime_type in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'application/vnd.ms-excel']:
                _, data_rows, _ = _parse_xlsx(tmp_path)
            elif mime_type in ['text/csv', 'application/csv']:
                _, data_rows, _ = _parse_csv(tmp_path)
        finally:
            import os
            try:
                os.unlink(tmp_path)
            except Exception:
                pass
    except Exception as e:
        logger.warning(f"Could not re-parse file for enhanced analysis: {e}")

    # 4. Extract unit numbers and compare to existing
    file_unit_numbers = _extract_unit_numbers_from_data(data_rows, unit_col_idx)

    comparison = None
    if existing_analysis['status'] != 'empty' and file_unit_numbers:
        comparison = compare_file_to_existing(
            file_unit_numbers,
            existing_analysis['existing_unit_numbers'],
        )

    # 5. Suggest import action
    suggested_action = suggest_import_action(existing_analysis, comparison)

    # 6. Analyze Plan field status
    plan_analysis = analyze_plan_field_status(project_id)

    # 7. Analyze unmapped columns for dynamic column offers
    dynamic_offers = analyze_unmapped_columns_for_dynamic(
        result.columns,
        data_rows,
        project_id,
    )

    # Build response
    base_result = discovery_result_to_dict(result)

    return {
        **base_result,
        'existing_data': existing_analysis,
        'comparison': comparison,
        'suggested_action': suggested_action,
        'plan_analysis': plan_analysis,
        'dynamic_column_offers': dynamic_offers,
        'summary': {
            'total_columns': len(result.columns),
            'auto_mapped_count': len([c for c in result.columns if c.action == MappingAction.AUTO.value]),
            'unmapped_count': len([c for c in result.columns if c.proposed_target is None]),
            'existing_unit_count': existing_analysis.get('unit_count', 0),
            'file_unit_count': len(file_unit_numbers),
            'new_columns_available': len(dynamic_offers),
        },
    }


def format_discovery_for_chat(discovery_result: dict) -> str:
    """
    Format discovery results as a compact string for Claude's tool response.
    Keeps total output under ~500 tokens to preserve attention on behavioral rules.

    Uses actual key names from discover_columns_enhanced() return structure.
    """
    parts = []

    # 1. Existing data — counts only, no unit lists
    existing = discovery_result.get('existing_data') or {}
    comparison = discovery_result.get('comparison') or {}
    suggested = discovery_result.get('suggested_action') or {}

    db_count = existing.get('unit_count', 0)
    file_count = comparison.get('file_count', 0)
    matching = comparison.get('matching_units', 0)
    new_units = comparison.get('new_units', 0)
    db_only_list = comparison.get('in_db_only', [])
    db_only_count = len(db_only_list)

    if db_count > 0 or file_count > 0:
        parts.append(
            f"EXISTING DATA: {db_count} units in DB, {file_count} in file. "
            f"{matching} matching, {new_units} new, {db_only_count} DB-only."
        )
        if db_only_count > 0 and db_only_list:
            unit_nums = db_only_list[:20]
            parts.append(f"DB-only units: {', '.join(str(u) for u in unit_nums)}")
    else:
        parts.append(f"EXISTING DATA: {existing.get('message', 'No existing units.')}")

    # Suggested action
    if suggested.get('message'):
        parts.append(f"Suggested: {suggested['message']}")
    if suggested.get('options'):
        for opt in suggested['options']:
            parts.append(f"  {opt['key']}) {opt['label']}")

    # 2. Column mapping — compact, no sample data
    columns = discovery_result.get('columns', [])
    auto_mapped = []
    needs_confirm = []
    unmapped = []
    for col in columns:
        name = col.get('source_column', '')
        target = col.get('proposed_target', '')
        confidence = col.get('confidence', '')
        if confidence == 'high' and target:
            auto_mapped.append(f"{name} -> {target}")
        elif confidence == 'medium' and target:
            needs_confirm.append(f"{name} -> {target}")
        elif not target:
            unmapped.append(name)

    if auto_mapped:
        parts.append(f"AUTO-MAPPED: {', '.join(auto_mapped)}")
    if needs_confirm:
        parts.append(f"NEEDS CONFIRM: {', '.join(needs_confirm)}")
    if unmapped:
        parts.append(f"UNMAPPED: {', '.join(unmapped)}")

    # 3. Plan field — counts only
    plan = discovery_result.get('plan_analysis') or {}
    populated = plan.get('populated', 0)
    blank = plan.get('empty', 0)
    derivable = plan.get('derivable_count', 0)
    if blank > 0:
        parts.append(f"PLAN: {populated} populated, {blank} blank. {derivable} can auto-derive from Bed/Bath.")
    elif populated > 0:
        parts.append(f"PLAN: All {populated} units have Plan populated.")

    # 4. Dynamic column offers — names and counts only
    dynamic = discovery_result.get('dynamic_column_offers', [])
    if dynamic:
        offers = []
        for d in dynamic:
            for proposed in d.get('proposed_columns', []):
                label = proposed.get('label', '')
                dtype = proposed.get('data_type', '')
                desc = proposed.get('description', '')
                offers.append(f"{label} ({desc}, {dtype})")
        if offers:
            parts.append(f"DYNAMIC COLUMN CANDIDATES: {', '.join(offers)}")

    return '\n'.join(parts)
