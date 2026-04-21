"""
CoStar Report Extractor

Handles deterministic extraction from CoStar Excel exports and
enhanced LLM extraction from CoStar PDF comp reports.

Two paths:
  Excel — Direct column-header mapping to tbl_sales_comparables.
          Confidence 0.95 (deterministic). No LLM cost.
  PDF   — Layout-aware system prompt injected before standard LLM extraction.

Detection: scans first 3000 chars for CoStar watermarks/header strings.
"""

import re
import logging
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# CoStar Excel column header → tbl_sales_comparables column name
#
# Keys are lowercased, whitespace-collapsed header strings.
# Values starting with '_recorded_' are held and applied only when the
# corresponding "True" value is absent (CoStar "True" > "Recorded").
# None values are intentionally unmapped (not written to DB).
# ─────────────────────────────────────────────────────────────────────────────

COSTAR_COLUMN_MAP: Dict[str, Optional[str]] = {
    # ── Identity ──────────────────────────────────────────────────────────────
    'property name':            'property_name',
    'property address':         'address',
    'address':                  'address',
    'city':                     'city',
    'state':                    'state',
    'zip':                      'zip',
    'postal code':              'zip',
    'zip code':                 'zip',
    'county':                   'county',
    'submarket':                'submarket',
    'market':                   'metro_market',
    'metro market':             'metro_market',
    'cbsa':                     'cbsa',
    'latitude':                 'latitude',
    'longitude':                'longitude',
    'distance from subject':    'distance_from_subject',
    'distance':                 'distance_from_subject',
    'costar property id':       'costar_comp_id',
    'property id':              'costar_comp_id',
    'costar id':                'costar_comp_id',

    # ── Property characteristics ───────────────────────────────────────────────
    'property type':            'property_type',
    'building type':            'property_type',
    'secondary type':           'property_subtype',
    'property subtype':         'property_subtype',
    'class':                    'building_class',
    'building class':           'building_class',
    'star rating':              'costar_star_rating',
    'costar star rating':       'costar_star_rating',
    'year built':               'year_built',
    'year renovated':           None,   # not in core table; skip
    'number of units':          'units',
    '# units':                  'units',
    'number of unit(s)':        'units',
    'units':                    'units',
    'building nra (sf)':        'building_sf',
    'nra (sf)':                 'building_sf',
    'building sf':              'building_sf',
    'rentable building area':   'building_sf',
    'gross building area (sf)': 'building_sf',
    'gba (sf)':                 'building_sf',
    'number of stories':        'num_floors',
    '# stories':                'num_floors',
    'stories':                  'num_floors',
    'number of buildings':      'num_buildings',
    '# buildings':              'num_buildings',
    'lot size (acres)':         'land_area_acres',
    'land area (acres)':        'land_area_acres',
    'lot size (sf)':            'land_area_sf',
    'land area (sf)':           'land_area_sf',
    'parking spaces':           'parking_spaces',
    'parking ratio':            'parking_ratio',
    'parking type':             'parking_type',
    'zoning':                   'zoning',
    'construction type':        'construction_type',
    'tenancy':                  'tenancy_type',
    'average unit size (sf)':   'avg_unit_size_sf',
    'avg unit size (sf)':       'avg_unit_size_sf',

    # ── Transaction — Recorded (lower priority; held until True is checked) ──
    'recorded sale date':       '_recorded_sale_date',
    'recorded sale price':      '_recorded_sale_price',
    'close of escrow':          '_recorded_sale_date',

    # ── Transaction — True (highest priority; overrides Recorded) ─────────────
    'true sale date':           'sale_date',
    'true sale price':          'sale_price',
    'sale date':                'sale_date',
    'sale price':               'sale_price',
    'transaction date':         'sale_date',

    # ── Pricing metrics ───────────────────────────────────────────────────────
    'price/unit':               'price_per_unit',
    'price per unit':           'price_per_unit',
    'price/sf (nra)':           'price_per_sf',
    'price/sf':                 'price_per_sf',
    'price per sf':             'price_per_sf',
    'price/acre':               'price_per_acre',
    'price per acre':           'price_per_acre',

    # ── Financial metrics ─────────────────────────────────────────────────────
    'cap rate':                 'cap_rate',
    'going-in cap rate':        'cap_rate',
    'overall cap rate':         'actual_cap_rate',
    'true equivalent rate':     'actual_cap_rate',
    'pro forma cap rate':       'pro_forma_cap_rate',
    'grm':                      'grm',
    'gross rent multiplier':    'grm',
    'gim':                      'gim',
    'noi':                      'noi_at_sale',
    'noi at sale':              'noi_at_sale',
    '% leased':                 'percent_leased_at_sale',
    'percent leased':           'percent_leased_at_sale',
    'percent leased at sale':   'percent_leased_at_sale',

    # ── Transaction details ───────────────────────────────────────────────────
    'sale conditions':          'sale_conditions',
    'conditions of sale':       'sale_conditions',
    'property rights':          'property_rights',
    'sale type':                'sale_type',
    'asking price':             'asking_price',
    'document number':          'document_number',
    'days on market':           'days_on_market',
    'is portfolio sale':        'is_portfolio_sale',
    'portfolio':                'portfolio_name',

    # ── Financing ─────────────────────────────────────────────────────────────
    'financing':                'financing_type',
    'financing type':           'financing_type',
    'lender':                   'financing_lender',
    'loan amount':              'financing_amount',
    'loan rate':                'financing_rate',

    # ── Buyer / Seller — Recorded ─────────────────────────────────────────────
    'buyer':                    'recorded_buyer',
    'seller':                   'recorded_seller',
    "buyer's broker":           'buyer_broker_company',
    "seller's broker":          'listing_broker_company',
    'selling broker':           'listing_broker_company',
    'buyer broker':             'buyer_broker_company',
    'listing broker':           'listing_broker_company',

    # ── Buyer / Seller — True ─────────────────────────────────────────────────
    'true buyer':               'true_buyer',
    'true seller':              'true_seller',
    'buyer type':               'buyer_type',
    'seller type':              'seller_type',

    # ── Verification ─────────────────────────────────────────────────────────
    'verified by':              'verification_source',
    'confirmed by':             'verification_source',
    'verification source':      'verification_source',
    'verification date':        'verification_date',

    # ── Notes ─────────────────────────────────────────────────────────────────
    'notes':                    'notes',
    'comments':                 'notes',
    'transaction notes':        'transaction_notes',
}

# Recorded → True fallback pairs.
# If a True value was not found but a Recorded one was, promote it.
_RECORDED_FALLBACKS: Dict[str, str] = {
    '_recorded_sale_date':  'sale_date',
    '_recorded_sale_price': 'sale_price',
}

# Minimum set of CoStar-specific column headers that identify the header row.
_COSTAR_HEADER_SIGNALS = {
    'property name',
    'property address',
    'recorded sale price',
    'true sale price',
    'costar property id',
    'building nra (sf)',
    'true equivalent rate',
    'price/unit',
    'price/sf (nra)',
}

# Text markers found in CoStar document headers / footers / watermarks.
_COSTAR_DOC_MARKERS = [
    'costar',
    'copyright costar',
    'costar realty information',
    'costar comp',
    'costar group',
    'costar.com',
]


# ─────────────────────────────────────────────────────────────────────────────
# Detection
# ─────────────────────────────────────────────────────────────────────────────

def is_costar_document(content: str) -> bool:
    """Return True if document content appears to originate from CoStar."""
    if not content:
        return False
    sample = content[:3000].lower()
    return any(marker in sample for marker in _COSTAR_DOC_MARKERS)


# ─────────────────────────────────────────────────────────────────────────────
# Excel parsing helpers
# ─────────────────────────────────────────────────────────────────────────────

def _normalize_header(val: Any) -> str:
    """Lowercase and collapse whitespace for column header matching."""
    return re.sub(r'\s+', ' ', str(val).strip().lower())


def _find_header_row(rows: list) -> Optional[int]:
    """
    Find the first row (within the first 10) that contains CoStar column headers.
    Requires at least 2 known CoStar header signals to avoid false positives.
    """
    for i, row in enumerate(rows[:10]):
        normalized = {_normalize_header(cell) for cell in row if cell is not None}
        matches = normalized & _COSTAR_HEADER_SIGNALS
        if len(matches) >= 2:
            return i
    return None


def _clean_value(val: Any) -> Optional[Any]:
    """Return None for null-like values; strip whitespace from strings."""
    if val is None:
        return None
    s = str(val).strip()
    if s.lower() in ('', 'none', 'null', 'n/a', 'na', '-', '--'):
        return None
    return s if isinstance(val, str) else val


def _parse_numeric(val: Any) -> Optional[float]:
    """Parse a value to float, stripping common formatting."""
    if val is None:
        return None
    try:
        return float(str(val).replace(',', '').replace('$', '').replace('%', '').strip())
    except (ValueError, TypeError):
        return None


# ─────────────────────────────────────────────────────────────────────────────
# Excel extraction
# ─────────────────────────────────────────────────────────────────────────────

def extract_costar_excel_comps(file_path: str) -> List[Dict[str, Any]]:
    """
    Parse a CoStar Excel export and return a list of comp dicts.

    Each dict uses tbl_sales_comparables column names as keys.
    'data_source' is set to 'CoStar' on every record.

    Returns an empty list if the file cannot be opened or contains no
    recognisable CoStar header rows.
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as exc:
        logger.error(f"CoStar Excel: failed to open '{file_path}': {exc}")
        return []

    all_comps: List[Dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            continue

        header_row_idx = _find_header_row(rows)
        if header_row_idx is None:
            logger.debug(f"CoStar Excel: no header row in sheet '{sheet_name}' — skipping")
            continue

        headers = [
            _normalize_header(h) if h is not None else ''
            for h in rows[header_row_idx]
        ]

        for data_row in rows[header_row_idx + 1:]:
            if not any(cell is not None for cell in data_row):
                continue  # entirely blank row

            comp: Dict[str, Any] = {'data_source': 'CoStar'}
            recorded_hold: Dict[str, Any] = {}

            for col_idx, header in enumerate(headers):
                if not header or col_idx >= len(data_row):
                    continue

                raw_val = data_row[col_idx]
                val = _clean_value(raw_val)
                if val is None:
                    continue

                # Look up the DB column name
                if header not in COSTAR_COLUMN_MAP:
                    continue  # Unmapped header — skip

                db_col = COSTAR_COLUMN_MAP[header]

                if db_col is None:
                    continue  # Explicitly unmapped

                if db_col.startswith('_recorded_'):
                    # Hold; apply only if True value absent
                    recorded_hold[db_col] = val
                else:
                    # Later duplicate headers for the same db_col are ignored
                    # (e.g. "True Sale Price" wins over any earlier "Sale Price")
                    if db_col not in comp:
                        comp[db_col] = val

            # Apply Recorded fallbacks where True value was not found
            for recorded_key, true_key in _RECORDED_FALLBACKS.items():
                if true_key not in comp and recorded_key in recorded_hold:
                    comp[true_key] = recorded_hold[recorded_key]

            # Normalise cap_rate: CoStar exports as percentage (5.25);
            # tbl_sales_comparables stores as decimal (0.0525).
            for cap_col in ('cap_rate', 'actual_cap_rate', 'pro_forma_cap_rate'):
                if cap_col in comp:
                    num = _parse_numeric(comp[cap_col])
                    if num is not None and num > 1:
                        comp[cap_col] = str(round(num / 100, 4))

            # Only keep rows that have at least a name, address, or price
            if comp.get('property_name') or comp.get('address') or comp.get('sale_price'):
                all_comps.append(comp)

    wb.close()
    logger.info(f"CoStar Excel: extracted {len(all_comps)} comp(s) from '{file_path}'")
    return all_comps


# ─────────────────────────────────────────────────────────────────────────────
# PDF: layout-aware system prompt
# ─────────────────────────────────────────────────────────────────────────────

COSTAR_PDF_SYSTEM_PROMPT = """\
You are extracting data from a CoStar comparable sales report PDF.

REPORT STRUCTURE
CoStar PDFs use a property card layout. Each card contains:
  • Header — property name, address, property type, photo strip
  • Transaction section — sale date, sale price, price/unit, price/SF, cap rate
  • Property section — year built, units, building SF, stories, lot size, class, submarket
  • Contacts section — buyer, seller, brokers (may be labeled "True Buyer", "True Seller")
  • Verification line — "Verified by [source] on [date]" or "Confirmed by …"

COSTAR TERMINOLOGY NOTES
  • "True Sale Price" and "True Sale Date" are the arms-length values.
    Always prefer them over "Recorded Sale Price" / "Recorded Sale Date".
  • "Price/SF (NRA)" = price per net rentable square foot.
  • "Building NRA (SF)" = building gross square footage.
  • "True Equivalent Rate" = the actual overall cap rate.
  • "Confirmed By" or "Verified By" = verification_source.
  • "CoStar Property ID" = costar_comp_id.

EXTRACTION RULES
  1. Extract ALL property cards found in the document.
  2. For each card, extract every available field — do not omit fields because
     they appear in a grid, table, or multi-column layout.
  3. Return null for fields you cannot find — never guess or infer values.
  4. Cap rates: return as a decimal (e.g. 0.0525 for 5.25 %).
  5. Prices: return as plain numbers without $ signs or commas.
"""


def get_costar_pdf_system_prompt() -> str:
    """Return the enhanced system prompt for CoStar PDF extraction."""
    return COSTAR_PDF_SYSTEM_PROMPT
